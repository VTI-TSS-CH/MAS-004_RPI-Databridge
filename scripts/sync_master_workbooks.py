from __future__ import annotations

import argparse
import shutil
from copy import copy
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook


ROOT_GIT_REPO = Path(__file__).resolve().parents[2]
RPI_REPO = Path(__file__).resolve().parents[1]
DEFAULT_PARAM_SOURCE = ROOT_GIT_REPO / "Parameterliste SAR41-MAS-004.xlsx"
DEFAULT_PARAM_REPO_COPY = RPI_REPO / "master_data" / "Parameterliste SAR41-MAS-004.xlsx"
DEFAULT_IO_SOURCE = Path(
    r"d:\Users\Egli_Erwin\Veralto\DE-SMD-Support-Switzerland - Documents\11_SAR\01_CH"
    r"\SAR41-MAS-004_Roche_TTO_LSR_Label und Kontrollsystem_V2\03_Dokumentation\SPS I_Os"
    r"\SAR41-MAS-004_SPS_I-Os.xlsx"
)
DEFAULT_IO_REPO_COPY = RPI_REPO / "master_data" / "SAR41-MAS-004_SPS_I-Os.xlsx"


SPECIAL_AI: dict[str, str] = {
    "MAP0001": "KI: Ich verstehe diesen Parameter als Etiketten- bzw. Traegerbandbreite in 1/10 mm. Der Wert definiert zugleich die Sollposition der beiden Etikettenanschlaege: der rechte Nullpunkt liegt von oben gesehen an der rechten Bandkante, daher entspricht 20.0 mm dem Sollwert 200 und 55.0 mm dem Sollwert 550 fuer Einlauf- und Auslaufanschlag.",
    "MAP0002": "KI: Ich verstehe diesen Parameter als Soll-Etikettenlaenge ohne Gap in 1/10 mm. Der ESP muss die real gemessene Etikettenlaenge mit MAP0040 vergleichen; bei zu kurzer Laenge wird MAE0025 gesetzt, bei zu langer Laenge MAE0026. Der Raspi behandelt diese beiden Fehler als Produktionspause statt als Purge/Not-Stop.",
    "MAP0003": "KI: Ich verstehe diesen Parameter als Druckposition in X-/Querrichtung auf dem Etikett. Fuer die Portalachse X gilt eine Umkehrlogik: ein positiver Druckpositionswert verschiebt den Tisch in negative X-Richtung, ein negativer Wert in positive X-Richtung. Die zusaetzliche Korrektur MAP0005 wird vorher auf den Druckpositionswert addiert.",
    "MAP0004": "KI: Ich verstehe diesen Parameter als Druckposition in Y-/Laufrichtung auf dem Etikett. Der konkrete Stopppunkt ergibt sich aus MAP0018 fuer Laser oder MAP0019 fuer TTO plus MAP0004 und plus Korrektur MAP0006.",
    "MAP0005": "KI: Ich verstehe diesen Parameter als zusaetzliche Korrektur auf die X-Druckposition. Er wird auf MAP0003 addiert, bevor daraus die invertierte X-Achskorrektur fuer den Tisch gebildet wird.",
    "MAP0006": "KI: Ich verstehe diesen Parameter als zusaetzliche Korrektur auf die Y-Druckposition. Er wird auf MAP0004 addiert und beeinflusst damit den exakten Druck-Stoppweg in Laufrichtung.",
    "MAP0007": "KI: Ich verstehe diesen Parameter als Materialdicke bzw. Produktions-Z-Position des Tisches in 1/10 mm. Zusammen mit der Nullpunktkorrektur MAP0028 ergibt er den Z-Sollwert fuer den Produktionsbetrieb.",
    "MAP0008": "KI: Ich verstehe diesen Parameter als Querposition des Etikettenerfassungs-Sensors im Produktionsbetrieb in mm. Fuer die Motor-/Achslogik wird er in 1/10 mm umgerechnet und mit der Nullpunktkorrektur MAP0029 kombiniert.",
    "MAP0009": "KI: Ich verstehe diesen Parameter als Querposition des Entnahme-/Kontrollsensors im Produktionsbetrieb in mm. Fuer die Motor-/Achslogik wird er in 1/10 mm umgerechnet und mit der Nullpunktkorrektur MAP0030 kombiniert.",
    "MAP0010": "KI: Ich verstehe diesen Parameter als X-/Quer-Leseposition der Material-Kontrollkamera. Er beschreibt die formatabhaengige Kameraposition; die mechanische Achskorrektur fuer diese Kamera wird separat ueber MAP0033 bzw. die Motor-Sollposition MAP0060 beruecksichtigt.",
    "MAP0011": "KI: Ich verstehe diesen Parameter als Y-/Laengs-Leseposition der Material-Kontrollkamera. Der Wert ist ein Prozessbezug fuer Trigger-/Auswertefenster und nicht die Motor-ID-5-Querachse selbst.",
    "MAP0012": "KI: Ich verstehe diesen Parameter als Y-/Laengs-Leseposition der OCR-/Verifizierungs-Kamera. Die Kamera ist aktuell nicht als eigene Verstellachse modelliert; der Wert bleibt aber fuer Prozessfenster und spaetere Verifikation relevant.",
    "MAP0013": "KI: Ich verstehe diesen Parameter als Rollenkern-Typ fuer Auf- und Abwicklung (0=76 mm, 1=100 mm). Die vorhandene Hardware kann den Kern nicht sicher direkt erkennen; eine Plausibilisierung ueber Wicklerbewegung und Rollenfuellstand ist theoretisch denkbar, aber noch nicht freigegeben. Deshalb wird der Wert vorerst nur als lesbarer Format-/Materialhinweis an den ESP gespiegelt und nicht aktiv fuer Regelentscheidungen verwendet.",
    "MAP0014": "KI: Ich verstehe diesen Parameter als Soll-Transportgeschwindigkeit des Etikettenbands. Im kontinuierlichen Lasermodus ist er die Druck-/Bandgeschwindigkeit; in allen getakteten Modi ist er die Vorzugsgeschwindigkeit zwischen den Druckpositionen. Der ESP nutzt ihn fuer Taktbewegungen und dynamische Nachregelung, der Raspi fuer Transparenz, Logging und Bedienung.",
    "MAP0015": "KI: Ich verstehe diesen Parameter als Rueckspulgeschwindigkeit fuer den Rueckspulbetrieb und fuer Ruecksetzbewegungen nach Fehlern bzw. Entnahmekontrolle. Dieser Wert ist prozesskritisch fuer den ESP, weil die Rueckspulung damit weich und reproduzierbar laufen soll.",
    "MAP0016": "KI: Ich verstehe diesen Parameter als Auswahl zwischen TTO und Laser. Es darf immer nur eines der beiden Drucksysteme fuehrend aktiv sein; davon haengen Triggerweg, Bypasslogik und die nachfolgenden Bewertungsbits im Label-Schieberegister ab.",
    "MAP0017": "KI: Ich verstehe diesen Parameter als Distanz vom Einlese-Sensor bis zur Material-Kontrollkamera. Diese Distanz ist ein fester Wegbezug im Label-Schieberegister, damit die Kamera fuer jedes bereits eingelesene Label an der richtigen Stelle getriggert wird.",
    "MAP0018": "KI: Ich verstehe diesen Parameter als Distanz vom Einlese-Sensor bis zur Druckposition des Lasers. Nur wenn MAP0016 auf Laser steht und kein Bypass aktiv ist, darf der ESP an genau dieser Wegposition den Laser-Trigger ausloesen.",
    "MAP0019": "KI: Ich verstehe diesen Parameter als Distanz vom Einlese-Sensor bis zur Druckposition des TTO. Nur wenn MAP0016 auf TTO steht und kein Bypass aktiv ist, darf der ESP an genau dieser Wegposition den TTO-Trigger ausloesen.",
    "MAP0020": "KI: Ich verstehe diesen Parameter als Distanz vom Einlese-Sensor bis zur OCR-/Verifizierungs-Kamera. Dieser Wegwert wird im Label-Schieberegister fuer den spaeteren OCR-Trigger und die Rueckmeldung von Verifizierung OK gebraucht.",
    "MAP0021": "KI: Ich verstehe diesen Parameter als Distanz vom Einlese-Sensor bis zum Kontrollsensor am Tischende. Dort wird geprueft, ob ein gutes Label vorhanden bleiben durfte bzw. ein schlechtes Label entnommen wurde.",
    "MAP0022": "KI: Ich verstehe diesen Parameter als Rueckspuldistanz, wenn ein Label am Kontrollsensor haette entnommen sein muessen, aber noch erkannt wurde. Der ESP muss dann das Band kontrolliert zuruecksetzen, ohne bereits bearbeitete Labels im Schieberegister fachlich noch einmal neu zu behandeln.",
    "MAP0023": "KI: Ich verstehe diesen Parameter als Vorwarnschwelle fuer den Abwickler in Prozent Restfuellstand. Wenn MAS0008 kleiner/gleich diesem Wert ist, ist die Abwickelrolle bald leer und die passende Wickler-/Materialwarnung muss aktiv werden. Der ESP liest diesen Wert fuer die Wickler-/Prozesslogik, Microtom liest ihn als eingestellte Schwelle.",
    "MAP0024": "KI: Ich verstehe diesen Parameter als Vorwarnschwelle fuer den Aufwickler in Prozent Rollenfuellstand. Wenn MAS0009 groesser/gleich diesem Wert ist, ist die Aufwickelrolle bald voll und die passende Wickler-/Materialwarnung muss aktiv werden. Default ist 95 %, also Warnung im Bereich 95-100 %. Alte/gespiegelte Werte kleiner/gleich 50 werden in der Wickler-Firmware uebergangsweise als Restreserve bis voll interpretiert, z.B. 5 bedeutet Warnung ab 95 %. Der ESP liest diesen Wert fuer die Wickler-/Prozesslogik, Microtom liest ihn als eingestellte Schwelle.",
    "MAP0025": "KI: Ich verstehe diesen Parameter als harte Produktions-Pausenschwelle fuer den Abwickler. Wird der verbleibende Fuellstand unterschritten, muss die Anlage in den Pausenzustand wechseln, damit kein unkontrollierter Produktionsabbruch entsteht.",
    "MAP0035": "KI: Ich verstehe diesen Parameter als Bypass fuer das aktive Drucksystem (TTO oder Laser). Wenn er gesetzt ist, wird der echte Drucker nicht getriggert und dessen Status-/Druckfertig-Signale werden ignoriert. Stattdessen wartet die ESP-Prozesslogik die passende Simulationsdauer MAP0069 (Laser) oder MAP0070 (TTO) ab und setzt das Druckergebnis intern als erledigt.",
    "MAP0036": "KI: Ich verstehe diesen Parameter als Bypass der Material-Kontrollkamera. Bei aktivem Bypass wird der Kameratrigger weiterhin wie im Produktionsbetrieb ausgegeben, aber Bereit-/Gut-Schlecht-Rueckmelde-IOs werden ignoriert. Das simulierte Gut/Schlecht-Ergebnis wird ueber MAP0067 gebildet.",
    "MAP0037": "KI: Ich verstehe diesen Parameter als Bypass der Druck-Verifikations-/OCR-Kamera. Bei aktivem Bypass wird der Kameratrigger weiterhin wie im Produktionsbetrieb ausgegeben, aber Bereit-/Gut-Schlecht-Rueckmelde-IOs werden ignoriert. Das simulierte Gut/Schlecht-Ergebnis wird ueber MAP0068 gebildet.",
    "MAP0038": "KI: Ich verstehe diesen Parameter als Bypass fuer den Etiketten-Kontrollsensor am Auslauf. Bei aktivem Bypass wird keine Entnahme-/Rueckspulpflicht fuer schlechte Labels erzwungen; die Statuswerte der Labels bleiben aber im Schieberegister erhalten.",
    "MAP0027": "KI: Ich verstehe diesen Parameter als Nullpunktkorrektur der Portalachse X. Die programmierte Motor-Nullposition bleibt die Grundstellung; dieser Wert verschiebt den fachlichen Nullpunkt fuer alle X-bezogenen Berechnungen.",
    "MAP0028": "KI: Ich verstehe diesen Parameter als Nullpunktkorrektur der Portalachse Z. Die programmierte Motor-Nullposition bleibt die Grundstellung; dieser Wert verschiebt den fachlichen Nullpunkt fuer alle Z-bezogenen Berechnungen.",
    "MAP0029": "KI: Ich verstehe diesen Parameter als Nullpunktkorrektur der Querachse fuer den Etikettenerfassungs-Sensor. Alle Sensorpositionen fuer diese Achse werden relativ zu diesem korrigierten Nullpunkt betrachtet.",
    "MAP0030": "KI: Ich verstehe diesen Parameter als Nullpunktkorrektur der Querachse fuer den Entnahme-/Kontrollsensor. Alle Sensorpositionen fuer diese Achse werden relativ zu diesem korrigierten Nullpunkt betrachtet.",
    "MAP0031": "KI: Ich verstehe diesen Parameter als Nullpunktkorrektur des Etikettenanschlags am Einlauf. Die aus MAP0001 abgeleitete Bandbreitenposition wird um diese Korrektur verschoben.",
    "MAP0032": "KI: Ich verstehe diesen Parameter als Nullpunktkorrektur des Etikettenanschlags am Auslauf. Die aus MAP0001 abgeleitete Bandbreitenposition wird um diese Korrektur verschoben.",
    "MAP0033": "KI: Ich verstehe diesen Parameter als Nullpunktkorrektur der Material-Kontrollkamera-Querachse. Die formatabhaengige Kameraposition wird relativ zu diesem korrigierten Nullpunkt betrachtet.",
    "MAP0034": "KI: Ich verstehe diesen Parameter als Nullpunktkorrektur des Laser-Schutzblechs. Die Schutzblechachse bleibt parametrierbar, auch wenn der Laserprozess spaeter finalisiert wird.",
    "MAP0039": "KI: Ich verstehe diesen Parameter als Freigabe fuer automatische Rueckspulung nach Produktionsstop. Wenn der Wert 1 ist, wird nach Produktionsstop und Rueckspul-Taster bis zum Ursprungspunkt zurueckgespult und erst danach die Produktion abgeschlossen; bei 0 gilt die Produktion ohne Rueckspulung als abgeschlossen.",
    "MAP0047": "KI: Ich verstehe diesen Parameter als Vorgabe, ob innen- oder aussengewickeltes Material verarbeitet wird. Daraus ergeben sich die Drehrichtungen fuer Auf- und Abwickler; der ESP soll diesen Wert lesen und fuer die Wicklerlogik umsetzen.",
    "MAP0056": "KI: Ich verstehe diesen Parameter als Sollposition der Portalachse X fuer Format bzw. Maschinenbewegung. Die eigentliche Anfahrt und Rueckmeldung erfolgt ueber den Oriental-Antrieb mit Motor-ID 1.",
    "MAP0057": "KI: Ich verstehe diesen Parameter als Sollposition der Portalachse Z fuer Format bzw. Maschinenbewegung. Die eigentliche Anfahrt und Rueckmeldung erfolgt ueber den Oriental-Antrieb mit Motor-ID 2.",
    "MAP0058": "KI: Ich verstehe diesen Parameter als Sollbewegung bzw. Sollposition des Etikettenantriebs in 1/10 mm. Dieser Parameter ist im Produktionsprozess besonders sensibel, weil daraus die genaue Bandposition fuer Druck und Kameratrigger entsteht.",
    "MAP0059": "KI: Ich verstehe diesen Parameter als Sollposition des Laser-Schutzblechs ueber Motor-ID 4. Solange der Laser fachlich noch nicht im Fokus steht, bleibt der Zusammenhang teilweise offen; die Achse muss aber trotzdem sauber parametrierbar bleiben.",
    "MAP0060": "KI: Ich verstehe diesen Parameter als Sollposition der Material-Kontrollkamera quer zum Band ueber Motor-ID 5. Dieser Wert ist formatabhaengig und bestimmt, wo die Kamera das Material liest.",
    "MAP0061": "KI: Ich verstehe diesen Parameter als Sollposition des Einlese-Sensors quer zum Band ueber Motor-ID 6. Der Sensor muss so positioniert werden, dass das Label sauber eingemessen wird.",
    "MAP0062": "KI: Ich verstehe diesen Parameter als Sollposition des Kontrollsensors am Auslauf quer zum Band ueber Motor-ID 7. Dieser Wert bestimmt, wo die Label-Entnahme bzw. Restanwesenheit geprueft wird.",
    "MAP0063": "KI: Ich verstehe diesen Parameter als Sollposition des linken bzw. Auslauf-Etikettenanschlags ueber Motor-ID 8. Dieser Wert folgt normalerweise der eingestellten Materialbreite.",
    "MAP0064": "KI: Ich verstehe diesen Parameter als Sollposition des rechten bzw. Einlauf-Etikettenanschlags ueber Motor-ID 9. Dieser Wert folgt normalerweise der eingestellten Materialbreite und sollte konsistent zu MAP0063 sein.",
    "MAP0065": "KI: Ich verstehe diesen Parameter als Bitmaske fuer die Freigabe der Maschinen-Tasten je nach Microtom-Benutzerlevel. Der Raspi nutzt diese Bitmaske, um Tastereingaben zu erlauben oder zu sperren und um die zugehoerigen Taster-LEDs nur dann zu signalisieren, wenn die Bedienhandlung fachlich und rechtebezogen erlaubt ist.",
    "MAP0066": "KI: Ich verstehe diesen Parameter als Distanz vom Einlese-Sensor bis zur ersten LED des externen LED-Streifens. Der ESP braucht diesen Wegbezug, um den Labelstatus entlang des Streifens korrekt anzuzeigen; der exakte Wert muss bei der realen Inbetriebnahme mechanisch verifiziert werden.",
    "MAP0067": "KI: Ich verstehe diesen Parameter als Simulationsmuster fuer die Material-Kontrollkamera bei aktivem MAP0036. 0=alle gut, 1=alle schlecht, n=jede n-te Etikette schlecht.",
    "MAP0068": "KI: Ich verstehe diesen Parameter als Simulationsmuster fuer die Druck-Verifikations-/OCR-Kamera bei aktivem MAP0037. 0=alle gut, 1=alle schlecht, n=jede n-te Etikette schlecht.",
    "MAP0069": "KI: Ich verstehe diesen Parameter als simulierte Laser-Druckdauer in ms bei aktivem Drucksystem-Bypass MAP0035.",
    "MAP0070": "KI: Ich verstehe diesen Parameter als simulierte TTO-Druckdauer in ms bei aktivem Drucksystem-Bypass MAP0035.",
    "MAP0071": "KI: Ich verstehe diesen Parameter als aktive Laenge des externen LED-Streifens in 1/10 mm. Aktuell ist der Streifen auf 520.0 mm gekuerzt; der ESP rendert dafuer 75 Pixel und sendet sie an den externen LED-Controller.",
    "MAP0072": "KI: Ich verstehe diesen Parameter als Freigabe des externen LED-Controller-Datenstroms. Bei 1 sendet der ESP32-PLC Prozess fertige RGB-Frames per UDP an den Olimex ESP32-POE-ISO-IND Controller, bei 0 bleibt der LED-Datenstrom aus.",
    "MAP0073": "KI: Ich verstehe diesen Parameter als letztes Oktett der Ziel-IP fuer den externen LED-Controller im ESP-Netz 192.168.2.x. 110 ist fuer den vorbereiteten Olimex ESP32-POE-ISO-IND Controller vorgesehen; der Default 255 nutzt Broadcast.",
    "MAP0074": "KI: Ich verstehe diesen Parameter als UDP-Zielport fuer das MAS004-LED-UDP/v1 Protokoll zum externen LED-Controller.",
    "MAP0075": "KI: Ich verstehe diesen Parameter als minimales Sendeintervall fuer LED-Frames in ms. Der Default 33 ms entspricht etwa 30 Hz und ist fuer die Positionsanzeige ausreichend schnell.",
    "MAP0076": "KI: Ich verstehe diesen Parameter als konstante Laengenkorrektur fuer den Etikettenerfassungssensor in 1/10 mm. Der ESP addiert diesen Wert auf die vom Sensorfenster gemessene HIGH-Laenge fuer Setup-/Anzeige-Diagnose; die Produktion entscheidet Laengenfehler auf roher Einlaufencoder-Laenge, damit die fixe Sensorfenster-Kompensation keine falschen Produktionsalarme erzeugt. Encoder-Rohwerte bleiben fuer Diagnose und Schlupfvergleich sichtbar.",
    "MAP0077": "KI: Ich verstehe diesen Parameter als festen Maschinen-Abgleichwert fuer den Wirkdurchmesser des Einlaufencoders in 1/1000 mm. Er ist nicht formatrelevant und wird nach der 2000-mm-Kalibrierfahrt vom Raspi an die ESP32-PLC synchronisiert.",
    "MAP0078": "KI: Ich verstehe diesen Parameter als festen Maschinen-Abgleichwert fuer den Wirkdurchmesser des Auslauf-/ID3-Encoders in 1/1000 mm. Er ist nicht formatrelevant und wird nach der 2000-mm-Kalibrierfahrt vom Raspi an die ESP32-PLC synchronisiert.",
    "MAS0001": "KI: Ich verstehe diesen Parameter als fuehrenden Anlagenstatus, den der Raspi gegenueber Microtom meldet. Die eigentliche Statuslogik entsteht aus Bedienbefehlen, Sicherheitsbedingungen und Maschinenfortschritt; der ESP liefert dafuer harte Prozessereignisse zu.",
    "MAS0002": "KI: Ich verstehe diesen Parameter nicht als Status, sondern als Befehlsbyte von Microtom an die Anlage. Der Raspi muss diesen Befehlswert in den passenden internen Zielzustand uebersetzen, zum Beispiel Start, Stop, Einrichten, Synchronisieren, Leerfahren, Rueckspulen oder Pause.",
    "MAS0003": "KI: Ich verstehe diesen Parameter als verdichtete Produktionsrueckmeldung pro fertig behandeltem Label. Der ESP baut die Einzelinformationen im Schieberegister auf; der Raspi packt sie in das definierte Bitfeld und meldet den Wert an Microtom sowie in das Label-Produktionslog.",
    "MAS0008": "KI: Ich verstehe diesen Parameter als Fuellstand des Abwicklers in Prozent. Microtom liest ihn nur; der Wert wird aus der Wicklerlogik abgeleitet.",
    "MAS0009": "KI: Ich verstehe diesen Parameter als Fuellstand des Aufwicklers in Prozent. Microtom liest ihn nur; der Wert wird aus der Wicklerlogik abgeleitet.",
    "MAS0026": "KI: Ich verstehe diesen Parameter als Status des Abwicklers. Der Wert wird von der Wicklersteuerung fuehrend geliefert und vom ESP/Raspi fuer die Gesamtanlage weiterverwendet.",
    "MAS0027": "KI: Ich verstehe diesen Parameter als Status des Aufwicklers. Der Wert wird von der Wicklersteuerung fuehrend geliefert und vom ESP/Raspi fuer die Gesamtanlage weiterverwendet.",
    "MAS0028": "KI: Ich verstehe diesen Parameter als Purge- bzw. Abbruchereignis, das eine saubere Fortsetzung der laufenden Produktion verhindert. Solche Ereignisse muessen in den Not-Stop-/Stoerungszustand fuehren und sind fachlich staerker als eine normale Pause.",
    "MAS0029": "KI: Ich verstehe diesen Parameter als aktuelle Produktions- bzw. Auftrags-ID fuer Logfiles und Rueckverfolgbarkeit. Gemass Masterliste ist er fuer den ESP nicht relevant; fuehrend verwaltet wird er auf Raspi-/Microtom-Seite.",
    "MAS0030": "KI: Ich verstehe diesen Parameter als Status, ob Produktionslogfiles der letzten Produktion noch abholbereit sind. Auch dieser Wert ist gemaess Masterliste Raspi-/Microtom-seitig und soll nicht in den ESP gespiegelt werden.",
    "TTS0001": "KI: Ich verstehe diesen Parameter als numerischen Gesamtstatus des TTO-Druckers gemaess ZBC-Statusabbild. Der Wert muss sowohl aktiv aus Druckerstatuswechseln zurueckgemeldet als auch - soweit vom Drucker unterstuetzt - vom ESP gesetzt werden koennen, zum Beispiel fuer Offline, Online oder Shutdown.",
}

LED_CONTROLLER_PARAM_SPECS = [
    {
        "key": "MAP0071",
        "ptype": "MAP",
        "pid": "0071",
        "min": "10",
        "max": "5200",
        "default": "5200",
        "unit": "1/10mm",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "name": "MAP0071 LED-Streifen aktive Laenge",
        "format": "YES",
        "message": "Aktive Laenge des externen LED-Streifens in 1/10 mm",
    },
    {
        "key": "MAP0072",
        "ptype": "MAP",
        "pid": "0072",
        "min": "0",
        "max": "1",
        "default": "1",
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "bool",
        "name": "MAP0072 LED-Controller UDP aktiv",
        "format": "YES",
        "message": "Freigabe des ESP32-PLC UDP-Datenstroms zum externen LED-Controller",
    },
    {
        "key": "MAP0073",
        "ptype": "MAP",
        "pid": "0073",
        "min": "1",
        "max": "255",
        "default": "255",
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint8",
        "name": "MAP0073 LED-Controller Ziel-IP letztes Oktett",
        "format": "NO",
        "message": "Letztes Oktett der LED-Controller-Ziel-IP im ESP-Netz 192.168.2.x",
    },
    {
        "key": "MAP0074",
        "ptype": "MAP",
        "pid": "0074",
        "min": "1",
        "max": "65535",
        "default": "3050",
        "unit": "port",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "name": "MAP0074 LED-Controller UDP-Port",
        "format": "NO",
        "message": "UDP-Zielport fuer MAS004-LED-UDP/v1 Frames",
    },
    {
        "key": "MAP0075",
        "ptype": "MAP",
        "pid": "0075",
        "min": "20",
        "max": "1000",
        "default": "33",
        "unit": "ms",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "name": "MAP0075 LED-Controller Frame-Intervall",
        "format": "NO",
        "message": "Minimales Sendeintervall fuer LED-Frames an den externen Controller",
    },
    {
        "key": "MAP0076",
        "ptype": "MAP",
        "pid": "0076",
        "min": "-50",
        "max": "50",
        "default": "8",
        "unit": "1/10mm",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "int16",
        "name": "MAP0076 Label-Laengenkompensation",
        "format": "NO",
        "message": "Konstante Kompensation der vom Etikettenerfassungssensor gemessenen HIGH-Laenge in 1/10 mm",
    },
    {
        "key": "MAP0077",
        "ptype": "MAP",
        "pid": "0077",
        "min": "1000",
        "max": "1000000",
        "default": "100765",
        "unit": "1/1000mm",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint32",
        "name": "MAP0077 Einlaufencoder Wirkdurchmesser",
        "format": "NO",
        "message": "Fester Maschinen-Abgleichwert fuer die Einlaufencoder-Skalierung in 1/1000 mm",
    },
    {
        "key": "MAP0078",
        "ptype": "MAP",
        "pid": "0078",
        "min": "1000",
        "max": "1000000",
        "default": "100649",
        "unit": "1/1000mm",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint32",
        "name": "MAP0078 Auslaufencoder Wirkdurchmesser",
        "format": "NO",
        "message": "Fester Maschinen-Abgleichwert fuer die Auslauf-/ID3-Encoder-Skalierung in 1/1000 mm",
    },
]


def row_key(ws, row_idx: int) -> str:
    ptype = str(ws.cell(row_idx, 1).value or "").strip().upper()
    pid = str(ws.cell(row_idx, 2).value or "").strip()
    return f"{ptype}{pid}" if ptype and pid else ""


def norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    for ch in ".:/\\_-?":
        text = text.replace(ch, " ")
    return " ".join(text.split())


def header_map(ws) -> dict[str, int]:
    return {
        norm_header(ws.cell(1, col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
        if norm_header(ws.cell(1, col_idx).value)
    }


def col_any(headers: dict[str, int], *names: str) -> int:
    for name in names:
        key = norm_header(name)
        if key in headers:
            return headers[key]
    raise RuntimeError(f"Missing expected workbook column. Tried: {', '.join(names)}")


def col_optional(headers: dict[str, int], *names: str) -> Optional[int]:
    for name in names:
        key = norm_header(name)
        if key in headers:
            return headers[key]
    return None


def normalize_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\r", "\n").replace("\n", " ").split()).strip()


def split_ai_cell(value: Any) -> tuple[str, str]:
    text = normalize_text(value)
    if not text:
        return "", ""
    marker = "KI:"
    idx = text.find(marker)
    if idx < 0:
        return text, ""
    return text[:idx].strip(), text[idx + len(marker):].strip()


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
        if src.number_format:
            dst.number_format = src.number_format


def access_text(rw: str, esp_rw: str) -> str:
    microtom_text = {
        "R": "Microtom soll diesen Wert nur lesen.",
        "W": "Microtom soll diesen Wert fuehrend schreiben koennen.",
        "R/W": "Microtom soll diesen Wert lesen und bei Bedarf auch schreiben koennen.",
        "N": "Dieser Wert ist fuer Microtom aktuell nicht freigegeben.",
    }.get(rw, f"Microtom-Rechte laut Tabelle: {rw or 'unbekannt'}.")
    esp_text = {
        "R": "Der ESP soll ihn aus der Raspi-/Microtom-Seite lesen bzw. gespiegelt bekommen, aber nicht fuehrend schreiben.",
        "W": "Der ESP darf ihn im Prozess aktiv verwenden und - falls fachlich vorgesehen - auch setzen oder rueckmelden.",
        "R/W": "Der ESP darf ihn lesen und schreiben.",
        "N": "Der ESP soll ihn gemaess Masterliste nicht verwenden.",
    }.get(esp_rw, f"ESP-Rechte laut Tabelle: {esp_rw or 'unbekannt'}.")
    return f"{microtom_text} {esp_text}"


def generic_ai_for_row(values: dict[str, str]) -> str:
    pkey = values["pkey"]
    ptype = values["ptype"]
    name = values["name"]
    message = values["message"]
    rw = values["rw"]
    esp_rw = values["esp_rw"]
    unit = values["unit"]
    fmt = values["format_relevant"]
    mapping = values["mapping"]
    existing = values["existing_ai"]
    operator_note = values.get("operator_note", "")

    if pkey in SPECIAL_AI:
        base = SPECIAL_AI[pkey]
    elif ptype == "MAP":
        desc = name or message or "einen Maschinen- bzw. Formatparameter"
        base = f"KI: Ich verstehe {pkey} als {desc}."
    elif ptype == "MAS":
        desc = name or message or "einen Maschinenstatus bzw. Maschinenwert"
        base = f"KI: Ich verstehe {pkey} als {desc}."
    elif ptype == "MAE":
        desc = name or message or "eine Stoerungsrueckmeldung"
        base = f"KI: Ich verstehe {pkey} als Stoerungsbit bzw. Stoerungsmeldung fuer {desc}."
    elif ptype == "MAW":
        desc = name or message or "eine Warnungsrueckmeldung"
        base = f"KI: Ich verstehe {pkey} als Warnungsbit bzw. Vorwarnung fuer {desc}."
    elif ptype == "TTP":
        desc = name or message or "einen TTO-Parameter"
        base = f"KI: Ich verstehe {pkey} als TTO-Betriebsparameter ({desc})."
    elif ptype in {"TTE", "TTW"}:
        desc = name or message or "eine TTO-Ereignis- bzw. Statusmeldung"
        base = f"KI: Ich verstehe {pkey} als vom TTO kommendes Ereignis bzw. Statussignal ({desc})."
    elif ptype == "TTS":
        desc = name or message or "einen TTO-Gesamtstatus"
        base = f"KI: Ich verstehe {pkey} als verdichteten TTO-Gesamtstatus ({desc})."
    elif ptype in {"LSE", "LSW"}:
        desc = name or message or "eine Laser-Ereignis- bzw. Statusmeldung"
        base = f"KI: Ich verstehe {pkey} als vom Laser kommendes Ereignis bzw. Statussignal ({desc})."
    else:
        desc = name or message or "einen derzeit noch nicht genauer beschriebenen Parameter"
        base = f"KI: Ich verstehe {pkey} als {desc}."

    extras: list[str] = []
    base_plain = base.removeprefix("KI:").strip()
    if operator_note:
        extras.append(f"Aus der ergaenzten Fachbeschreibung wurde ausgewertet: {operator_note}.")
    extras.append(access_text(rw, esp_rw))
    if unit:
        extras.append(f"Einheit bzw. Wertebereich laut Tabelle: {unit}.")
    if fmt:
        extras.append(
            "Der Wert ist formatrelevant und muss beim Formatwechsel konsistent mitgezogen werden."
            if fmt.upper() == "YES"
            else "Der Wert ist laut Tabelle nicht formatrelevant und eher als Maschinen-/Systemeinstellung zu sehen."
        )
    if mapping:
        extras.append(f"Technische Zuordnung laut Tabelle: {mapping}.")
    if existing and existing != base_plain and not existing.startswith("Ich verstehe "):
        extras.append(f"Vorhandene Projekt-Notiz: {existing}.")
    if not name and not message:
        extras.append("Die fachliche Bedeutung ist in der Tabelle aktuell noch nicht ausgeschrieben; das muss im Feldtest bzw. ueber Herstellerspezifikation weiter verifiziert werden.")

    return " ".join(part.strip() for part in [base, *extras] if part and part.strip())


def ensure_map0066(ws):
    headers = header_map(ws)
    c_type = col_any(headers, "Params_Type.:", "Params_Type")
    c_id = col_any(headers, "Param. ID.:", "Param ID")
    c_min = col_any(headers, "Min.:", "Min")
    c_max = col_any(headers, "Max.:", "Max")
    c_default = col_any(headers, "Default Value:", "Default Value")
    c_unit = col_any(headers, "Einheit:", "Einheit", "Unit")
    c_user_range = col_optional(headers, "Microtom User Range:", "Microtom User Range")
    c_rw = col_any(headers, "R/W:", "R/W")
    c_esp_rw = col_any(headers, "ESP32 R/W:", "ESP32 R/W")
    c_dtype = col_any(headers, "Data Type:", "Data Type")
    c_name = col_any(headers, "Name:", "Name")
    c_zbc_mapping = col_optional(headers, "ZBC Mapping:", "ZBC Mapping")
    c_format = col_any(headers, "Format relevant?:", "Format relevant")
    c_message = col_any(headers, "Message:", "Message")
    c_cause = col_any(headers, "Possible Cause:", "Possible Cause")
    c_effects = col_any(headers, "Effects:", "Effects")
    c_remedy = col_any(headers, "Remedy:", "Remedy")
    c_ai = col_any(headers, "KI-Anweisungen:", "KI-Anweisungen", "AI Instructions")

    existing_row = None
    map0065_row = None
    for row_idx in range(2, ws.max_row + 1):
        key = row_key(ws, row_idx)
        if key == "MAP0065":
            map0065_row = row_idx
        if key == "MAP0066":
            existing_row = row_idx
            break

    if existing_row is None:
        if map0065_row is None:
            raise RuntimeError("MAP0065 not found - cannot insert MAP0066 cleanly")
        target_row = map0065_row + 1
        if row_key(ws, target_row):
            ws.insert_rows(target_row, 1)
        copy_row_style(ws, map0065_row, target_row, ws.max_column)
        existing_row = target_row

    values_by_col = {
        c_type: "MAP",
        c_id: "0066",
        c_min: "0",
        c_max: "30000",
        c_default: "8000",
        c_unit: "1/10mm",
        c_rw: "W",
        c_esp_rw: "R",
        c_dtype: "uint16",
        c_name: "MAP0066 Distanz Etikettenerfassung - erste LED LED-Streifen",
        c_format: "NO",
        c_message: "Distanz von der Etikettenerfassung bis zur ersten LED des LED-Streifens",
        c_cause: None,
        c_effects: None,
        c_remedy: None,
        c_ai: SPECIAL_AI["MAP0066"],
    }
    if c_user_range is not None:
        values_by_col[c_user_range] = None
    if c_zbc_mapping is not None:
        values_by_col[c_zbc_mapping] = None

    for col_idx, value in values_by_col.items():
        ws.cell(existing_row, col_idx).value = value


def ensure_led_controller_params(ws):
    headers = header_map(ws)
    c_type = col_any(headers, "Params_Type.:", "Params_Type")
    c_id = col_any(headers, "Param. ID.:", "Param ID")
    c_min = col_any(headers, "Min.:", "Min")
    c_max = col_any(headers, "Max.:", "Max")
    c_default = col_any(headers, "Default Value:", "Default Value")
    c_unit = col_any(headers, "Einheit:", "Einheit", "Unit")
    c_user_range = col_optional(headers, "Microtom User Range:", "Microtom User Range")
    c_rw = col_any(headers, "R/W:", "R/W")
    c_esp_rw = col_any(headers, "ESP32 R/W:", "ESP32 R/W")
    c_dtype = col_any(headers, "Data Type:", "Data Type")
    c_name = col_any(headers, "Name:", "Name")
    c_zbc_mapping = col_optional(headers, "ZBC Mapping:", "ZBC Mapping")
    c_format = col_any(headers, "Format relevant?:", "Format relevant")
    c_message = col_any(headers, "Message:", "Message")
    c_cause = col_any(headers, "Possible Cause:", "Possible Cause")
    c_effects = col_any(headers, "Effects:", "Effects")
    c_remedy = col_any(headers, "Remedy:", "Remedy")
    c_ai = col_any(headers, "KI-Anweisungen:", "KI-Anweisungen", "AI Instructions")

    def find_row(key: str) -> Optional[int]:
        for row_idx in range(2, ws.max_row + 1):
            if row_key(ws, row_idx) == key:
                return row_idx
        return None

    previous_key = "MAP0070"
    for spec in LED_CONTROLLER_PARAM_SPECS:
        existing_row = find_row(spec["key"])
        if existing_row is None:
            anchor = find_row(previous_key) or find_row("MAP0066")
            if anchor is None:
                raise RuntimeError("MAP0066/MAP0070 not found - cannot insert LED controller params cleanly")
            target_row = anchor + 1
            if row_key(ws, target_row):
                ws.insert_rows(target_row, 1)
            copy_row_style(ws, anchor, target_row, ws.max_column)
            existing_row = target_row

        values_by_col = {
            c_type: spec["ptype"],
            c_id: spec["pid"],
            c_min: spec["min"],
            c_max: spec["max"],
            c_default: spec["default"],
            c_unit: spec["unit"],
            c_rw: spec["rw"],
            c_esp_rw: spec["esp_rw"],
            c_dtype: spec["dtype"],
            c_name: spec["name"],
            c_format: spec["format"],
            c_message: spec["message"],
            c_cause: None,
            c_effects: None,
            c_remedy: None,
            c_ai: SPECIAL_AI[spec["key"]],
        }
        if c_user_range is not None:
            values_by_col[c_user_range] = None
        if c_zbc_mapping is not None:
            values_by_col[c_zbc_mapping] = None

        for col_idx, value in values_by_col.items():
            ws.cell(existing_row, col_idx).value = value
        previous_key = spec["key"]


def sync_parameter_workbook(path: Path):
    wb = load_workbook(path)
    ws = wb["Parameter"]
    ensure_map0066(ws)
    ensure_led_controller_params(ws)
    headers = header_map(ws)
    c_type = col_any(headers, "Params_Type.:", "Params_Type")
    c_id = col_any(headers, "Param. ID.:", "Param ID")
    c_unit = col_any(headers, "Einheit:", "Einheit", "Unit")
    c_rw = col_any(headers, "R/W:", "R/W")
    c_esp_rw = col_any(headers, "ESP32 R/W:", "ESP32 R/W")
    c_name = col_any(headers, "Name:", "Name")
    c_mapping = col_optional(headers, "ZBC Mapping:", "ZBC Mapping", "ESP Key")
    c_format = col_any(headers, "Format relevant?:", "Format relevant")
    c_message = col_any(headers, "Message:", "Message")
    c_ai = col_any(headers, "KI-Anweisungen:", "KI-Anweisungen", "AI Instructions")

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        ptype = str(ws.cell(row_idx, c_type).value or "").strip().upper()
        pid = str(ws.cell(row_idx, c_id).value or "").strip()
        if not ptype or not pid:
            continue
        operator_note, existing_ai = split_ai_cell(ws.cell(row_idx, c_ai).value)
        values = {
            "pkey": f"{ptype}{pid}",
            "ptype": ptype,
            "name": normalize_text(ws.cell(row_idx, c_name).value),
            "mapping": normalize_text(ws.cell(row_idx, c_mapping).value) if c_mapping else "",
            "format_relevant": normalize_text(ws.cell(row_idx, c_format).value),
            "message": normalize_text(ws.cell(row_idx, c_message).value),
            "rw": normalize_text(ws.cell(row_idx, c_rw).value).upper(),
            "esp_rw": normalize_text(ws.cell(row_idx, c_esp_rw).value).upper(),
            "unit": normalize_text(ws.cell(row_idx, c_unit).value),
            "existing_ai": existing_ai,
            "operator_note": operator_note,
        }
        ai_text = generic_ai_for_row(values)
        if ws.cell(row_idx, c_ai).value != ai_text:
            ws.cell(row_idx, c_ai).value = ai_text
            updated += 1

    wb.save(path)
    return updated


def main():
    ap = argparse.ArgumentParser(description="Sync MAS-004 master workbooks and enrich KI texts.")
    ap.add_argument("--param-source", default=str(DEFAULT_PARAM_SOURCE))
    ap.add_argument("--param-repo-copy", default=str(DEFAULT_PARAM_REPO_COPY))
    ap.add_argument("--io-source", default=str(DEFAULT_IO_SOURCE))
    ap.add_argument("--io-repo-copy", default=str(DEFAULT_IO_REPO_COPY))
    args = ap.parse_args()

    param_source = Path(args.param_source)
    param_repo_copy = Path(args.param_repo_copy)
    io_source = Path(args.io_source)
    io_repo_copy = Path(args.io_repo_copy)

    if not param_source.exists():
        raise RuntimeError(f"Parameter workbook not found: {param_source}")
    if not io_source.exists():
        raise RuntimeError(f"IO workbook not found: {io_source}")

    updated_source = sync_parameter_workbook(param_source)
    param_repo_copy.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(param_source, param_repo_copy)

    io_repo_copy.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(io_source, io_repo_copy)

    print(f"Parameter workbook synced: {param_source}")
    print(f"Repo workbook copy updated: {param_repo_copy}")
    print(f"KI entries refreshed: {updated_source}")
    print(f"IO workbook copied: {io_repo_copy}")


if __name__ == "__main__":
    main()
