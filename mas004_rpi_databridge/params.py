from __future__ import annotations

import io
import re
from typing import Optional, Dict, Any, Tuple

import openpyxl

from mas004_rpi_databridge.db import DB, now_ts

SHEET_NAME = "Parameter"
DEVICE_PUSH_PTYPES = {"TTE", "TTW", "LSE", "LSW", "MAE", "MAW"}


def _to_str(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _to_clean_str(v) -> Optional[str]:
    s = _to_str(v)
    if s is None:
        return None
    s = s.strip()
    return s if s else None


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None


def _to_int(v) -> Optional[int]:
    if v is None:
        return None
    s = _to_clean_str(v)
    if not s:
        return None
    try:
        if s.lower().startswith("0x"):
            return int(s, 16)
        return int(float(s))
    except Exception:
        return None


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def _normalize_microtom_rw(value: str | None) -> Optional[str]:
    raw = _to_clean_str(value)
    if raw is None:
        return None
    norm = raw.upper().replace("_", "/")
    if norm == "RW":
        norm = "R/W"
    if norm in ("R", "W", "R/W", "N"):
        return norm
    return raw.upper()


def _normalize_esp_rw(value: str | None) -> str:
    raw = _to_clean_str(value)
    if raw is None:
        return "W"
    norm = raw.upper().replace("_", "/")
    if norm == "RW":
        norm = "R/W"
    if norm in ("R", "W", "R/W", "N"):
        return norm
    return "W"


class ParamStore:
    def __init__(self, db: DB):
        self.db = db

    def import_xlsx(self, file_path: str) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f"Excel-Sheet '{SHEET_NAME}' nicht gefunden. Vorhanden: {wb.sheetnames}")
        ws = wb[SHEET_NAME]

        inserted = 0
        updated = 0
        skipped = 0

        headers_raw = [(_to_str(ws.cell(1, c).value) or "").strip() for c in range(1, ws.max_column + 1)]
        header_map: Dict[str, int] = {}
        for i, h in enumerate(headers_raw, start=1):
            nh = _norm_header(h)
            if nh:
                header_map[nh] = i

        def col_any(*names: str) -> Optional[int]:
            for n in names:
                nn = _norm_header(n)
                if nn in header_map:
                    return header_map[nn]
            for want in names:
                want2 = (want or "").strip().lower()
                for idx, h in enumerate(headers_raw, start=1):
                    if (h or "").strip().lower() == want2:
                        return idx
            return None

        c_type = col_any("Params_Type.", "Params_Type.:", "Params Type", "Params_Type")
        c_id = col_any("Param ID.", "Param. ID.:", "Param. ID.", "Param ID", "Param_ID", "Param. ID")

        c_min = col_any("Min.", "Min.:", "Min")
        c_max = col_any("Max.", "Max.:", "Max")
        c_def = col_any("Default Value", "Default Value:", "Default")
        c_unit = col_any("Einheit", "Unit")
        c_rw = col_any("R/W:", "R/W", "RW", "R_W")
        c_esp_rw = col_any("ESP32 R/W:", "ESP32 R/W", "ESP R/W:", "ESP R/W", "ESP_RW")
        c_dtype = col_any("Data Type", "DataType", "Datatype")
        c_name = col_any("Name")
        c_fmt = col_any("Format relevant?", "Format relevant", "Format")
        c_msg = col_any("Message")
        c_cause = col_any("Possible Cause", "Possible cause")
        c_eff = col_any("Effects", "Effect")
        c_rem = col_any("Remedy")
        c_ai = col_any("KI-Anweisungen:", "KI-Anweisungen", "KI Anweisungen", "AI Instructions")

        # Optional mapping columns for live device protocols.
        c_esp_key = col_any("ESP Key", "ESP_Key", "ESP Param", "ESP Parameter")
        c_zbc_mapping = col_any("ZBC Mapping", "ZBC Mapping:")
        c_zbc_msg_id = col_any("VJ6530 Msg ID", "VJ6530 Message ID", "ZBC Message ID", "ZBC Msg ID")
        c_zbc_cmd_id = col_any("VJ6530 Cmd ID", "VJ6530 Command ID", "ZBC Command ID", "ZBC Cmd ID")
        c_zbc_codec = col_any("VJ6530 Codec", "ZBC Codec", "ZBC Value Codec")
        c_zbc_scale = col_any("VJ6530 Scale", "ZBC Scale")
        c_zbc_offset = col_any("VJ6530 Offset", "ZBC Offset")
        c_ult_set = col_any("VJ3350 Set Cmd", "Ultimate Set Cmd", "Ultimate Set Command")
        c_ult_get = col_any("VJ3350 Get Cmd", "Ultimate Get Cmd", "Ultimate Get Command")
        c_ult_var = col_any("VJ3350 Var", "Ultimate Var", "Ultimate Variable")

        has_map_cols = any(
            c is not None
            for c in (
                c_esp_key,
                c_zbc_mapping,
                c_zbc_msg_id,
                c_zbc_cmd_id,
                c_zbc_codec,
                c_zbc_scale,
                c_zbc_offset,
                c_ult_set,
                c_ult_get,
                c_ult_var,
            )
        )

        if not c_type or not c_id:
            raise RuntimeError(
                "Pflichtspalten fehlen. Erwartet (varianten): 'Params_Type' und 'Param ID'. "
                f"Gefunden (normalisiert): {sorted(list(header_map.keys()))}"
            )

        with self.db._conn() as c:
            for r in range(2, ws.max_row + 1):
                ptype = (_to_str(ws.cell(r, c_type).value) or "").strip().upper()
                pid = (_to_str(ws.cell(r, c_id).value) or "").strip()

                if not ptype and not pid:
                    continue
                if not ptype or not pid:
                    skipped += 1
                    continue

                pkey = f"{ptype}{pid}"

                min_v = _to_float(ws.cell(r, c_min).value) if c_min else None
                max_v = _to_float(ws.cell(r, c_max).value) if c_max else None
                default_v = _to_str(ws.cell(r, c_def).value) if c_def else None
                unit = _to_str(ws.cell(r, c_unit).value) if c_unit else None
                rw = _normalize_microtom_rw(_to_str(ws.cell(r, c_rw).value) if c_rw else None)
                esp_rw = _normalize_esp_rw(_to_str(ws.cell(r, c_esp_rw).value) if c_esp_rw else "W")
                dtype = _to_str(ws.cell(r, c_dtype).value) if c_dtype else None
                name = _to_str(ws.cell(r, c_name).value) if c_name else None
                fmt = _to_str(ws.cell(r, c_fmt).value) if c_fmt else None
                msg = _to_str(ws.cell(r, c_msg).value) if c_msg else None
                cause = _to_str(ws.cell(r, c_cause).value) if c_cause else None
                eff = _to_str(ws.cell(r, c_eff).value) if c_eff else None
                rem = _to_str(ws.cell(r, c_rem).value) if c_rem else None
                ai = _to_str(ws.cell(r, c_ai).value) if c_ai else None

                ts = now_ts()

                exists = c.execute("SELECT 1 FROM params WHERE pkey=?", (pkey,)).fetchone() is not None
                if exists:
                    c.execute(
                        """UPDATE params SET
                           ptype=?, pid=?, min_v=?, max_v=?, default_v=?, unit=?, rw=?, esp_rw=?, dtype=?,
                           name=?, format_relevant=?, message=?, possible_cause=?, effects=?, remedy=?,
                           ai_instructions=?,
                           updated_ts=?
                           WHERE pkey=?""",
                        (
                            ptype,
                            pid,
                            min_v,
                            max_v,
                            default_v,
                            unit,
                            rw,
                            esp_rw,
                            dtype,
                            name,
                            fmt,
                            msg,
                            cause,
                            eff,
                            rem,
                            ai,
                            ts,
                            pkey,
                        ),
                    )
                    updated += 1
                else:
                    c.execute(
                        """INSERT INTO params(
                           pkey,ptype,pid,min_v,max_v,default_v,unit,rw,esp_rw,dtype,name,format_relevant,
                           message,possible_cause,effects,remedy,ai_instructions,updated_ts
                           )
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            pkey,
                            ptype,
                            pid,
                            min_v,
                            max_v,
                            default_v,
                            unit,
                            rw,
                            esp_rw,
                            dtype,
                            name,
                            fmt,
                            msg,
                            cause,
                            eff,
                            rem,
                            ai,
                            ts,
                        ),
                    )
                    inserted += 1

                if has_map_cols:
                    map_row = {
                        "esp_key": _to_clean_str(ws.cell(r, c_esp_key).value) if c_esp_key else None,
                        "zbc_mapping": _to_clean_str(ws.cell(r, c_zbc_mapping).value) if c_zbc_mapping else None,
                        "zbc_message_id": _to_int(ws.cell(r, c_zbc_msg_id).value) if c_zbc_msg_id else None,
                        "zbc_command_id": _to_int(ws.cell(r, c_zbc_cmd_id).value) if c_zbc_cmd_id else None,
                        "zbc_value_codec": _to_clean_str(ws.cell(r, c_zbc_codec).value) if c_zbc_codec else None,
                        "zbc_scale": _to_float(ws.cell(r, c_zbc_scale).value) if c_zbc_scale else None,
                        "zbc_offset": _to_float(ws.cell(r, c_zbc_offset).value) if c_zbc_offset else None,
                        "ultimate_set_cmd": _to_clean_str(ws.cell(r, c_ult_set).value) if c_ult_set else None,
                        "ultimate_get_cmd": _to_clean_str(ws.cell(r, c_ult_get).value) if c_ult_get else None,
                        "ultimate_var_name": _to_clean_str(ws.cell(r, c_ult_var).value) if c_ult_var else None,
                    }
                    has_any_map = any(v is not None for v in map_row.values())

                    if has_any_map:
                        c.execute(
                            """INSERT INTO param_device_map(
                               pkey, esp_key, zbc_mapping, zbc_message_id, zbc_command_id, zbc_value_codec,
                               zbc_scale, zbc_offset, ultimate_set_cmd, ultimate_get_cmd,
                               ultimate_var_name, updated_ts
                               )
                               VALUES(?,?,?,?,?,?,?,?,?,?,?,?)
                               ON CONFLICT(pkey) DO UPDATE SET
                                 esp_key=excluded.esp_key,
                                 zbc_mapping=excluded.zbc_mapping,
                                 zbc_message_id=excluded.zbc_message_id,
                                 zbc_command_id=excluded.zbc_command_id,
                                 zbc_value_codec=excluded.zbc_value_codec,
                                 zbc_scale=excluded.zbc_scale,
                                 zbc_offset=excluded.zbc_offset,
                                 ultimate_set_cmd=excluded.ultimate_set_cmd,
                                 ultimate_get_cmd=excluded.ultimate_get_cmd,
                                 ultimate_var_name=excluded.ultimate_var_name,
                                 updated_ts=excluded.updated_ts""",
                            (
                                pkey,
                                map_row["esp_key"],
                                map_row["zbc_mapping"],
                                map_row["zbc_message_id"],
                                map_row["zbc_command_id"],
                                map_row["zbc_value_codec"],
                                map_row["zbc_scale"],
                                map_row["zbc_offset"],
                                map_row["ultimate_set_cmd"],
                                map_row["ultimate_get_cmd"],
                                map_row["ultimate_var_name"],
                                ts,
                            ),
                        )
                    else:
                        c.execute("DELETE FROM param_device_map WHERE pkey=?", (pkey,))

        return {"ok": True, "inserted": inserted, "updated": updated, "skipped": skipped}

    def get_meta(self, pkey: str) -> Optional[Dict[str, Any]]:
        with self.db._conn() as c:
            row = c.execute(
                """SELECT pkey,ptype,pid,min_v,max_v,default_v,unit,rw,esp_rw,dtype,name,format_relevant,message,ai_instructions
                   FROM params WHERE pkey=?""",
                (pkey,),
            ).fetchone()
        if not row:
            return None
        keys = [
            "pkey",
            "ptype",
            "pid",
            "min_v",
            "max_v",
            "default_v",
            "unit",
            "rw",
            "esp_rw",
            "dtype",
            "name",
            "format_relevant",
            "message",
            "ai_instructions",
        ]
        return dict(zip(keys, row))

    def get_device_map(self, pkey: str) -> Dict[str, Any]:
        with self.db._conn() as c:
            row = c.execute(
                """SELECT esp_key,zbc_mapping,zbc_message_id,zbc_command_id,zbc_value_codec,zbc_scale,zbc_offset,
                          ultimate_set_cmd,ultimate_get_cmd,ultimate_var_name
                   FROM param_device_map
                   WHERE pkey=?""",
                (pkey,),
            ).fetchone()

        if not row:
            return {}

        keys = [
            "esp_key",
            "zbc_mapping",
            "zbc_message_id",
            "zbc_command_id",
            "zbc_value_codec",
            "zbc_scale",
            "zbc_offset",
            "ultimate_set_cmd",
            "ultimate_get_cmd",
            "ultimate_var_name",
        ]
        return dict(zip(keys, row))

    def get_value(self, pkey: str) -> Optional[str]:
        with self.db._conn() as c:
            row = c.execute("SELECT value FROM param_values WHERE pkey=?", (pkey,)).fetchone()
        return row[0] if row else None

    def get_effective_value(self, pkey: str) -> str:
        v = self.get_value(pkey)
        if v is not None:
            return v
        meta = self.get_meta(pkey)
        dv = (meta or {}).get("default_v")
        return dv if dv is not None else "0"

    def actor_access(self, pkey: str, actor: str = "microtom") -> str:
        meta = self.get_meta(pkey)
        if not meta:
            return "N"
        actor_name = (actor or "microtom").strip().lower()
        if actor_name in ("esp", "esp32", "esp-plc"):
            return _normalize_esp_rw(meta.get("esp_rw"))
        return _normalize_microtom_rw(meta.get("rw")) or "W"

    def can_actor_read(self, pkey: str, actor: str = "microtom") -> bool:
        access = self.actor_access(pkey, actor=actor)
        return access != "N"

    def can_actor_write(self, pkey: str, actor: str = "microtom") -> bool:
        access = self.actor_access(pkey, actor=actor)
        return access in ("W", "R/W")

    def validate_read(self, pkey: str, actor: str = "microtom") -> Tuple[bool, str]:
        meta = self.get_meta(pkey)
        if not meta:
            return False, "NAK_UnknownParam"
        if not self.can_actor_read(pkey, actor=actor):
            return False, "NAK_NoAccess"
        return True, "OK"

    def validate_write(self, pkey: str, value: str, actor: str = "microtom") -> Tuple[bool, str]:
        meta = self.get_meta(pkey)
        if not meta:
            return False, "NAK_UnknownParam"

        if not self.can_actor_write(pkey, actor=actor):
            access = self.actor_access(pkey, actor=actor)
            return False, "NAK_NoAccess" if access == "N" else "NAK_ReadOnly"

        min_v = meta.get("min_v")
        max_v = meta.get("max_v")
        try:
            fv = float(value)
            if min_v is not None and fv < float(min_v):
                return False, "NAK_OutOfRange"
            if max_v is not None and fv > float(max_v):
                return False, "NAK_OutOfRange"
        except Exception:
            # non-numeric values are allowed when no numeric bounds are enforced
            pass
        return True, "OK"

    def set_value(self, pkey: str, value: str, actor: str = "microtom") -> Tuple[bool, str]:
        ok, msg = self.validate_write(pkey, value, actor=actor)
        if not ok:
            return ok, msg

        ts = now_ts()
        with self.db._conn() as c:
            c.execute(
                "INSERT INTO param_values(pkey,value,updated_ts) VALUES(?,?,?) "
                "ON CONFLICT(pkey) DO UPDATE SET value=excluded.value, updated_ts=excluded.updated_ts",
                (pkey, str(value), ts),
            )
            # Fuer den Testbetrieb soll ein erfolgreiches Write auch den Default-Wert fortschreiben.
            c.execute(
                "UPDATE params SET default_v=?, updated_ts=? WHERE pkey=?",
                (str(value), ts, pkey),
            )
        return True, "OK"

    def apply_device_value(self, pkey: str, value: str, promote_default: bool = False) -> Tuple[bool, str]:
        """
        Speichert einen von Device-Seite gemeldeten Wert (TTO/Laser/ESP) lokal.
        Im Unterschied zu set_value() wird rw NICHT blockiert, da ReadOnly hier nur
        fuer Microtom-Schreibrechte gilt, nicht fuer eingehende Status-/Fehlermeldungen.
        """
        meta = self.get_meta(pkey)
        if not meta:
            return False, "NAK_UnknownParam"

        rw = (meta.get("rw") or "").strip().upper()
        ptype = (meta.get("ptype") or "").strip().upper()
        can_update_default = bool(promote_default) or rw in ("W", "R/W") or ptype in DEVICE_PUSH_PTYPES

        ts = now_ts()
        with self.db._conn() as c:
            c.execute(
                "INSERT INTO param_values(pkey,value,updated_ts) VALUES(?,?,?) "
                "ON CONFLICT(pkey) DO UPDATE SET value=excluded.value, updated_ts=excluded.updated_ts",
                (pkey, str(value), ts),
            )
            if can_update_default:
                c.execute(
                    "UPDATE params SET default_v=?, updated_ts=? WHERE pkey=?",
                    (str(value), ts, pkey),
                )
        return True, "OK"

    def update_meta(
        self,
        pkey: str,
        default_v: Optional[str] = None,
        min_v: Optional[float] = None,
        max_v: Optional[float] = None,
        rw: Optional[str] = None,
        esp_rw: Optional[str] = None,
    ) -> Tuple[bool, str]:
        meta = self.get_meta(pkey)
        if not meta:
            return False, "NAK_UnknownParam"

        new_min = meta.get("min_v") if min_v is None else min_v
        new_max = meta.get("max_v") if max_v is None else max_v
        new_def = meta.get("default_v") if default_v is None else str(default_v)
        new_rw = meta.get("rw") if rw is None else str(rw).strip()
        new_esp_rw = meta.get("esp_rw") if esp_rw is None else str(esp_rw).strip()

        # rw normalisieren
        if new_rw is not None:
            rwu = _normalize_microtom_rw(new_rw)
            if rwu not in ("R", "W", "R/W", "N", None):
                return False, "NAK_BadRW"
            new_rw = rwu
        if new_esp_rw is not None:
            erwu = _normalize_esp_rw(new_esp_rw)
            if erwu not in ("R", "W", "R/W", "N"):
                return False, "NAK_BadEspRW"
            new_esp_rw = erwu

        # min/max check
        if new_min is not None and new_max is not None:
            if float(new_min) > float(new_max):
                return False, "NAK_MinGreaterThanMax"

        # default range check (nur wenn numeric)
        try:
            fv = float(new_def) if new_def is not None else None
            if fv is not None:
                if new_min is not None and fv < float(new_min):
                    return False, "NAK_DefaultOutOfRange"
                if new_max is not None and fv > float(new_max):
                    return False, "NAK_DefaultOutOfRange"
        except Exception:
            pass

        with self.db._conn() as c:
            c.execute(
                """UPDATE params
                   SET default_v=?, min_v=?, max_v=?, rw=?, esp_rw=?, updated_ts=?
                   WHERE pkey=?""",
                (new_def, new_min, new_max, new_rw, new_esp_rw, now_ts(), pkey),
            )
        return True, "OK"

    def list_params(self, ptype: Optional[str] = None, q: Optional[str] = None, limit: int = 200, offset: int = 0):
        limit = max(1, min(int(limit), 1000))
        offset = max(0, int(offset))
        where = []
        args = []

        if ptype:
            where.append("p.ptype=?")
            args.append(ptype)

        if q:
            q2 = f"%{q}%"
            where.append("(p.pkey LIKE ? OR p.name LIKE ? OR p.message LIKE ? OR p.ai_instructions LIKE ?)")
            args.extend([q2, q2, q2, q2])

        wsql = ("WHERE " + " AND ".join(where)) if where else ""
        sql = f"""SELECT p.pkey,p.ptype,p.pid,p.min_v,p.max_v,p.default_v,p.unit,p.rw,p.esp_rw,p.dtype,p.name,p.message,
                         p.ai_instructions,
                         m.esp_key,m.zbc_mapping,m.zbc_message_id,m.zbc_command_id,m.zbc_value_codec,m.zbc_scale,m.zbc_offset,
                         m.ultimate_set_cmd,m.ultimate_get_cmd,m.ultimate_var_name
                  FROM params p
                  LEFT JOIN param_device_map m ON m.pkey = p.pkey
                  {wsql}
                  ORDER BY p.ptype ASC, p.pid ASC
                  LIMIT ? OFFSET ?"""
        args.extend([limit, offset])

        with self.db._conn() as c:
            rows = c.execute(sql, args).fetchall()

        out = []
        for r in rows:
            pkey = r[0]
            cur = self.get_value(pkey)
            eff = cur if cur is not None else (r[5] if r[5] is not None else "0")
            out.append(
                {
                    "pkey": r[0],
                    "ptype": r[1],
                    "pid": r[2],
                    "min_v": r[3],
                    "max_v": r[4],
                    "default_v": r[5],
                    "current_v": cur,
                    "effective_v": eff,
                    "unit": r[6],
                    "rw": r[7],
                    "esp_rw": r[8],
                    "dtype": r[9],
                    "name": r[10],
                    "message": r[11],
                    "ai_instructions": r[12],
                    "esp_key": r[13],
                    "zbc_mapping": r[14],
                    "zbc_message_id": r[15],
                    "zbc_command_id": r[16],
                    "zbc_value_codec": r[17],
                    "zbc_scale": r[18],
                    "zbc_offset": r[19],
                    "ultimate_set_cmd": r[20],
                    "ultimate_get_cmd": r[21],
                    "ultimate_var_name": r[22],
                }
            )
        return out

    def export_xlsx_bytes(self, ptype: Optional[str] = None, q: Optional[str] = None) -> bytes:
        rows = self.list_params(ptype=ptype, q=q, limit=100000, offset=0)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        headers = [
            "Params_Type.:",
            "Param. ID.:",
            "Min.:",
            "Max.:",
            "Default Value",
            "Current Value",
            "Effective Value",
            "Einheit",
            "R/W:",
            "ESP32 R/W:",
            "Data Type",
            "Name",
            "Message",
            "KI-Anweisungen:",
            "ESP Key",
            "ZBC Mapping",
            "VJ6530 Msg ID",
            "VJ6530 Cmd ID",
            "VJ6530 Codec",
            "VJ6530 Scale",
            "VJ6530 Offset",
            "VJ3350 Set Cmd",
            "VJ3350 Get Cmd",
            "VJ3350 Var",
        ]
        ws.append(headers)

        for r in rows:
            ws.append(
                [
                    r.get("ptype"),
                    r.get("pid"),
                    r.get("min_v"),
                    r.get("max_v"),
                    r.get("default_v"),
                    r.get("current_v"),
                    r.get("effective_v"),
                    r.get("unit"),
                    r.get("rw"),
                    r.get("esp_rw"),
                    r.get("dtype"),
                    r.get("name"),
                    r.get("message"),
                    r.get("ai_instructions"),
                    r.get("esp_key"),
                    r.get("zbc_mapping"),
                    r.get("zbc_message_id"),
                    r.get("zbc_command_id"),
                    r.get("zbc_value_codec"),
                    r.get("zbc_scale"),
                    r.get("zbc_offset"),
                    r.get("ultimate_set_cmd"),
                    r.get("ultimate_get_cmd"),
                    r.get("ultimate_var_name"),
                ]
            )

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()
