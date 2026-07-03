from typing import Optional, Dict, Any

from fastapi import FastAPI, Request, HTTPException, Header, UploadFile, File, Query
from fastapi.responses import HTMLResponse, Response, FileResponse, RedirectResponse
from fastapi.openapi.docs import get_swagger_ui_html
from pydantic import BaseModel
from datetime import datetime
import hashlib
import hmac
import json
import html
import math
import subprocess
import os
import re
import secrets
import shutil
import tempfile
import threading
import time
import uuid

try:
    import multipart  # type: ignore  # noqa: F401

    MULTIPART_AVAILABLE = True
except Exception:
    MULTIPART_AVAILABLE = False

from mas004_rpi_databridge.config import Settings, DEFAULT_CFG_PATH
from mas004_rpi_databridge.db import DB
from mas004_rpi_databridge.outbox import Outbox
from mas004_rpi_databridge.inbox import Inbox
from mas004_rpi_databridge.params import ParamStore
from mas004_rpi_databridge.logstore import LogStore
from mas004_rpi_databridge.netconfig import IfaceCfg, apply_static, get_current_ip_info
from mas004_rpi_databridge.esp_motors import EspMotorClient
from mas004_rpi_databridge.io_master import IoStore
from mas004_rpi_databridge.io_runtime import IoRuntime
from mas004_rpi_databridge.machine_runtime import MachineRuntime, mark_external_purge_clear
from mas004_rpi_databridge.machine_control_ui import build_machine_control_ui_html
from mas004_rpi_databridge.device_clients import EspPlcClient
from mas004_rpi_databridge.format_semantics import build_format_plan
from mas004_rpi_databridge.production_setup_ui import build_production_setup_ui_html
from mas004_rpi_databridge.production_visualization_ui import build_production_visualization_ui_html
from mas004_rpi_databridge.mae0048_diagnostics import collect_mae0048_diagnostics
from mas004_rpi_databridge.mae0048_diagnostics_ui import build_mae0048_diagnostics_ui_html
from mas004_rpi_databridge.motor3_calibration_ui import build_motor3_calibration_ui_html
from mas004_rpi_databridge.commissioning import CommissioningStore
from mas004_rpi_databridge.commissioning_ui import build_commissioning_ui_html
from mas004_rpi_databridge.machine_backups import MachineBackupManager
from mas004_rpi_databridge.backup_ui import build_backup_ui_html
from mas004_rpi_databridge.format_profiles import FormatProfileStore, normalize_profile_name
from mas004_rpi_databridge.motor_catalog import merge_motor_payload, motor_catalog
from mas004_rpi_databridge.motor_bindings import build_motor_bindings
from mas004_rpi_databridge.motor_master_sync import (
    apply_motor_setup_master_config_to_client,
    reapply_motor_setup_master_to_params,
    sync_motor_master_values,
)
from mas004_rpi_databridge.motor_setup_lock import (
    clear_motor_setup_manual_lock,
    touch_motor_setup_manual_lock,
)
from mas004_rpi_databridge.motor_state_store import MotorStateStore
from mas004_rpi_databridge.protocol import normalize_pid
from mas004_rpi_databridge.peers import peer_urls
from mas004_rpi_databridge.production_logs import ProductionLogManager
from mas004_rpi_databridge.router import Router
from mas004_rpi_databridge.smart_wickler_client import SmartWicklerClient, normalize_winder_role
from mas004_rpi_databridge.smart_wickler_ui import build_winder_ui_html
from mas004_rpi_databridge.timeutil import format_local_timestamp, local_from_timestamp

ASSET_DIR = os.path.join(os.path.dirname(__file__), "assets")
VIDEOJET_LOGO_PATH = os.path.join(ASSET_DIR, "videojet-logo.jpg")
REPO_MASTER_PARAMS_XLSX = os.path.join(os.path.dirname(os.path.dirname(__file__)), "master_data", "Parameterliste SAR41-MAS-004.xlsx")
REPO_MASTER_IOS_XLSX = os.path.join(os.path.dirname(os.path.dirname(__file__)), "master_data", "SAR41-MAS-004_SPS_I-Os.xlsx")
WORKSPACE_MASTER_PARAMS_XLSX = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "Parameterliste SAR41-MAS-004.xlsx")
FALLBACK_VIDEOJET_LOGO_PATH = os.path.join("/opt", "MAS-004_RPI-Databridge", "mas004_rpi_databridge", "assets", "videojet-logo.jpg")
FALLBACK_REPO_MASTER_PARAMS_XLSX = os.path.join("/opt", "MAS-004_RPI-Databridge", "master_data", "Parameterliste SAR41-MAS-004.xlsx")
FALLBACK_REPO_MASTER_IOS_XLSX = os.path.join("/opt", "MAS-004_RPI-Databridge", "master_data", "SAR41-MAS-004_SPS_I-Os.xlsx")
MACHINE_SETUP_USER = "Admin"
MACHINE_SETUP_PASSWORD = "VideojetMAS004!"
MACHINE_SETUP_COOKIE = "mas004_machine_setup"
MACHINE_SETUP_SESSION_MAX_AGE_S = 12 * 60 * 60
MOTOR_REFRESH_CACHE_TTL_S = 1.5
MOTOR_REFRESH_BUSY_WAIT_S = 0.15
MACHINE_BYPASS_PARAM_DEFS = (
    {
        "pkey": "MAP0035",
        "ptype": "MAP",
        "pid": "0035",
        "default_v": "0",
        "min_v": 0,
        "max_v": 1,
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "bool",
        "name": "Bypass Drucksystem",
        "message": "Bypass fuer das aktive Drucksystem. Der echte Drucker blockiert den Prozess dann nicht.",
    },
    {
        "pkey": "MAP0036",
        "ptype": "MAP",
        "pid": "0036",
        "default_v": "0",
        "min_v": 0,
        "max_v": 1,
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "bool",
        "name": "Bypass Material-Kontrollkamera",
        "message": "Bei aktivem Bypass wird die Materialkamera getriggert, Rueckmeldungen werden simuliert.",
    },
    {
        "pkey": "MAP0037",
        "ptype": "MAP",
        "pid": "0037",
        "default_v": "0",
        "min_v": 0,
        "max_v": 1,
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "bool",
        "name": "Bypass Druck-Verifikationskamera",
        "message": "Bei aktivem Bypass wird die OCR/Verifikationskamera getriggert, Rueckmeldungen werden simuliert.",
    },
    {
        "pkey": "MAP0038",
        "ptype": "MAP",
        "pid": "0038",
        "default_v": "0",
        "min_v": 0,
        "max_v": 1,
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "bool",
        "name": "Bypass Etiketten-Entnahmesensor",
        "message": "Bei aktivem Bypass wird keine Entnahmekontrolle/Rueckspulung erzwungen.",
    },
    {
        "pkey": "MAP0067",
        "ptype": "MAP",
        "pid": "0067",
        "default_v": "0",
        "min_v": 0,
        "max_v": 9999,
        "unit": "n",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "name": "Materialkamera Bypass-Simulation",
        "message": "0=alle Labels gut, 1=alle schlecht, n=jede n-te Etikette schlecht.",
    },
    {
        "pkey": "MAP0068",
        "ptype": "MAP",
        "pid": "0068",
        "default_v": "0",
        "min_v": 0,
        "max_v": 9999,
        "unit": "n",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "name": "Verifikationskamera Bypass-Simulation",
        "message": "0=alle Labels gut, 1=alle schlecht, n=jede n-te Etikette schlecht.",
    },
    {
        "pkey": "MAP0069",
        "ptype": "MAP",
        "pid": "0069",
        "default_v": "2000",
        "min_v": 0,
        "max_v": 10000,
        "unit": "ms",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "name": "Laser Bypass-Druckdauer",
        "message": "Simulierte Druckdauer in ms, wenn Laser/Drucksystem gebypasst ist.",
    },
    {
        "pkey": "MAP0070",
        "ptype": "MAP",
        "pid": "0070",
        "default_v": "2000",
        "min_v": 0,
        "max_v": 10000,
        "unit": "ms",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "name": "TTO Bypass-Druckdauer",
        "message": "Simulierte Druckdauer in ms, wenn TTO/Drucksystem gebypasst ist.",
    },
    {
        "pkey": "MAP0079",
        "ptype": "MAP",
        "pid": "0079",
        "default_v": "0",
        "min_v": 0,
        "max_v": 1,
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "bool",
        "name": "Laser Testmodus Parallelbetrieb",
        "message": "Bei aktivem TTO wird Laser Start Q0.3 beim TTO Online/Offline-Schalten fuer 100 ms gepulst; Q0.1 triggert 100 ms parallel zu Q0.0.",
    },
)

MACHINE_LED_PARAM_DEFS = (
    {
        "pkey": "MAP0071",
        "ptype": "MAP",
        "pid": "0071",
        "default_v": "5200",
        "min_v": 10,
        "max_v": 5200,
        "unit": "0.1mm",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "format_relevant": "YES",
        "name": "LED-Streifen aktive Laenge",
        "message": "Aktiv genutzte LED-Streifenlaenge. Aktuell 520.0 mm, der ESP sendet dafuer 75 Pixel.",
    },
    {
        "pkey": "MAP0072",
        "ptype": "MAP",
        "pid": "0072",
        "default_v": "8",
        "min_v": 0,
        "max_v": 1,
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "bool",
        "format_relevant": "NO",
        "name": "LED-Controller UDP aktiv",
        "message": "1=ESP32-PLC sendet fertige LED-Frames per UDP an den Olimex ESP32-POE-ISO-IND LED-Controller.",
    },
    {
        "pkey": "MAP0073",
        "ptype": "MAP",
        "pid": "0073",
        "default_v": "255",
        "min_v": 1,
        "max_v": 255,
        "unit": "",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint8",
        "format_relevant": "NO",
        "name": "LED-Controller Ziel-IP letztes Oktett",
        "message": "Letztes Oktett im ESP-Netz 192.168.2.x. 110 ist der vorbereitete Olimex-Controller, 255 bedeutet Broadcast.",
    },
    {
        "pkey": "MAP0074",
        "ptype": "MAP",
        "pid": "0074",
        "default_v": "3050",
        "min_v": 1,
        "max_v": 65535,
        "unit": "port",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "format_relevant": "NO",
        "name": "LED-Controller UDP-Port",
        "message": "UDP-Port fuer MAS004-LED-UDP/v1 Frames zum externen LED-Controller.",
    },
    {
        "pkey": "MAP0075",
        "ptype": "MAP",
        "pid": "0075",
        "default_v": "100",
        "min_v": 20,
        "max_v": 1000,
        "unit": "ms",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint16",
        "format_relevant": "NO",
        "name": "LED-Controller Frame-Intervall",
        "message": "Minimaler Sendeabstand der LED-Frames. 100 ms entspricht 10 Hz.",
    },
)

MACHINE_PROCESS_PARAM_DEFS = (
    {
        "pkey": "MAP0076",
        "ptype": "MAP",
        "pid": "0076",
        "default_v": "8",
        "min_v": -50,
        "max_v": 50,
        "unit": "1/10mm",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "int16",
        "format_relevant": "NO",
        "name": "Label-Laengenkompensation",
        "message": (
            "Konstante Kompensation der vom Etikettenerfassungssensor gemessenen HIGH-Laenge. "
            "Default 8 = +0.8 mm, Laengenfehler brauchen Roh- und kompensierten Wert ausserhalb derselben Grenze."
        ),
    },
    {
        "pkey": "MAP0077",
        "ptype": "MAP",
        "pid": "0077",
        "default_v": "100765",
        "min_v": 1000,
        "max_v": 1000000,
        "unit": "1/1000mm",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint32",
        "format_relevant": "NO",
        "name": "Einlaufencoder Wirkdurchmesser",
        "message": "Fester Maschinen-Abgleichwert fuer die Einlaufencoder-Skalierung. Default 100765 = 100.765 mm.",
    },
    {
        "pkey": "MAP0078",
        "ptype": "MAP",
        "pid": "0078",
        "default_v": "100649",
        "min_v": 1000,
        "max_v": 1000000,
        "unit": "1/1000mm",
        "rw": "W",
        "esp_rw": "R",
        "dtype": "uint32",
        "format_relevant": "NO",
        "name": "Auslaufencoder Wirkdurchmesser",
        "message": "Fester Maschinen-Abgleichwert fuer die Auslauf-/ID3-Encoder-Skalierung. Default 100649 = 100.649 mm.",
    },
)


def resolve_videojet_logo_path() -> Optional[str]:
    for candidate in (VIDEOJET_LOGO_PATH, FALLBACK_VIDEOJET_LOGO_PATH):
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def resolve_repo_master_ios_xlsx() -> Optional[str]:
    for candidate in (REPO_MASTER_IOS_XLSX, FALLBACK_REPO_MASTER_IOS_XLSX):
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def require_token(x_token: Optional[str], cfg: Settings):
    if cfg.ui_token and x_token != cfg.ui_token:
        raise HTTPException(status_code=401, detail="Unauthorized")


def require_shared_secret(x_shared_secret: Optional[str], cfg: Settings):
    if (cfg.shared_secret or "") and x_shared_secret != cfg.shared_secret:
        raise HTTPException(status_code=401, detail="Unauthorized (shared secret)")


def require_token_or_shared_secret(x_token: Optional[str], x_shared_secret: Optional[str], cfg: Settings):
    if cfg.ui_token and x_token == cfg.ui_token:
        return
    require_shared_secret(x_shared_secret, cfg)


def sanitize_machine_setup_target(path: Optional[str]) -> str:
    target = (path or "").strip()
    if not target.startswith("/ui/machine-setup"):
        return "/ui/machine-setup/motors"
    return target


def machine_setup_secret(cfg: Settings) -> bytes:
    seed = f"mas004-machine-setup|{cfg.shared_secret}|{cfg.ui_token}|{cfg.db_path}"
    return hashlib.sha256(seed.encode("utf-8")).digest()


def build_machine_setup_cookie(cfg: Settings, issued_at: Optional[int] = None) -> str:
    ts = int(issued_at or time.time())
    payload = f"{MACHINE_SETUP_USER}|{ts}"
    sig = hmac.new(machine_setup_secret(cfg), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{payload}|{sig}"


def has_machine_setup_session(request: Request, cfg: Settings) -> bool:
    raw = (request.cookies.get(MACHINE_SETUP_COOKIE) or "").strip()
    if not raw:
        return False
    try:
        username, issued_raw, signature = raw.split("|", 2)
        issued_at = int(issued_raw)
    except Exception:
        return False
    if username != MACHINE_SETUP_USER:
        return False
    payload = f"{username}|{issued_at}"
    expected = hmac.new(machine_setup_secret(cfg), payload.encode("utf-8"), hashlib.sha256).hexdigest()
    if not secrets.compare_digest(expected, signature):
        return False
    return (int(time.time()) - issued_at) <= MACHINE_SETUP_SESSION_MAX_AGE_S


def require_machine_setup_session(request: Request, cfg: Settings):
    if not has_machine_setup_session(request, cfg):
        raise HTTPException(status_code=401, detail="Machine-Setup login required")


def machine_setup_cookie_kwargs(request: Request, cfg: Settings) -> dict[str, Any]:
    scheme = (request.url.scheme or "").lower()
    return {
        "httponly": True,
        "max_age": MACHINE_SETUP_SESSION_MAX_AGE_S,
        "samesite": "lax",
        "secure": bool(cfg.webui_https or scheme == "https"),
        "path": "/",
    }


def get_current_time_info() -> dict[str, Any]:
    info: dict[str, Any] = {
        "ok": True,
        "local_time": "",
        "universal_time": "",
        "rtc_time": "",
        "time_zone": "",
        "system_clock_synchronized": "",
        "ntp_service": "",
        "rtc_in_local_tz": "",
        "timesync_server": "",
        "timesync_address": "",
    }

    try:
        proc = subprocess.run(
            ["timedatectl"],
            capture_output=True,
            text=True,
            timeout=5,
            check=False,
        )
        output = (proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")
        for raw_line in output.splitlines():
            line = raw_line.strip()
            if ":" not in line:
                continue
            key, value = [part.strip() for part in line.split(":", 1)]
            key_l = key.lower()
            if key_l == "local time":
                info["local_time"] = value
            elif key_l == "universal time":
                info["universal_time"] = value
            elif key_l == "rtc time":
                info["rtc_time"] = value
            elif key_l == "time zone":
                info["time_zone"] = value
            elif key_l == "system clock synchronized":
                info["system_clock_synchronized"] = value
            elif key_l == "ntp service":
                info["ntp_service"] = value
            elif key_l == "rtc in local tz":
                info["rtc_in_local_tz"] = value
    except Exception as e:
        info["ok"] = False
        info["error"] = f"timedatectl failed: {e}"
        return info

    try:
        proc = subprocess.run(
            ["timedatectl", "show-timesync", "--all"],
            capture_output=True,
            text=True,
            timeout=5,
            check=False,
        )
        if proc.returncode == 0:
            for raw_line in (proc.stdout or "").splitlines():
                line = raw_line.strip()
                if "=" not in line:
                    continue
                key, value = [part.strip() for part in line.split("=", 1)]
                if key == "ServerName":
                    info["timesync_server"] = value
                elif key == "ServerAddress":
                    info["timesync_address"] = value
    except Exception:
        pass

    return info


class ConfigUpdate(BaseModel):
    # Microtom
    peer_base_url: Optional[str] = None
    peer_base_url_secondary: Optional[str] = None
    diclient_adapter_key: Optional[str] = None
    peer_watchdog_host: Optional[str] = None
    peer_health_path: Optional[str] = None

    # HTTP/tls
    tls_verify: Optional[bool] = None
    http_timeout_s: Optional[float] = None
    eth0_source_ip: Optional[str] = None
    ntp_server: Optional[str] = None
    ntp_sync_interval_min: Optional[int] = None

    # webui
    webui_port: Optional[int] = None
    ui_token: Optional[str] = None
    shared_secret: Optional[str] = None
    backup_root_path: Optional[str] = None
    machine_serial_number: Optional[str] = None
    machine_name: Optional[str] = None

    # device endpoints
    esp_host: Optional[str] = None
    esp_port: Optional[int] = None
    esp_simulation: Optional[bool] = None
    esp_watchdog_host: Optional[str] = None
    esp_io_poll_interval_s: Optional[float] = None
    raspi_plc_model: Optional[str] = None
    raspi_io_simulation: Optional[bool] = None
    raspi_io_poll_interval_s: Optional[float] = None
    light_curtain_auto_reset_enabled: Optional[bool] = None
    moxa1_host: Optional[str] = None
    moxa1_port: Optional[int] = None
    moxa1_simulation: Optional[bool] = None
    moxa2_host: Optional[str] = None
    moxa2_port: Optional[int] = None
    moxa2_simulation: Optional[bool] = None
    moxa3_host: Optional[str] = None
    moxa3_port: Optional[int] = None
    moxa3_simulation: Optional[bool] = None
    moxa_poll_interval_s: Optional[float] = None
    vj3350_host: Optional[str] = None
    vj3350_port: Optional[int] = None
    vj3350_simulation: Optional[bool] = None
    vj6530_host: Optional[str] = None
    vj6530_port: Optional[int] = None
    vj6530_simulation: Optional[bool] = None
    vj6530_poll_interval_s: Optional[float] = None
    vj6530_async_enabled: Optional[bool] = None
    smart_unwinder_host: Optional[str] = None
    smart_unwinder_port: Optional[int] = None
    smart_unwinder_simulation: Optional[bool] = None
    smart_rewinder_host: Optional[str] = None
    smart_rewinder_port: Optional[int] = None
    smart_rewinder_simulation: Optional[bool] = None

    # daily logfile retention
    logs_keep_days_all: Optional[int] = None
    logs_keep_days_esp: Optional[int] = None
    logs_keep_days_tto: Optional[int] = None
    logs_keep_days_laser: Optional[int] = None
    machine_audit_keep_hours: Optional[int] = None


class NetworkUpdate(BaseModel):
    eth0_ip: str
    eth0_prefix: int
    eth0_gateway: str = ""
    eth0_dns: str = ""
    eth1_ip: str
    eth1_prefix: int
    eth1_gateway: str = ""
    eth1_dns: str = ""
    apply_now: bool = False  # wenn true -> versucht Netzwerk live umzustellen


class OutboxEnqueue(BaseModel):
    method: str = "POST"
    path: str = "/api/inbox"
    url: Optional[str] = None
    headers: Dict[str, Any] = {}
    body: Optional[Dict[str, Any]] = None
    idempotency_key: Optional[str] = None


class ParamEdit(BaseModel):
    pkey: str
    default_v: Optional[str] = None
    min_v: Optional[float] = None
    max_v: Optional[float] = None
    rw: Optional[str] = None
    esp_rw: Optional[str] = None


class TestSendReq(BaseModel):
    source: str
    msg: str
    ptype_hint: Optional[str] = None


class MotorMoveReq(BaseModel):
    mode: str
    value: float


class MotorPositionReq(BaseModel):
    value: float


class MotorConfigReq(BaseModel):
    steps_per_mm: Optional[float] = None
    speed_mm_s: Optional[float] = None
    accel_mm_s2: Optional[float] = None
    decel_mm_s2: Optional[float] = None
    current_pct: Optional[float] = None
    hold_current_pct: Optional[float] = None
    invert_direction: Optional[bool] = None
    min_tenths_mm: Optional[int] = None
    max_tenths_mm: Optional[int] = None
    min_enabled: Optional[bool] = None
    max_enabled: Optional[bool] = None
    position_mm: Optional[float] = None


class MotorSimulationReq(BaseModel):
    enabled: bool


class MotorPollReq(BaseModel):
    enabled: bool


class EspCommandReq(BaseModel):
    line: str
    read_timeout_s: float = 2.0
    read_limit: int = 8192
    priority: bool = False


class IoWriteReq(BaseModel):
    value: int


class IoOverrideReq(BaseModel):
    value: int


class MachineSetupLoginReq(BaseModel):
    username: str
    password: str
    next: Optional[str] = "/ui/machine-setup/motors"


class MachineButtonReq(BaseModel):
    button: str


class MachineLedTestReq(BaseModel):
    action: str = "start"
    duration_ms: int = 0


class ProductionVisualizationComponentReq(BaseModel):
    key: str
    mm: float


class Motor3CalibrationApplyReq(BaseModel):
    actual_travel_mm: float
    actual_label_length_mm: Optional[float] = None


class MachineAuditRetentionReq(BaseModel):
    keep_hours: int


class MachineBypassReq(BaseModel):
    values: Dict[str, Any]


class FormatProfileReq(BaseModel):
    name: str
    values: Dict[str, Any]
    note: str = ""


class FormatSendReq(BaseModel):
    name: str = ""
    values: Dict[str, Any]


class CommissioningStartReq(BaseModel):
    mode: str = "full"


class CommissioningStepReq(BaseModel):
    status: str
    note: str = ""


class BackupCreateReq(BaseModel):
    backup_type: str
    name: str
    note: str = ""


class BackupIdentityReq(BaseModel):
    machine_serial_number: str
    machine_name: str = ""


def build_app(cfg_path: str = DEFAULT_CFG_PATH) -> FastAPI:
    app = FastAPI(title="MAS-004_RPI-Databridge", version="0.3.0", docs_url=None)
    
    cfg = Settings.load(cfg_path)
    db = DB(cfg.db_path)
    outbox = Outbox(db)
    inbox = Inbox(db)
    params = ParamStore(db)
    io_store = IoStore(db)
    logs = LogStore(db)
    production_logs = ProductionLogManager(db, cfg=cfg, outbox=outbox)
    format_profiles = FormatProfileStore(db)
    if not os.path.exists(cfg.master_params_xlsx_path) and os.path.exists(REPO_MASTER_PARAMS_XLSX):
        os.makedirs(os.path.dirname(cfg.master_params_xlsx_path), exist_ok=True)
        shutil.copyfile(REPO_MASTER_PARAMS_XLSX, cfg.master_params_xlsx_path)
    repo_master_ios_xlsx = resolve_repo_master_ios_xlsx()
    if not os.path.exists(cfg.master_ios_xlsx_path) and repo_master_ios_xlsx:
        os.makedirs(os.path.dirname(cfg.master_ios_xlsx_path), exist_ok=True)
        shutil.copyfile(repo_master_ios_xlsx, cfg.master_ios_xlsx_path)
    if io_store.count_points() == 0 and os.path.exists(cfg.master_ios_xlsx_path):
        io_store.import_xlsx(cfg.master_ios_xlsx_path)
    test_sources = {"raspi", "esp-plc", "vj3350", "vj6530"}
    default_ptype_hint = {"raspi": "", "esp-plc": "MAS", "vj3350": "LSE", "vj6530": "TTE"}
    catalog_ids = [int(item["id"]) for item in motor_catalog()]
    motor_state_store = MotorStateStore(cfg)
    motor_bindings_cache: dict[str, Any] = {"ts": 0.0, "value": {}, "loaded": False}
    motor_refresh_guard = threading.Lock()
    motor_refresh_locks: dict[int, threading.Lock] = {}
    motor_refresh_cache: dict[int, tuple[float, dict[str, Any]]] = {}
    motor3_calibration_guard = threading.Lock()
    motor3_calibration_process: dict[str, Any] = {"proc": None, "mode": "", "started_ts": 0.0}

    def get_motor_client() -> EspMotorClient:
        return EspMotorClient(Settings.load(cfg_path))

    def get_motor_state_store() -> MotorStateStore:
        return motor_state_store

    def get_motor_refresh_lock(motor_id: int) -> threading.Lock:
        mid = int(motor_id)
        with motor_refresh_guard:
            lock = motor_refresh_locks.get(mid)
            if lock is None:
                lock = threading.Lock()
                motor_refresh_locks[mid] = lock
            return lock

    def get_cached_motor_refresh(motor_id: int, max_age_s: float) -> Optional[dict[str, Any]]:
        now = time.monotonic()
        with motor_refresh_guard:
            cached = motor_refresh_cache.get(int(motor_id))
            if not cached:
                return None
            ts, payload = cached
            if (now - ts) > max_age_s:
                return None
            out = json.loads(json.dumps(payload))
        if isinstance(out, dict):
            out["cached"] = True
        return out

    def remember_motor_refresh(motor_id: int, payload: dict[str, Any]) -> None:
        with motor_refresh_guard:
            motor_refresh_cache[int(motor_id)] = (time.monotonic(), json.loads(json.dumps(payload)))

    def last_known_motor_payload(motor_id: int) -> Optional[dict[str, Any]]:
        motor = get_motor_state_store().cached_motors().get(int(motor_id))
        return json.loads(json.dumps(motor)) if isinstance(motor, dict) else None

    def get_commissioning_store(cfg2: Optional[Settings] = None) -> CommissioningStore:
        return CommissioningStore(db, cfg2 or Settings.load(cfg_path), cfg_path)

    def get_backup_manager(cfg2: Optional[Settings] = None) -> MachineBackupManager:
        return MachineBackupManager(db, cfg2 or Settings.load(cfg_path), cfg_path)

    def get_motor_bindings() -> dict[int, dict[str, Any]]:
        now = time.monotonic()
        if bool(motor_bindings_cache.get("loaded")) and (now - float(motor_bindings_cache.get("ts") or 0.0)) < 30.0:
            return motor_bindings_cache.get("value") or {}
        with db._conn() as c:
            rows = c.execute(
                """SELECT p.pkey,p.ptype,p.pid,p.unit,p.rw,p.esp_rw,p.dtype,p.name,p.message,p.ai_instructions
                   FROM params p
                   WHERE p.ptype IN ('MAP','MAS','MAE')
                     AND COALESCE(TRIM(p.ai_instructions), '') <> ''
                   ORDER BY p.ptype ASC, p.pid ASC"""
            ).fetchall()
        compact_rows = [
            {
                "pkey": row[0],
                "ptype": row[1],
                "pid": row[2],
                "unit": row[3],
                "rw": row[4],
                "esp_rw": row[5],
                "dtype": row[6],
                "name": row[7],
                "message": row[8],
                "ai_instructions": row[9],
            }
            for row in rows
        ]
        value = {int(item["motor_id"]): item for item in build_motor_bindings(compact_rows)}
        motor_bindings_cache["ts"] = now
        motor_bindings_cache["value"] = value
        motor_bindings_cache["loaded"] = True
        return value

    def get_simulated_motor_ids(cfg2: Optional[Settings] = None) -> set[int]:
        cfg_local = cfg2 or Settings.load(cfg_path)
        stored = get_motor_state_store().simulation_ids()
        if bool(getattr(cfg_local, "esp_simulation", False)):
            return set(catalog_ids)
        return stored

    def motor_live_allowed(motor_id: int, cfg2: Optional[Settings] = None) -> bool:
        return int(motor_id) not in get_simulated_motor_ids(cfg2)

    def require_live_motor_or_raise(motor_id: int, cfg2: Optional[Settings] = None) -> Settings:
        cfg_local = cfg2 or Settings.load(cfg_path)
        if not motor_live_allowed(motor_id, cfg_local):
            raise HTTPException(status_code=409, detail=f"Motor {int(motor_id)} ist im Simulationsbetrieb")
        client = EspMotorClient(cfg_local)
        if not client.available():
            raise HTTPException(status_code=503, detail="ESP motor endpoint missing or simulation enabled")
        return cfg_local

    def get_motor_overview_payload() -> dict[str, Any]:
        cfg2 = Settings.load(cfg_path)
        client = get_motor_client()
        store = get_motor_state_store()
        cached_motors = store.cached_motors()
        simulated_ids = get_simulated_motor_ids(cfg2)
        live_target_ids = [mid for mid in catalog_ids if mid not in simulated_ids]
        live_items: list[dict[str, Any]] = []
        live_successes: list[dict[str, Any]] = []
        live_error_count = 0
        live_error = ""
        motor_poll: Optional[dict[str, Any]] = None

        if live_target_ids and not client.available():
            live_error = "ESP-Motor-Endpoint nicht erreichbar"

        if live_successes:
            store.remember_motors(live_successes)
            cached_motors = store.cached_motors()
        live_available = bool(live_successes)

        merged = merge_motor_payload(
            {
                "ok": True,
                "motors": live_items,
                "error": live_error,
                "live_available": live_available,
                "motor_poll": motor_poll,
            },
            get_motor_bindings(),
            simulated_ids=simulated_ids,
            cached_motors=cached_motors,
        )
        if live_target_ids and live_error and not live_successes:
            count = len(live_target_ids)
            merged["message"] = f"{count} Live-Motor{'en' if count != 1 else ''} derzeit nicht erreichbar"
        elif live_error_count:
            merged["message"] = f"{live_error_count} Live-Motor{'en' if live_error_count != 1 else ''} derzeit nicht erreichbar"
        elif live_target_ids:
            merged["message"] = "Motorstatus ist gecached; Live-Werte pro Motor mit Status aktualisieren oder 1s Polling lesen"
        else:
            merged["message"] = ""
        merged["simulated_ids"] = sorted(simulated_ids)
        merged["motor_poll"] = motor_poll
        return merged

    def get_motor_poll_payload(cfg2: Optional[Settings] = None) -> dict[str, Any]:
        cfg_local = cfg2 or Settings.load(cfg_path)
        client = EspMotorClient(cfg_local)
        if not client.available():
            return {
                "ok": False,
                "available": False,
                "auto_poll": False,
                "error": "ESP motor endpoint missing or simulation enabled",
            }
        return {
            "ok": True,
            "available": True,
            "auto_poll": False,
            "interval_ms": None,
            "next_motor": None,
            "message": "Globales ESP-Motor-Round-Robin ist deaktiviert; Livewerte nur explizit pro Motor.",
        }

    def touch_motor_setup_authority(motor_id: Optional[int] = None, reason: str = "motor_setup_webui") -> None:
        try:
            touch_motor_setup_manual_lock(db, motor_id=motor_id, reason=reason, source="webui")
        except Exception as exc:
            logs.log("machine", "warning", f"Motor-Setup-Sperre konnte nicht gesetzt werden: {exc}")

    def require_motor_ack(result: dict[str, Any], action: str) -> None:
        if not isinstance(result, dict):
            raise HTTPException(status_code=502, detail=f"ESP returned invalid reply for {action}")
        if not bool(result.get("ok", True)):
            detail = result.get("reply") or result.get("error") or f"ESP rejected {action}"
            raise HTTPException(status_code=502, detail=str(detail))

    def motor_client_call(action: str, fn):
        try:
            return fn()
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(
                status_code=502,
                detail=f"ESP motor communication failed during {action}: {exc}",
            ) from exc

    def refresh_motor_snapshot(
        client: EspMotorClient,
        motor_id: int,
        *,
        allow_recent_cache: bool = True,
    ) -> dict[str, Any]:
        if allow_recent_cache:
            cached = get_cached_motor_refresh(motor_id, MOTOR_REFRESH_CACHE_TTL_S)
            if cached is not None:
                return cached

        lock = get_motor_refresh_lock(motor_id)
        acquired = lock.acquire(blocking=False)
        if not acquired:
            # Another browser tab/action is already doing the expensive Modbus
            # read for this motor. Wait briefly for that result instead of
            # queueing a second REFRESH against the single ESP TCP endpoint.
            acquired = lock.acquire(timeout=MOTOR_REFRESH_BUSY_WAIT_S)
            if not acquired:
                cached = get_cached_motor_refresh(motor_id, 10.0)
                if cached is not None:
                    cached["refresh_in_progress"] = True
                    return cached
                stored = last_known_motor_payload(motor_id)
                payload: dict[str, Any] = {
                    "ok": True,
                    "refresh_in_progress": True,
                    "cached": bool(stored),
                    "message": f"Motor {int(motor_id)} Refresh laeuft bereits",
                }
                if stored is not None:
                    payload["motor"] = stored
                return payload
            if allow_recent_cache:
                cached = get_cached_motor_refresh(motor_id, MOTOR_REFRESH_CACHE_TTL_S)
                if cached is not None:
                    lock.release()
                    return cached

        try:
            if allow_recent_cache:
                cached = get_cached_motor_refresh(motor_id, MOTOR_REFRESH_CACHE_TTL_S)
                if cached is not None:
                    return cached
            payload = client.refresh(motor_id)
            raw_motor = payload.get("motor") if isinstance(payload, dict) else None
            if isinstance(raw_motor, dict):
                get_motor_state_store().remember_motors([raw_motor])
            if isinstance(payload, dict):
                remember_motor_refresh(motor_id, payload)
            return payload
        finally:
            if acquired:
                lock.release()

    def _to_int_or_none(value: Any) -> Optional[int]:
        try:
            return int(float(str(value).strip()))
        except Exception:
            return None

    def guard_motor_position_move(
        client: EspMotorClient,
        motor_id: int,
        mode: str,
        value: float,
    ) -> None:
        if int(motor_id) == 3:
            return
        cfg_payload = motor_client_call("CONFIG", lambda: client.config(motor_id))
        motor_cfg = (cfg_payload or {}).get("config") or {}
        is_positional = bool(motor_cfg.get("min_enabled", True)) or bool(motor_cfg.get("max_enabled", True))
        if not is_positional:
            return
        refresh_payload = refresh_motor_snapshot(client, motor_id, allow_recent_cache=False)
        motor = (refresh_payload or {}).get("motor") or {}
        state = motor.get("state") or motor or {}
        feedback_tenths = _to_int_or_none(state.get("feedback_tenths_mm"))
        if feedback_tenths is None:
            raise HTTPException(status_code=409, detail=f"Motor {int(motor_id)} Istposition nicht lesbar")
        if bool(state.get("alarm")):
            raise HTTPException(
                status_code=409,
                detail=f"Motor {int(motor_id)} ist im Alarm ({state.get('alarm_code')}); Bewegung gesperrt",
            )
        if bool(state.get("hwto")):
            raise HTTPException(status_code=409, detail=f"Motor {int(motor_id)} HWTO/Sicherheitskreis aktiv")

        min_enabled = bool(motor_cfg.get("min_enabled", True))
        max_enabled = bool(motor_cfg.get("max_enabled", True))
        min_tenths = _to_int_or_none(motor_cfg.get("min_tenths_mm"))
        max_tenths = _to_int_or_none(motor_cfg.get("max_tenths_mm"))
        if min_enabled and min_tenths is not None and feedback_tenths < min_tenths:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Motor {int(motor_id)} steht unter Min-Grenze "
                    f"({feedback_tenths / 10.0:.1f}mm < {min_tenths / 10.0:.1f}mm); Bewegung gesperrt"
                ),
            )
        if max_enabled and max_tenths is not None and feedback_tenths > max_tenths:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Motor {int(motor_id)} steht ueber Max-Grenze "
                    f"({feedback_tenths / 10.0:.1f}mm > {max_tenths / 10.0:.1f}mm); Bewegung gesperrt"
                ),
            )

        mode_norm = (mode or "").strip().lower()
        if mode_norm == "absolute_mm":
            target_tenths = int(round(float(value) * 10.0))
        elif mode_norm == "relative_mm":
            target_tenths = feedback_tenths + int(round(float(value) * 10.0))
        elif mode_norm == "relative_steps":
            steps_per_mm = float(motor_cfg.get("steps_per_mm") or 0.0)
            if steps_per_mm <= 0.0:
                raise HTTPException(status_code=409, detail=f"Motor {int(motor_id)} steps/mm ungueltig")
            target_tenths = feedback_tenths + int(round((float(value) / steps_per_mm) * 10.0))
        else:
            raise HTTPException(status_code=400, detail="Unsupported motor move mode")

        if min_enabled and min_tenths is not None and target_tenths < min_tenths:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Motor {int(motor_id)} Ziel unter Min-Grenze "
                    f"({target_tenths / 10.0:.1f}mm < {min_tenths / 10.0:.1f}mm); Bewegung gesperrt"
                ),
            )
        if max_enabled and max_tenths is not None and target_tenths > max_tenths:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Motor {int(motor_id)} Ziel ueber Max-Grenze "
                    f"({target_tenths / 10.0:.1f}mm > {max_tenths / 10.0:.1f}mm); Bewegung gesperrt"
                ),
            )

    def motor_action_response(
        client: EspMotorClient,
        motor_id: int,
        action: str,
        result: dict[str, Any],
        *,
        refresh: bool = True,
    ) -> dict[str, Any]:
        require_motor_ack(result, action)
        response: dict[str, Any] = {
            "ok": True,
            "action": action,
            "reply": result.get("reply") if isinstance(result, dict) else "",
        }
        if refresh:
            try:
                refresh_payload = refresh_motor_snapshot(client, motor_id, allow_recent_cache=False)
                response["refresh"] = refresh_payload
                raw_motor = refresh_payload.get("motor") if isinstance(refresh_payload, dict) else None
                if isinstance(raw_motor, dict):
                    response["motor"] = raw_motor
            except Exception as exc:
                # Command ACK is still useful; the UI shows if the follow-up refresh failed.
                response["refresh_error"] = str(exc)
        return response

    def motor_status_action_response(
        client: EspMotorClient,
        motor_id: int,
        action: str,
        result: dict[str, Any],
    ) -> dict[str, Any]:
        require_motor_ack(result, action)
        response: dict[str, Any] = {
            "ok": True,
            "action": action,
            "reply": result.get("reply") if isinstance(result, dict) else "",
        }
        try:
            status_payload = client.status(motor_id)
            response["status"] = status_payload
            raw_motor = status_payload.get("motor") if isinstance(status_payload, dict) else None
            if isinstance(raw_motor, dict):
                get_motor_state_store().remember_motors([raw_motor])
                response["motor"] = raw_motor
        except Exception as exc:
            response["status_error"] = str(exc)
        return response

    def motor_config_action_response(
        cfg2: Settings,
        client: EspMotorClient,
        motor_id: int,
        action: str,
        result: dict[str, Any],
        *,
        config_payload: Optional[dict[str, Any]] = None,
    ) -> dict[str, Any]:
        require_motor_ack(result, action)
        response: dict[str, Any] = {
            "ok": True,
            "action": action,
            "reply": result.get("reply") if isinstance(result, dict) else "",
        }
        cfg_payload = motor_client_call("CONFIG", lambda: client.config(motor_id))
        response["config"] = cfg_payload
        current_config = cfg_payload.get("config") if isinstance(cfg_payload, dict) else None
        motor: dict[str, Any] = {"id": int(motor_id)}
        try:
            refresh_payload = refresh_motor_snapshot(client, motor_id, allow_recent_cache=False)
            response["refresh"] = refresh_payload
            raw_motor = refresh_payload.get("motor") if isinstance(refresh_payload, dict) else None
            if isinstance(raw_motor, dict):
                motor = raw_motor
        except Exception as exc:
            response["refresh_error"] = str(exc)
            try:
                status_payload = client.status(motor_id)
                response["status"] = status_payload
                raw_motor = status_payload.get("motor") if isinstance(status_payload, dict) else None
                if isinstance(raw_motor, dict):
                    get_motor_state_store().remember_motors([raw_motor])
                    motor = raw_motor
            except Exception as status_exc:
                response["status_error"] = str(status_exc)
                motor = last_known_motor_payload(motor_id) or motor
        if isinstance(current_config, dict):
            motor["config"] = current_config
        response["motor"] = motor
        if config_payload is not None:
            verify_motor_config_applied(motor_id, response, config_payload)
        return sync_motor_master_from_response(cfg2, motor_id, action, response)

    def _motor_config_values_equal(actual: Any, expected: Any) -> bool:
        if isinstance(expected, bool):
            if isinstance(actual, bool):
                return actual is expected
            return str(actual).strip().lower() in ("1", "true", "yes", "on") if expected else str(actual).strip().lower() in ("0", "false", "no", "off")
        try:
            if isinstance(expected, int) and not isinstance(expected, bool):
                return int(round(float(actual))) == int(expected)
            return abs(float(actual) - float(expected)) <= 0.001
        except Exception:
            return str(actual).strip() == str(expected).strip()

    def verify_motor_config_applied(motor_id: int, response: dict[str, Any], payload: dict[str, Any]) -> None:
        motor = response.get("motor") if isinstance(response, dict) else None
        config = motor.get("config") if isinstance(motor, dict) else None
        if not isinstance(config, dict):
            raise HTTPException(
                status_code=502,
                detail=f"Motor {int(motor_id)} Parameter gespeichert, aber Refresh/Config nicht lesbar",
            )
        mismatches: list[str] = []
        for key, expected in payload.items():
            if expected is None:
                continue
            if key not in config or not _motor_config_values_equal(config.get(key), expected):
                mismatches.append(f"{key}: soll {expected}, ist {config.get(key)}")
        if mismatches:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Motor {int(motor_id)} Parameter wurden vom ESP nach SAVE nicht bestaetigt: "
                    + "; ".join(mismatches)
                ),
            )

    def verify_motor_feedback_position(motor_id: int, response: dict[str, Any], expected_mm: float) -> None:
        motor = response.get("motor") if isinstance(response, dict) else None
        state = motor.get("state") if isinstance(motor, dict) else None
        feedback = _to_int_or_none((state or {}).get("feedback_tenths_mm")) if isinstance(state, dict) else None
        expected_tenths = int(round(float(expected_mm) * 10.0))
        if feedback is None:
            raise HTTPException(
                status_code=502,
                detail=f"Motor {int(motor_id)} Istposition gespeichert, aber Refresh/Istwert nicht lesbar",
            )
        if abs(int(feedback) - expected_tenths) > 1:
            raise HTTPException(
                status_code=409,
                detail=(
                    f"Motor {int(motor_id)} Istposition wurde nicht uebernommen: "
                    f"soll {expected_tenths / 10.0:.1f}mm, ist {int(feedback) / 10.0:.1f}mm"
                ),
            )

    def save_and_refresh_motor_master(
        cfg2: Settings,
        client: EspMotorClient,
        motor_id: int,
        action: str,
        result: dict[str, Any],
        *,
        config_payload: Optional[dict[str, Any]] = None,
        expected_position_mm: Optional[float] = None,
        do_save: bool = True,
        refresh_after_save: bool = True,
    ) -> dict[str, Any]:
        require_motor_ack(result, action)
        reply_parts = [str(result.get("reply") or "").strip()]
        if do_save:
            save_result = motor_client_call(
                "SAVE",
                lambda: client.save(motor_id, allow_machine_setup_write=True),
            )
            require_motor_ack(save_result, "SAVE")
            reply_parts.append(str(save_result.get("reply") or "").strip())
        combined = {"ok": True, "reply": "; ".join(part for part in reply_parts if part)}
        if not refresh_after_save:
            response = motor_config_action_response(
                cfg2,
                client,
                motor_id,
                action,
                combined,
                config_payload=config_payload,
            )
            if expected_position_mm is not None:
                verify_motor_feedback_position(motor_id, response, expected_position_mm)
            return response
        response = motor_action_response(client, motor_id, action, combined)
        if config_payload is not None:
            verify_motor_config_applied(motor_id, response, config_payload)
        if expected_position_mm is not None:
            verify_motor_feedback_position(motor_id, response, expected_position_mm)
        return sync_motor_master_from_response(cfg2, motor_id, action, response)

    def motor_master_workbook_paths(cfg2: Settings) -> list[str]:
        paths: list[str] = []
        if cfg2.master_params_xlsx_path:
            paths.append(cfg2.master_params_xlsx_path)
        paths.extend(
            [
                REPO_MASTER_PARAMS_XLSX,
                WORKSPACE_MASTER_PARAMS_XLSX,
                FALLBACK_REPO_MASTER_PARAMS_XLSX,
            ]
        )
        return paths

    def restore_motor_master_config_to_esp(
        cfg2: Settings,
        client: EspMotorClient,
        motor_id: int,
        *,
        restore_position: bool = True,
    ) -> dict[str, Any]:
        restore = motor_client_call(
            "RESTORE_MASTER_CONFIG",
            lambda: apply_motor_setup_master_config_to_client(
                params,
                client,
                motor_id,
                restore_position=restore_position,
            ),
        )
        require_motor_ack(restore, "RESTORE_MASTER_CONFIG")
        payload = restore.get("payload") or {}
        if not payload:
            return restore
        combined = {
            "ok": True,
            "reply": str(restore.get("reply") or "").strip(),
        }
        response = motor_action_response(client, motor_id, "RESTORE_MASTER_CONFIG", combined)
        try:
            verify_payload = dict(payload)
            if restore.get("position_restored"):
                # SET_POSITION_MM deliberately recalculates the ESP zero offset
                # from the current AZD feedback.  The restored position is the
                # source of truth, so the old stored zero offset must not make
                # this verification fail immediately after a correct restore.
                verify_payload.pop("zero_offset_steps", None)
            verify_motor_config_applied(motor_id, response, verify_payload)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc
        response = sync_motor_master_from_response(cfg2, motor_id, "RESTORE_MASTER_CONFIG", response)
        response["restored"] = True
        return response

    def sync_motor_master_from_response(
        cfg2: Settings,
        motor_id: int,
        action: str,
        response: dict[str, Any],
    ) -> dict[str, Any]:
        raw_motor = response.get("motor") if isinstance(response, dict) else None
        if not isinstance(raw_motor, dict):
            return response
        try:
            sync_result = sync_motor_master_values(
                params,
                cfg2,
                motor_id,
                raw_motor,
                get_motor_bindings(),
                workbook_paths=motor_master_workbook_paths(cfg2),
                sync_position_default=action in (
                    "SET_POSITION_MM",
                    "ZERO",
                    "SET_CONFIG",
                    "SAVE",
                    "SET_MIN",
                    "SET_MAX",
                ),
                allow_protected_position_param_write=True,
            )
            response["master_sync"] = sync_result
            if sync_result.get("updated_pkeys"):
                motor_bindings_cache["loaded"] = False
                logs.log(
                    "machine",
                    "info",
                    (
                        f"Motor {int(motor_id)} Setup-Master nach {action} synchronisiert: "
                        + ", ".join(sync_result.get("updated_pkeys") or [])
                    ),
                )
            if sync_result.get("db_errors"):
                logs.log(
                    "machine",
                    "warning",
                    f"Motor {int(motor_id)} Setup-Master teilweise nicht synchronisiert: {sync_result.get('db_errors')}",
                )
        except Exception as exc:
            response["master_sync_error"] = str(exc)
            logs.log("machine", "warning", f"Motor {int(motor_id)} Setup-Master-Sync fehlgeschlagen: {exc}")
        return response

    def get_winder_state(role: str) -> dict[str, Any]:
        normalized = normalize_winder_role(role)
        return SmartWicklerClient(Settings.load(cfg_path), normalized).fetch_state()

    def normalize_test_source(source: str) -> str:
        s = (source or "").strip().lower()
        if s not in test_sources:
            raise HTTPException(status_code=400, detail=f"Unknown source '{source}'")
        return s

    def normalize_test_line(raw_msg: str, ptype_hint: Optional[str]) -> str:
        s = (raw_msg or "").strip()
        if not s:
            raise HTTPException(status_code=400, detail="Empty message")

        m_full = re.match(r"^\s*([A-Za-z]{3})([0-9A-Za-z_]+)\s*=\s*(.+?)\s*$", s)
        if m_full:
            return f"{m_full.group(1).upper()}{m_full.group(2)}={m_full.group(3).strip()}"

        hint = (ptype_hint or "").strip().upper()
        if hint and not re.match(r"^[A-Z]{3}$", hint):
            raise HTTPException(status_code=400, detail="ptype_hint must be 3 letters (e.g. TTE, MAP, MAS)")

        m_short = re.match(r"^\s*([0-9A-Za-z_]+)\s*=\s*(.+?)\s*$", s)
        if m_short and hint:
            return f"{hint}{m_short.group(1)}={m_short.group(2).strip()}"

        return s

    def split_test_messages(raw_msg: str) -> list[str]:
        text = (raw_msg or "").strip()
        if not text:
            raise HTTPException(status_code=400, detail="Empty message")
        normalized = text.replace("\r\n", "\n").replace("\r", "\n")
        parts = [p.strip() for p in re.split(r"[,\n;]+", normalized) if p.strip()]
        if not parts:
            raise HTTPException(status_code=400, detail="Empty message")
        return parts

    def logo_html() -> str:
        return """
<div style="display:flex; align-items:center; justify-content:flex-start; margin-bottom:8px;">
  <img src="/ui/assets/videojet-logo.jpg" alt="Videojet" style="display:block; height:72px; width:auto; max-width:100%; object-fit:contain;"/>
</div>
"""

    def nav_button_html(key: str, href: str, label: str, active: str) -> str:
        is_active = key == active
        cls = "navbtn active" if is_active else "navbtn"
        base_style = (
            "display:inline-flex; align-items:center; justify-content:center; "
            "min-height:34px; padding:8px 12px; border:1px solid #cbd7e5; "
            "border-radius:8px; background:#ffffff; color:#1f2933; "
            "font-weight:700; font-size:14px; line-height:1.1; "
            "text-decoration:none; white-space:nowrap; box-sizing:border-box;"
        )
        active_style = " background:#005eb8; color:#ffffff; border-color:#005eb8;" if is_active else ""
        current = ' aria-current="page"' if is_active else ""
        return f'<a class="{cls}" href="{href}" style="{base_style}{active_style}"{current}>{label}</a>'

    def nav_html(active: str) -> str:
        items = [
            ("home", "/", "Home"),
            ("params", "/ui/params", "Parameter"),
            ("test", "/ui/test", "Test UI"),
            ("docs", "/docs", "API Docs"),
            ("settings", "/ui/settings", "Settings"),
            ("machine_setup", "/ui/machine-setup", "Machine-Setup"),
        ]
        links = []
        for key, href, label in items:
            links.append(nav_button_html(key, href, label, active))
        return (
            '<div style="position:sticky; top:0; z-index:60; background:#f4f6f9; '
            'padding:8px 0 10px 0; margin-bottom:8px;">'
            + logo_html()
            + '<nav class="topnav" style="display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin:0 0 2px 0;">'
            + "".join(links)
            + "</nav></div>"
        )

    def machine_setup_nav_html(active: str) -> str:
        items = [
            ("commissioning", "/ui/machine-setup/commissioning", "Commissioning"),
            ("backups", "/ui/machine-setup/backups", "Backups"),
            ("production", "/ui/machine-setup/production", "Produktion"),
            ("visualization", "/ui/machine-setup/visualization", "Visualisierung"),
            ("mae0048", "/ui/machine-setup/mae0048", "MAE0048"),
            ("calibration", "/ui/machine-setup/calibration", "Kalibrierung"),
            ("process", "/ui/machine-setup/process", "Control / Audit"),
            ("io", "/ui/machine-setup/io", "I/O"),
            ("motors", "/ui/machine-setup/motors", "Motors"),
            ("unwinder", "/ui/machine-setup/winders/unwinder", "Abwickler"),
            ("rewinder", "/ui/machine-setup/winders/rewinder", "Aufwickler"),
        ]
        links = []
        for key, href, label in items:
            links.append(nav_button_html(key, href, label, active))
        links.append(nav_button_html("logout", "/ui/machine-setup/logout", "Logout", active))
        return (
            '<div style="margin:-2px 0 14px 0;">'
            '<div style="font-size:12px; font-weight:700; color:#5f6b7a; text-transform:uppercase; letter-spacing:0; margin-bottom:8px;">'
            'Machine-Setup'
            '</div>'
            '<nav class="topnav" style="display:flex; gap:6px; flex-wrap:wrap; align-items:center; margin:0;">'
            + "".join(links)
            + "</nav></div>"
        )

    def machine_setup_login_html(next_path: str, error: str = "") -> str:
        nav = nav_html("machine_setup")
        error_html = ""
        if error:
            error_html = f'<div style="padding:10px 12px; border:1px solid #f3c4c4; background:#fdecec; color:#8a1c1c; border-radius:10px; margin-bottom:12px;">{html.escape(error)}</div>'
        hidden_next = html.escape(next_path, quote=True)
        return f"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>Machine-Setup Login</title>
  <style>
    body{{margin:0; font-family:Segoe UI,Arial,sans-serif; background:#f4f6f9; color:#1f2933}}
    .wrap{{max-width:1500px; margin:0 auto; padding:16px}}
    .topnav{{display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px}}
    .navbtn{{padding:8px 12px; border:1px solid #d6dde7; border-radius:8px; background:#fff; color:#1f2933; text-decoration:none}}
    .navbtn.active{{background:#005eb8; color:#fff; border-color:#005eb8}}
    .card{{max-width:480px; background:#fff; border:1px solid #d6dde7; border-radius:12px; padding:18px}}
    .field{{display:flex; flex-direction:column; gap:4px; margin-bottom:12px}}
    .field label{{font-size:12px; color:#5f6b7a; font-weight:700}}
    input{{width:100%; min-height:40px; padding:10px 12px; border:1px solid #d6dde7; border-radius:10px}}
    button{{min-height:40px; padding:9px 14px; border:1px solid #a8bfd8; border-radius:10px; background:#e8f0f8; color:#17324b; font-weight:700; cursor:pointer}}
    button:hover{{background:#d9e7f5}}
    .muted{{color:#5f6b7a}}
  </style>
</head>
<body>
  <div class="wrap">
    {nav}
    <div class="card">
      <h2 style="margin-top:0">Machine-Setup Login</h2>
      <p class="muted">Dieser Bereich ist separat geschuetzt. Bitte mit dem Machine-Setup-Admin anmelden.</p>
      {error_html}
      <form onsubmit="return submitLogin(event)">
        <input type="hidden" name="next" value="{hidden_next}"/>
        <div class="field">
          <label>Benutzer</label>
          <input id="machineSetupUser" name="username" autocomplete="username" required/>
        </div>
        <div class="field">
          <label>Passwort</label>
          <input id="machineSetupPassword" name="password" type="password" autocomplete="current-password" required/>
        </div>
        <button type="submit">Anmelden</button>
      </form>
      <div id="machineSetupError" style="display:none; padding:10px 12px; border:1px solid #f3c4c4; background:#fdecec; color:#8a1c1c; border-radius:10px; margin-top:12px;"></div>
    </div>
  </div>
  <script>
    async function submitLogin(event) {{
      event.preventDefault();
      const payload = {{
        username: document.getElementById("machineSetupUser").value.trim(),
        password: document.getElementById("machineSetupPassword").value,
        next: document.querySelector('input[name="next"]').value || "/ui/machine-setup/motors"
      }};
      const errorBox = document.getElementById("machineSetupError");
      errorBox.style.display = "none";
      errorBox.textContent = "";
      const r = await fetch("/ui/machine-setup/login", {{
        method: "POST",
        headers: {{"Content-Type":"application/json"}},
        body: JSON.stringify(payload)
      }});
      const txt = await r.text();
      let data = {{}};
      try {{ data = JSON.parse(txt); }} catch(e) {{}}
      if(!r.ok) {{
        errorBox.textContent = (data && data.detail) ? data.detail : (`HTTP ${{r.status}}`);
        errorBox.style.display = "block";
        return false;
      }}
      window.location.href = (data && data.redirect) ? data.redirect : payload.next;
      return false;
    }}
  </script>
</body>
</html>
"""

    @app.get("/ui/assets/videojet-logo.jpg", include_in_schema=False)
    def ui_logo_asset():
        logo_path = resolve_videojet_logo_path()
        if not logo_path:
            raise HTTPException(status_code=404, detail="Logo asset missing")
        return FileResponse(logo_path, media_type="image/jpeg")

    # -----------------------------
    # Home
    # -----------------------------
    @app.get("/docs/swagger", include_in_schema=False)
    def docs_swagger():
        return get_swagger_ui_html(openapi_url=app.openapi_url, title=f"{app.title} - Swagger")

    @app.get("/docs", response_class=HTMLResponse, include_in_schema=False)
    def docs_page():
        nav = nav_html("docs")
        return f"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>API Docs</title>
  <style>
    body{{margin:0; font-family:Segoe UI,Arial,sans-serif; background:#f4f6f9; color:#1f2933}}
    .wrap{{max-width:1500px; margin:0 auto; padding:16px}}
    .topnav{{display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px}}
    .navbtn{{padding:8px 12px; border:1px solid #d6dde7; border-radius:8px; background:#fff; color:#1f2933; text-decoration:none}}
    .navbtn.active{{background:#005eb8; color:#fff; border-color:#005eb8}}
    .card{{background:#fff; border:1px solid #d6dde7; border-radius:10px; overflow:hidden}}
    .card h2{{margin:0; padding:12px 14px; border-bottom:1px solid #d6dde7}}
    iframe{{width:100%; height:calc(100vh - 170px); border:0}}
  </style>
</head>
<body>
  <div class="wrap">
    {nav}
    <div class="card">
      <h2>API Documentation</h2>
      <iframe src="/docs/swagger"></iframe>
    </div>
  </div>
</body>
</html>
"""

    @app.get("/", response_class=HTMLResponse)
    def home():
        cfg2 = Settings.load(cfg_path)
        nav = nav_html("home")
        
        def build_home_log_panel(channel: str, title: str, limit: int = 180) -> str:
            items = logs.list_logs(channel, limit=limit)
            lines = []
            for it in items:
                ts = float(it.get("ts") or 0.0)
                direction = str(it.get("direction") or "").upper()
                msg = str(it.get("message") or "")
                if channel == "all":
                    src = str(it.get("channel") or "")
                    lines.append(f"[{format_local_timestamp(ts, include_ms=False)}] [{src}] {direction} {msg}")
                else:
                    lines.append(f"[{format_local_timestamp(ts, include_ms=False)}] {direction} {msg}")
            body = "\n".join(lines) if lines else "(keine Eintraege)"
            return (
                '<section class="log-card">'
                f"<h3>{html.escape(title)}</h3>"
                f"<pre>{html.escape(body)}</pre>"
                "</section>"
            )

        home_log_panels = "".join(
            [
                build_home_log_panel("all", "All Channels"),
                build_home_log_panel("raspi", "Raspi"),
                build_home_log_panel("esp-plc", "ESP32-PLC"),
                build_home_log_panel("vj6530", "VJ6530 (TTO)"),
                build_home_log_panel("vj3350", "VJ3350 (Laser)"),
            ]
        )
        return f"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>MAS-004 Home</title>
  <style>
    body{{margin:0; font-family:Segoe UI,Arial,sans-serif; background:#f4f6f9; color:#1f2933}}
    .wrap{{max-width:1500px; margin:0 auto; padding:16px}}
    .topnav{{display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px}}
    .navbtn{{padding:8px 12px; border:1px solid #d6dde7; border-radius:8px; background:#fff; color:#1f2933; text-decoration:none}}
    .navbtn.active{{background:#005eb8; color:#fff; border-color:#005eb8}}
    .card{{background:#fff; border:1px solid #d6dde7; border-radius:10px; padding:14px}}
    .grid{{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:10px}}
    .logs-grid{{display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:10px; margin-top:10px}}
    .log-card{{border:1px solid #d6dde7; border-radius:10px; background:#fbfdff; padding:10px}}
    .log-card h3{{margin:0 0 8px 0; font-size:15px}}
    .log-card pre{{margin:0; background:#f7faff; border:1px solid #d6dde7; border-radius:8px; padding:8px; max-height:280px; overflow:auto; white-space:pre-wrap; word-break:break-word; font-size:12px; line-height:1.35; font-family:Consolas, "Courier New", monospace}}
    @media(max-width:900px){{.grid{{grid-template-columns:1fr;}}}}
    @media(max-width:1100px){{.logs-grid{{grid-template-columns:1fr;}}}}
  </style>
</head>
<body>
  <div class="wrap">
    {nav}
    <div class="card">
      <h2>MAS-004_RPI-Databridge</h2>
      <div class="grid">
        <div><b>eth0</b>: {cfg2.eth0_ip}</div>
        <div><b>eth1</b>: {cfg2.eth1_ip}</div>
        <div><b>Outbox</b>: <span id="home_outbox">{outbox.count()}</span></div>
        <div><b>Inbox pending</b>: <span id="home_inbox">{inbox.count_pending()}</span></div>
        <div><b>Peer</b>: {cfg2.peer_base_url}</div>
        <div><b>Peer (parallel)</b>: {(cfg2.peer_base_url_secondary or "-")}</div>
        <div><b>Watchdog host</b>: {cfg2.peer_watchdog_host}</div>
      </div>
    </div>
    <div class="card" style="margin-top:12px;">
      <h2>Logs (Read-only)</h2>
      <div class="logs-grid">
        {home_log_panels}
      </div>
    </div>
  </div>
  <script>
    const HOME_REFRESH_MS = 2000;
    let homeLiveTimer = null;

    async function refreshHomeCounters() {{
      try {{
        const r = await fetch("/api/ui/status/public");
        if (!r.ok) return;
        const j = await r.json();
        const outboxNode = document.getElementById("home_outbox");
        const inboxNode = document.getElementById("home_inbox");
        if (outboxNode) outboxNode.textContent = String(j.outbox_count ?? "-");
        if (inboxNode) inboxNode.textContent = String(j.inbox_pending ?? "-");
      }} catch (_e) {{
      }}
    }}

    function startHomeLiveCounters() {{
      if (homeLiveTimer) return;
      homeLiveTimer = setInterval(() => {{
        if (document.hidden) return;
        refreshHomeCounters();
      }}, HOME_REFRESH_MS);
    }}

    refreshHomeCounters();
    startHomeLiveCounters();
    document.addEventListener("visibilitychange", () => {{
      if (!document.hidden) refreshHomeCounters();
    }});
  </script>
</body>
</html>
        """

    def get_master_workbook_info() -> dict[str, Any]:
        cfg2 = Settings.load(cfg_path)
        path = cfg2.master_params_xlsx_path
        exists = os.path.exists(path)
        stat = os.stat(path) if exists else None
        return {
            "path": path,
            "exists": exists,
            "size_bytes": int(stat.st_size) if stat else 0,
            "mtime_iso": local_from_timestamp(stat.st_mtime).isoformat(timespec="seconds") if stat else None,
        }

    def get_io_workbook_info() -> dict[str, Any]:
        cfg2 = Settings.load(cfg_path)
        path = cfg2.master_ios_xlsx_path
        exists = os.path.exists(path)
        stat = os.stat(path) if exists else None
        return {
            "path": path,
            "exists": exists,
            "size_bytes": int(stat.st_size) if stat else 0,
            "mtime_iso": local_from_timestamp(stat.st_mtime).isoformat(timespec="seconds") if stat else None,
            "meta": io_store.master_info(),
        }

    def io_device_configured_simulation(cfg_local: Settings, device_code: str) -> bool:
        code = str(device_code or "").strip().lower()
        if code == "esp32_plc58":
            return bool(getattr(cfg_local, "esp_simulation", True))
        if code == "raspi_plc21":
            return bool(getattr(cfg_local, "raspi_io_simulation", True))
        if code in {"moxa_e1211_1", "moxa_e1213_1"}:
            return bool(getattr(cfg_local, "moxa1_simulation", True))
        if code in {"moxa_e1211_2", "moxa_e1213_2"}:
            return bool(getattr(cfg_local, "moxa2_simulation", True))
        if code == "moxa_e1213_3":
            return bool(getattr(cfg_local, "moxa3_simulation", True))
        return False

    def get_io_snapshot(*, live: bool = False) -> dict[str, Any]:
        cfg2 = Settings.load(cfg_path)
        runtime = IoRuntime(cfg2, io_store)
        if live:
            payload = runtime.refresh()
            device_status = {item["device_code"]: item for item in (payload.get("devices") or [])}
        else:
            points = io_store.list_points(include_reserved=True)
            payload = {
                "ok": True,
                "changed": 0,
                "devices": [],
                "points": points,
                "snapshot_only": True,
            }
            device_status = {}
            grouped_points: dict[str, list[dict[str, Any]]] = {}
            for point in points:
                grouped_points.setdefault(str(point.get("device_code") or ""), []).append(point)
            for device_code, device_points in grouped_points.items():
                qualities = {str(point.get("quality") or "unknown").lower() for point in device_points}
                host, port = runtime._device_address(device_code)
                simulation = io_device_configured_simulation(cfg2, device_code)
                reachable = (not simulation) and bool(qualities & {"live", "override"})
                device_status[device_code] = {
                    "device_code": device_code,
                    "host": host,
                    "port": port,
                    "simulation": simulation,
                    "reachable": reachable,
                    "error": "" if simulation or reachable else "kein aktueller Live-Snapshot",
                }
            payload["devices"] = list(device_status.values())
        catalog = io_store.list_devices()
        for item in catalog:
            status = device_status.get(item["device_code"], {})
            item["host"] = status.get("host", "")
            item["port"] = status.get("port", 0)
            item["simulation"] = bool(status.get("simulation", False))
            item["reachable"] = bool(status.get("reachable", False))
            item["error"] = status.get("error", "")
        payload["device_catalog"] = catalog
        payload["master_workbook"] = get_io_workbook_info()
        return payload

    def get_machine_overview() -> dict[str, Any]:
        cfg2 = Settings.load(cfg_path)
        runtime = MachineRuntime(cfg2, db, params, io_store, logs, outbox)
        return runtime.snapshot()

    def get_param_value_map(prefixes: tuple[str, ...] = ("MAP", "MAS", "MAE", "MAW")) -> dict[str, str]:
        placeholders = ",".join("?" for _ in prefixes)
        with db._conn() as c:
            rows = c.execute(
                f"""SELECT p.pkey, COALESCE(v.value, p.default_v, '0')
                    FROM params p
                    LEFT JOIN param_values v ON v.pkey = p.pkey
                    WHERE p.ptype IN ({placeholders})""",
                tuple(prefixes),
            ).fetchall()
        return {str(row[0]): str(row[1] if row[1] is not None else "0") for row in rows}

    def safe_param_float(values: dict[str, str], key: str, default: float = 0.0) -> float:
        try:
            return float(str(values.get(key, default)).strip())
        except Exception:
            return float(default)

    def safe_param_int(values: dict[str, str], key: str, default: int = 0) -> int:
        try:
            return int(float(str(values.get(key, default)).strip()))
        except Exception:
            return int(default)

    def update_param_workbook_defaults(cfg2: Settings, pkey: str, value: str) -> list[str]:
        updated: list[str] = []
        ptype = str(pkey or "")[:3].upper()
        pid = normalize_pid(ptype, str(pkey or "")[3:])
        paths: list[str] = []
        for candidate in (
            getattr(cfg2, "master_params_xlsx_path", ""),
            REPO_MASTER_PARAMS_XLSX,
            WORKSPACE_MASTER_PARAMS_XLSX,
            FALLBACK_REPO_MASTER_PARAMS_XLSX,
        ):
            if candidate and candidate not in paths:
                paths.append(candidate)
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception:
            return updated
        for path in paths:
            if not path or not os.path.exists(path):
                continue
            try:
                wb = load_workbook(path)
                changed = False
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        if len(row) < 5:
                            continue
                        row_ptype = str(row[0].value or "").strip().upper()
                        row_pid = str(row[1].value or "").strip().zfill(4)
                        if row_ptype == ptype and row_pid == pid:
                            row[4].value = str(value)
                            changed = True
                            break
                    if changed:
                        break
                if changed:
                    wb.save(path)
                    updated.append(path)
            except Exception as exc:
                try:
                    logs.log(
                        "machine",
                        "WARN",
                        f"Master-Excel Default fuer {pkey} konnte nicht aktualisiert werden: {path}: {exc}",
                    )
                except Exception:
                    pass
        return updated

    def set_visualization_component_position(cfg2: Settings, body: ProductionVisualizationComponentReq) -> dict[str, Any]:
        key = str(body.key or "").strip().lower()
        try:
            mm = float(body.mm)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Ungueltiger mm-Wert") from exc
        if not math.isfinite(mm):
            raise HTTPException(status_code=400, detail="Ungueltiger mm-Wert")
        values = get_param_value_map()
        use_laser = str(values.get("MAP0016", "0")).strip() not in ("", "0", "false", "False", "FALSE")
        print_key = "MAP0018" if use_laser else "MAP0019"
        recipes: dict[str, dict[str, Any]] = {
            "material": {"param": "MAP0017", "offset_params": ("MAP0011",), "label": "Materialkamera"},
            "print": {"param": print_key, "offset_params": ("MAP0004", "MAP0006"), "label": "Druck"},
            "verify": {"param": "MAP0020", "offset_params": ("MAP0012",), "label": "Verifikation"},
            "control": {"param": "MAP0021", "offset_params": (), "label": "Entnahmesensor"},
            "led_start": {"param": "MAP0066", "offset_params": (), "label": "LED 1"},
            "led_end": {"param": "MAP0071", "offset_params": ("MAP0066",), "label": "LED Ende"},
            "exit": {"param": "MAP0021", "offset_mm": 50.0, "label": "Bahn verlassen"},
        }
        recipe = recipes.get(key)
        if not recipe:
            raise HTTPException(status_code=400, detail=f"Marker {key!r} ist nicht editierbar")
        offset_tenths = 0
        for offset_key in recipe.get("offset_params", ()):
            sign = -1 if str(offset_key).startswith("-") else 1
            actual_key = str(offset_key).lstrip("-")
            offset_tenths += sign * safe_param_int(values, actual_key, 0)
        if "offset_mm" in recipe:
            offset_tenths += int(round(float(recipe.get("offset_mm") or 0.0) * 10.0))
        value_tenths = int(round(mm * 10.0)) - offset_tenths
        param_key = str(recipe["param"])
        meta = params.get_meta(param_key)
        if not meta:
            raise HTTPException(status_code=400, detail=f"{param_key} ist unbekannt")
        min_v = meta.get("min_v")
        max_v = meta.get("max_v")
        if min_v is not None and value_tenths < int(float(min_v)):
            raise HTTPException(status_code=400, detail=f"{param_key} waere unter Minimum {min_v}")
        if max_v is not None and value_tenths > int(float(max_v)):
            raise HTTPException(status_code=400, detail=f"{param_key} waere ueber Maximum {max_v}")
        ok, msg = params.set_value(param_key, str(value_tenths), actor="machine-setup-ui")
        if not ok:
            raise HTTPException(status_code=400, detail=msg)
        workbook_paths = update_param_workbook_defaults(cfg2, param_key, str(value_tenths))
        try:
            logs.log(
                "machine",
                "INFO",
                (
                    f"Visualisierung Marker {recipe.get('label')} gespeichert: "
                    f"{mm:.1f} mm -> {param_key}={value_tenths}"
                ),
            )
        except Exception:
            pass
        payload = get_production_visualization_payload()
        payload["saved"] = {
            "key": key,
            "label": recipe.get("label"),
            "mm": mm,
            "pkey": param_key,
            "value": str(value_tenths),
            "workbooks": workbook_paths,
        }
        return payload

    def parse_esp_json_reply(reply: str, *, command: str = "") -> dict[str, Any]:
        text = str(reply or "").strip()
        if text.upper().startswith("JSON "):
            text = text[5:].strip()
        if text and not text.endswith(("}", "]")):
            detail = f"truncated/non-complete JSON reply ({len(text)} chars)"
            if command:
                detail = f"{command}: {detail}"
            detail += f", tail={text[-80:]!r}"
            raise ValueError(detail)
        try:
            return json.loads(text or "{}")
        except json.JSONDecodeError as exc:
            detail = f"invalid ESP JSON ({len(text)} chars): {exc}"
            if command:
                detail = f"{command}: {detail}"
            detail += f", tail={text[-80:]!r}"
            raise ValueError(detail) from exc

    def read_esp_json(client: EspPlcClient, command: str, *, read_timeout_s: float = 2.0, read_limit: int = 4096) -> dict[str, Any]:
        reply = client.exchange_line(command, read_timeout_s=read_timeout_s, read_limit=read_limit)
        stripped = str(reply or "").strip()
        if stripped.upper().startswith("NAK"):
            raise RuntimeError(f"{command}: {stripped}")
        return parse_esp_json_reply(reply, command=command)

    def read_compact_esp_visualization(client: EspPlcClient, *, allow_fallback: bool = True) -> tuple[dict[str, Any], list[str]]:
        visualization_error = ""
        try:
            snapshot = read_esp_json(client, "PROCESS VISUALIZATION?", read_timeout_s=2.0, read_limit=65536)
            if isinstance(snapshot, dict) and snapshot:
                return snapshot, []
        except Exception as exc:
            visualization_error = str(exc)

        snapshot: dict[str, Any] = {
            "ok": True,
            "labels": [],
        }
        errors: list[str] = []
        if not allow_fallback:
            if visualization_error:
                snapshot["ok"] = False
                snapshot["fallback_reason"] = visualization_error
                errors.append(visualization_error)
            return snapshot, errors
        success_count = 0
        for key, command in (
            ("production", "PROCESS PRODUCTION STATUS?"),
            ("setup_measure", "PROCESS SETUP_MEASURE STATUS?"),
            ("led_test", "PROCESS LED_TEST STATUS?"),
        ):
            try:
                snapshot[key] = read_esp_json(client, command, read_timeout_s=2.0, read_limit=8192)
                success_count += 1
            except Exception as exc:
                errors.append(str(exc))
        production = snapshot.get("production") if isinstance(snapshot.get("production"), dict) else {}
        setup_measure = snapshot.get("setup_measure") if isinstance(snapshot.get("setup_measure"), dict) else {}
        snapshot["infeed_mm"] = setup_measure.get("infeed_mm", production.get("label_acquire_mm", 0.0))
        snapshot["drive_mm"] = setup_measure.get("drive_mm", 0.0)
        snapshot["infeed_speed_mm_s"] = setup_measure.get("command_speed_mm_s", production.get("speed_mm_s", 0.0))
        snapshot["drive_speed_mm_s"] = snapshot.get("infeed_speed_mm_s", 0.0)
        snapshot["faults"] = {
            "label_short": False,
            "label_long": False,
            "sensor_fault": False,
            "band_break": False,
        }
        snapshot["compact_fallback"] = True
        if visualization_error:
            snapshot["fallback_reason"] = visualization_error
        if success_count > 0:
            return snapshot, errors
        if visualization_error:
            errors.insert(0, visualization_error)
        return snapshot, errors

    def recent_completed_labels(limit: int = 80) -> list[dict[str, Any]]:
        with db._conn() as c:
            rows = c.execute(
                """SELECT production_label,label_no,created_ts,completed_ts,zero_mm,exit_mm,
                          material_ok,print_ok,verify_ok,removed,production_ok,payload_json
                   FROM label_register
                   ORDER BY COALESCE(completed_ts, created_ts) DESC, label_no DESC
                   LIMIT ?""",
                (int(limit),),
            ).fetchall()
        out: list[dict[str, Any]] = []
        for row in rows:
            try:
                payload = json.loads(row[11] or "{}")
            except Exception:
                payload = {}
            out.append(
                {
                    "production_label": row[0],
                    "label_no": int(row[1] or 0),
                    "created_ts": float(row[2] or 0.0),
                    "completed_ts": float(row[3] or 0.0),
                    "zero_mm": float(row[4] or 0.0),
                    "exit_mm": float(row[5] or 0.0),
                    "material_ok": bool(row[6]),
                    "print_ok": bool(row[7]),
                    "verify_ok": bool(row[8]),
                    "removed": bool(row[9]),
                    "production_ok": bool(row[10]),
                    "payload": payload,
                }
            )
        return out

    def build_production_track(values: dict[str, str]) -> dict[str, Any]:
        material_mm = (safe_param_float(values, "MAP0017", 2200.0) + safe_param_float(values, "MAP0011", 0.0)) / 10.0
        use_laser = str(values.get("MAP0016", "0")).strip() not in ("", "0", "false", "False", "FALSE")
        print_key = "MAP0018" if use_laser else "MAP0019"
        print_mm = (
            safe_param_float(values, print_key, 7000.0 if use_laser else 11000.0)
            + safe_param_float(values, "MAP0004", 100.0)
            + safe_param_float(values, "MAP0006", 0.0)
        ) / 10.0
        verify_mm = (safe_param_float(values, "MAP0020", 12700.0) + safe_param_float(values, "MAP0012", 0.0)) / 10.0
        control_mm = safe_param_float(values, "MAP0021", 19300.0) / 10.0
        exit_mm = control_mm + 50.0
        led_offset_mm = safe_param_float(values, "MAP0066", 8000.0) / 10.0
        led_length_mm = max(1.0, min(520.0, safe_param_float(values, "MAP0071", 5200.0) / 10.0))
        components = [
            {"key": "detect", "kind": "detect", "label": "Erfassung", "mm": 0.0, "editable": False},
            {"key": "material", "kind": "material", "label": "Materialkamera", "mm": material_mm, "editable": True, "param": "MAP0017"},
            {"key": "print", "kind": "print", "label": "Druck", "mm": print_mm, "editable": True, "param": print_key},
            {"key": "verify", "kind": "verify", "label": "Verifikation", "mm": verify_mm, "editable": True, "param": "MAP0020"},
            {"key": "control", "kind": "control", "label": "Entnahmesensor", "mm": control_mm, "editable": True, "param": "MAP0021"},
            {"key": "exit", "kind": "exit", "label": "Bahn verlassen", "mm": exit_mm, "editable": True, "param": "MAP0021"},
            {"key": "led_start", "kind": "detect", "label": "LED 1", "mm": led_offset_mm, "editable": True, "param": "MAP0066"},
            {"key": "led_end", "kind": "detect", "label": "LED Ende", "mm": led_offset_mm + led_length_mm, "editable": True, "param": "MAP0071"},
        ]
        length_mm = max(item["mm"] for item in components) + 120.0
        return {
            "length_mm": length_mm,
            "components": sorted(components, key=lambda item: float(item["mm"])),
        }

    def get_production_visualization_payload() -> dict[str, Any]:
        cfg2 = Settings.load(cfg_path)
        machine = get_machine_overview()
        values = get_param_value_map()
        format_plan = build_format_plan(values)
        label = format_plan.get("label") or {}
        process = format_plan.get("process") or {}
        esp_snapshot: dict[str, Any] = {}
        esp_errors: list[str] = []
        machine_info = dict((machine or {}).get("info") or {})
        production_info = dict(machine_info.get("production_runtime") or {})
        machine_state = int((machine or {}).get("current_state") or 0)
        production_window = machine_state == 5
        critical_motion_window = machine_state in (2, 3, 4, 6) or bool(production_info.get("pending_start"))
        if critical_motion_window:
            esp_errors.append("ESP-Liveabfrage waehrend Einrichten/Uebergang zur Kanalentlastung pausiert")
        elif bool(getattr(cfg2, "esp_simulation", False)) or not str(getattr(cfg2, "esp_host", "") or "").strip():
            esp_errors.append("ESP-Simulation aktiv" if bool(getattr(cfg2, "esp_simulation", False)) else "ESP-Endpunkt fehlt")
        else:
            try:
                timeout_s = float(getattr(cfg2, "http_timeout_s", 1.5) or 1.5)
                client = EspPlcClient(cfg2.esp_host, int(cfg2.esp_port), timeout_s=timeout_s)
                esp_snapshot, esp_errors = read_compact_esp_visualization(client, allow_fallback=not production_window)
            except Exception as exc:
                esp_errors.append(repr(exc))
        active_labels = list(esp_snapshot.get("labels") or [])
        esp_error = "; ".join(str(item) for item in esp_errors if str(item).strip())
        led_offset_mm = safe_param_float(values, "MAP0066", 8000.0) / 10.0
        led_length_mm = max(1.0, min(520.0, safe_param_float(values, "MAP0071", 5200.0) / 10.0))
        led_pitch_mm = 6.95
        led_count = max(1, min(75, int(math.ceil(led_length_mm / led_pitch_mm))))
        led_physical_length_mm = led_count * led_pitch_mm
        led_controller_last_octet = safe_param_int(values, "MAP0073", 255)
        led_controller_port = safe_param_int(values, "MAP0074", 3050)
        led_controller_enabled = safe_param_int(values, "MAP0072", 1) != 0
        return {
            "ok": not bool(esp_error and not esp_snapshot),
            "ts": time.time(),
            "machine": machine,
            "esp_snapshot": esp_snapshot,
            "esp_error": esp_error,
            "active_labels": active_labels,
            "completed_labels": recent_completed_labels(limit=80),
            "format": {
                "label_length_mm": float(label.get("length_tenths_mm") or 0) / 10.0,
                "label_tolerance_mm": float(label.get("length_tolerance_tenths_mm") or 0) / 10.0,
                "transport_speed_mm_s": process.get("transport_speed_mm_s"),
                "active_printer": (format_plan.get("printer") or {}).get("active"),
                "print_distance_mm": float((format_plan.get("printer") or {}).get("stop_distance_tenths_mm") or 0) / 10.0,
            },
            "track": build_production_track(values),
            "led": {
                "pin": None,
                "pin_label": "Olimex ESP32-POE-ISO-IND LED-Controller",
                "documented_io_pin": "UDP",
                "documented_io_pin_id": "MAS004-LED-UDP/v1",
                "documented_io_pin_label": "ESP32-PLC Schieberegister -> UDP -> Olimex WS2812-Controller",
                "fastled_requires_raw_gpio": False,
                "documented_io_pin_fastled_capable": False,
                "output": "external_udp",
                "protocol": "MAS004-LED-UDP/v1",
                "controller_source": "ESP32-PLC Prozess-Schieberegister",
                "controller_target": f"192.168.2.{led_controller_last_octet}:{led_controller_port}",
                "controller_broadcast": led_controller_last_octet == 255,
                "controller_enabled": led_controller_enabled,
                "frame_interval_ms": safe_param_int(values, "MAP0075", 100),
                "warning": (
                    "Die PLC58 treibt keinen WS2812-Pin mehr. Sie sendet fertige RGB-Frames "
                    "direkt aus dem ESP-Schieberegister per UDP an den Olimex ESP32-POE-ISO-IND LED-Controller."
                ),
                "count": led_count,
                "max_count": 75,
                "pitch_mm": led_pitch_mm,
                "offset_mm": led_offset_mm,
                "length_mm": led_length_mm,
                "physical_length_mm": led_physical_length_mm,
                "first_led_distance_tenths_mm": safe_param_int(values, "MAP0066", 8000),
                "strip_length_tenths_mm": safe_param_int(values, "MAP0071", 5200),
            },
        }

    def execute_led_test_command(cfg2: Settings, body: MachineLedTestReq) -> dict[str, Any]:
        action = (body.action or "start").strip().lower()
        if action == "stop":
            command = "PROCESS LED_TEST STOP"
        elif action == "status":
            command = "PROCESS LED_TEST STATUS?"
        elif action in ("high", "static_high", "tx1_high"):
            duration_ms = int(body.duration_ms or 0)
            if duration_ms <= 0:
                command = "PROCESS LED_TEST HIGH DURATION_MS=0"
            else:
                duration_ms = max(250, min(10000, duration_ms))
                command = f"PROCESS LED_TEST HIGH DURATION_MS={duration_ms}"
        elif action in ("", "start", "red"):
            duration_ms = int(body.duration_ms or 0)
            if duration_ms <= 0:
                command = "PROCESS LED_TEST RED DURATION_MS=0"
            else:
                duration_ms = max(250, min(10000, duration_ms))
                command = f"PROCESS LED_TEST RED DURATION_MS={duration_ms}"
        else:
            raise HTTPException(status_code=400, detail=f"Unknown LED test action {body.action!r}")

        if bool(getattr(cfg2, "esp_simulation", False)) or not str(getattr(cfg2, "esp_host", "") or "").strip():
            return {
                "ok": True,
                "simulation": True,
                "command": command,
                "reply": "SIM_PROCESS_LED_TEST",
            }

        try:
            timeout_s = float(getattr(cfg2, "http_timeout_s", 1.5) or 1.5)
            client = EspPlcClient(cfg2.esp_host, int(cfg2.esp_port), timeout_s=timeout_s)
            reply = client.exchange_line(command, read_timeout_s=2.0, read_limit=4096)
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"ESP LED test failed: {exc!r}") from exc

        payload: dict[str, Any] = {"ok": True, "command": command, "reply": reply}
        if str(reply or "").strip().upper().startswith("NAK"):
            raise HTTPException(status_code=409, detail=payload)
        if str(reply or "").strip().upper().startswith("JSON "):
            try:
                payload["status"] = parse_esp_json_reply(reply)
            except Exception:
                pass
        return payload

    def ensure_machine_bypass_params() -> None:
        ts = time.time()
        with db._conn() as c:
            for spec in MACHINE_BYPASS_PARAM_DEFS:
                exists = c.execute("SELECT 1 FROM params WHERE pkey=?", (spec["pkey"],)).fetchone()
                if not exists:
                    c.execute(
                        """INSERT INTO params(
                           pkey,ptype,pid,min_v,max_v,default_v,unit,rw,esp_rw,dtype,name,format_relevant,
                           message,possible_cause,effects,remedy,ai_instructions,updated_ts
                           )
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            spec["pkey"],
                            spec["ptype"],
                            spec["pid"],
                            spec.get("min_v"),
                            spec.get("max_v"),
                            spec.get("default_v"),
                            spec.get("unit"),
                            spec.get("rw"),
                            spec.get("esp_rw"),
                            spec.get("dtype"),
                            spec.get("name"),
                            "NO",
                            spec.get("message"),
                            "",
                            "",
                            "",
                            (
                                "KI: Ich verstehe diesen Parameter als produktionsrelevante "
                                "Bypass-/Simulationsvorgabe. Er wird ueber die Prozessseite "
                                "oder Microtom gesetzt und zur ESP32-PLC gespiegelt."
                            ),
                            ts,
                        ),
                    )
                c.execute(
                    "UPDATE params SET rw=?, esp_rw=?, format_relevant=?, name=?, message=?, updated_ts=? WHERE pkey=?",
                    (
                        spec.get("rw"),
                        spec.get("esp_rw"),
                        "NO",
                        spec.get("name"),
                        spec.get("message"),
                        ts,
                        spec["pkey"],
                    ),
                )

    def ensure_machine_led_params() -> None:
        ts = time.time()
        with db._conn() as c:
            for spec in MACHINE_LED_PARAM_DEFS:
                exists = c.execute("SELECT 1 FROM params WHERE pkey=?", (spec["pkey"],)).fetchone()
                if not exists:
                    c.execute(
                        """INSERT INTO params(
                           pkey,ptype,pid,min_v,max_v,default_v,unit,rw,esp_rw,dtype,name,format_relevant,
                           message,possible_cause,effects,remedy,ai_instructions,updated_ts
                           )
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            spec["pkey"],
                            spec["ptype"],
                            spec["pid"],
                            spec.get("min_v"),
                            spec.get("max_v"),
                            spec.get("default_v"),
                            spec.get("unit"),
                            spec.get("rw"),
                            spec.get("esp_rw"),
                            spec.get("dtype"),
                            spec.get("name"),
                            spec.get("format_relevant", "NO"),
                            spec.get("message"),
                            "",
                            "",
                            "",
                            (
                                "KI: Ich verstehe diesen Parameter als LED-Controller-Vorgabe. "
                                "Der ESP32-PLC Prozess rendert daraus fertige UDP-Frames fuer "
                                "einen externen WS2812-Controller."
                            ),
                            ts,
                        ),
                    )
                c.execute(
                    """UPDATE params
                       SET rw=?, esp_rw=?, format_relevant=?, name=?, message=?, min_v=?, max_v=?, default_v=?, updated_ts=?
                       WHERE pkey=?""",
                    (
                        spec.get("rw"),
                        spec.get("esp_rw"),
                        spec.get("format_relevant", "NO"),
                        spec.get("name"),
                        spec.get("message"),
                        spec.get("min_v"),
                        spec.get("max_v"),
                        spec.get("default_v"),
                        ts,
                        spec["pkey"],
                    ),
                )

    def ensure_machine_process_params() -> None:
        ts = time.time()
        with db._conn() as c:
            for spec in MACHINE_PROCESS_PARAM_DEFS:
                exists = c.execute("SELECT 1 FROM params WHERE pkey=?", (spec["pkey"],)).fetchone()
                if not exists:
                    c.execute(
                        """INSERT INTO params(
                           pkey,ptype,pid,min_v,max_v,default_v,unit,rw,esp_rw,dtype,name,format_relevant,
                           message,possible_cause,effects,remedy,ai_instructions,updated_ts
                           )
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            spec["pkey"],
                            spec["ptype"],
                            spec["pid"],
                            spec.get("min_v"),
                            spec.get("max_v"),
                            spec.get("default_v"),
                            spec.get("unit"),
                            spec.get("rw"),
                            spec.get("esp_rw"),
                            spec.get("dtype"),
                            spec.get("name"),
                            spec.get("format_relevant", "NO"),
                            spec.get("message"),
                            "",
                            "",
                            "",
                            (
                                "KI: Ich verstehe diesen Parameter als Prozess-/Sensor-Kalibrierwert. "
                                "Er wird vom Raspi zur ESP32-PLC gespiegelt und beeinflusst die "
                                "bewertete Etikettenlaenge, nicht die Encoder-Wegskalierung."
                            ),
                            ts,
                        ),
                    )
                c.execute(
                    """UPDATE params
                       SET rw=?, esp_rw=?, format_relevant=?, name=?, message=?, min_v=?, max_v=?, default_v=?, updated_ts=?
                       WHERE pkey=?""",
                    (
                        spec.get("rw"),
                        spec.get("esp_rw"),
                        spec.get("format_relevant", "NO"),
                        spec.get("name"),
                        spec.get("message"),
                        spec.get("min_v"),
                        spec.get("max_v"),
                        spec.get("default_v"),
                        ts,
                        spec["pkey"],
                    ),
                )

    ensure_machine_bypass_params()
    ensure_machine_led_params()
    ensure_machine_process_params()

    def get_machine_bypass_payload() -> dict[str, Any]:
        ensure_machine_bypass_params()
        items = []
        for spec in MACHINE_BYPASS_PARAM_DEFS:
            pkey = str(spec["pkey"])
            meta = params.get_meta(pkey) or {}
            items.append(
                {
                    "pkey": pkey,
                    "value": params.get_effective_value(pkey),
                    "default_v": meta.get("default_v"),
                    "min_v": meta.get("min_v"),
                    "max_v": meta.get("max_v"),
                    "unit": meta.get("unit") or spec.get("unit") or "",
                    "dtype": meta.get("dtype") or spec.get("dtype") or "",
                    "name": meta.get("name") or spec.get("name") or pkey,
                    "message": meta.get("message") or spec.get("message") or "",
                    "can_write_microtom": params.can_actor_write(pkey, actor="microtom"),
                }
            )
        return {"ok": True, "parameters": items}

    def write_machine_bypass_values(values: Dict[str, Any]) -> dict[str, Any]:
        ensure_machine_bypass_params()
        allowed = {str(item["pkey"]) for item in MACHINE_BYPASS_PARAM_DEFS}
        router = Router(Settings.load(cfg_path), inbox, outbox, params, logs)
        results = []
        for raw_key, raw_value in sorted((values or {}).items()):
            pkey = str(raw_key or "").strip().upper()
            if pkey not in allowed:
                results.append({"pkey": pkey, "ok": False, "error": "NAK_NotBypassParam"})
                continue
            meta = params.get_meta(pkey) or {}
            if not params.can_actor_write(pkey, actor="microtom"):
                results.append({"pkey": pkey, "ok": False, "error": "NAK_ReadOnly"})
                continue
            try:
                value = int(float(str(raw_value).strip()))
            except Exception:
                results.append({"pkey": pkey, "ok": False, "error": "NAK_Syntax"})
                continue
            min_v = meta.get("min_v")
            max_v = meta.get("max_v")
            if min_v is not None and value < int(float(min_v)):
                results.append({"pkey": pkey, "ok": False, "error": "NAK_OutOfRange"})
                continue
            if max_v is not None and value > int(float(max_v)):
                results.append({"pkey": pkey, "ok": False, "error": "NAK_OutOfRange"})
                continue
            line = f"{pkey}={value}"
            response = router.handle_microtom_line(line, correlation=f"machine-bypass:{pkey}")
            results.append({"pkey": pkey, "line": line, "response": response, "ok": bool(response) and "NAK" not in str(response).upper()})
        return {"ok": all(bool(item.get("ok")) for item in results), "results": results, "bypass": get_machine_bypass_payload()}

    def motor3_calibration_script_path() -> str:
        repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        return os.path.join(repo_root, "scripts", "diagnose_motor3_travel_2000.py")

    def motor3_calibration_result_path() -> str:
        return "/tmp/mas004_motor3_travel_2000_result.json"

    def read_motor3_calibration_result() -> dict[str, Any]:
        path = motor3_calibration_result_path()
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as fh:
                payload = json.load(fh)
            return payload if isinstance(payload, dict) else {}
        except Exception as exc:
            return {"error": repr(exc), "path": path}

    def esp_json(cfg2: Settings, line: str, timeout_s: float = 5.0) -> dict[str, Any]:
        client = EspPlcClient(
            cfg2.esp_host,
            cfg2.esp_port,
            timeout_s=cfg2.get_float("esp_connect_timeout_s", 1.5),
        )
        text = client.exchange_line(line, read_timeout_s=timeout_s)
        raw = str(text or "").strip()
        if raw.upper().startswith("JSON "):
            raw = raw[5:].strip()
        try:
            return dict(json.loads(raw))
        except Exception:
            return {"raw": raw}

    def motor3_calibration_status(cfg2: Settings) -> dict[str, Any]:
        with motor3_calibration_guard:
            proc = motor3_calibration_process.get("proc")
            if proc is not None and proc.poll() is not None:
                motor3_calibration_process["returncode"] = proc.returncode
                motor3_calibration_process["proc"] = None
                proc = None
            running = proc is not None
            process_info = {
                "running": running,
                "mode": motor3_calibration_process.get("mode") or "",
                "started_ts": motor3_calibration_process.get("started_ts") or 0.0,
                "returncode": motor3_calibration_process.get("returncode"),
            }
        ensure_machine_process_params()
        result = read_motor3_calibration_result()
        live: dict[str, Any] = {}
        motor3: dict[str, Any] = {}
        try:
            live = esp_json(cfg2, "PROCESS TRAVEL_DIAG STATUS?", timeout_s=5.0)
        except Exception as exc:
            live = {"error": repr(exc)}
        try:
            motor3_payload = esp_json(cfg2, "MOTOR 3 REFRESH", timeout_s=5.0)
            motor3 = motor3_payload.get("motor") if isinstance(motor3_payload.get("motor"), dict) else motor3_payload
        except Exception as exc:
            motor3 = {"error": repr(exc)}
        return {
            "ok": True,
            **process_info,
            "script": motor3_calibration_script_path(),
            "result_path": motor3_calibration_result_path(),
            "result": result,
            "travel_diag": live,
            "motor3": motor3,
            "params": {
                key: params.get_effective_value(key)
                for key in ("MAP0076", "MAP0077", "MAP0078")
            },
        }

    def start_motor3_calibration_process(mode: str) -> dict[str, Any]:
        script = motor3_calibration_script_path()
        if not os.path.exists(script):
            raise RuntimeError(f"Kalibrier-Script nicht gefunden: {script}")
        if mode not in {"prepare", "run"}:
            raise RuntimeError(f"Ungueltiger Kalibriermodus: {mode}")
        with motor3_calibration_guard:
            proc = motor3_calibration_process.get("proc")
            if proc is not None and proc.poll() is None:
                raise RuntimeError(f"Kalibrierung laeuft bereits: {motor3_calibration_process.get('mode')}")
            python_exe = shutil.which("python3") or shutil.which("python") or "python3"
            proc = subprocess.Popen(
                [python_exe, script, mode],
                cwd=os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                start_new_session=True,
            )
            motor3_calibration_process.update(
                {"proc": proc, "mode": mode, "started_ts": time.time(), "returncode": None}
            )
            return {"ok": True, "mode": mode, "pid": proc.pid, "script": script}

    def _safe_float_value(value: Any, default: float = 0.0) -> float:
        try:
            return float(str(value).strip())
        except Exception:
            return float(default)

    def apply_motor3_calibration(cfg2: Settings, body: Motor3CalibrationApplyReq) -> dict[str, Any]:
        ensure_machine_process_params()
        result = read_motor3_calibration_result()
        final_snapshot = result.get("final_snapshot") if isinstance(result.get("final_snapshot"), dict) else {}
        final_motor = result.get("final_motor") if isinstance(result.get("final_motor"), dict) else {}
        start_motor = result.get("start_motor") if isinstance(result.get("start_motor"), dict) else {}
        actual_travel = float(body.actual_travel_mm)
        if not math.isfinite(actual_travel) or actual_travel < 100.0:
            raise RuntimeError("Real gefahrene Strecke ungueltig")
        target_mm = _safe_float_value(result.get("target_mm"), 2000.0)
        measured_infeed = _safe_float_value(final_snapshot.get("infeed_mm"), 0.0)
        measured_drive = _safe_float_value(final_snapshot.get("drive_mm"), 0.0)
        if abs(measured_infeed) < 100.0 or abs(measured_drive) < 100.0:
            raise RuntimeError(f"Keine nutzbare 2000-mm-Messung vorhanden: {final_snapshot}")
        current_infeed_um = _safe_float_value(params.get_effective_value("MAP0077"), 100765.0)
        current_drive_um = _safe_float_value(params.get_effective_value("MAP0078"), 100649.0)
        config = final_motor.get("config") if isinstance(final_motor.get("config"), dict) else {}
        start_config = start_motor.get("config") if isinstance(start_motor.get("config"), dict) else {}
        current_steps = _safe_float_value(
            config.get("steps_per_mm") or final_motor.get("steps_per_mm") or start_config.get("steps_per_mm"),
            31.627315,
        )
        new_infeed_um = int(round(current_infeed_um * (actual_travel / measured_infeed)))
        new_drive_um = int(round(current_drive_um * (actual_travel / measured_drive)))
        new_steps = current_steps * (target_mm / actual_travel)

        labels = result.get("labels_first_10") if isinstance(result.get("labels_first_10"), list) else []
        raw_lengths = [
            _safe_float_value(item.get("raw_length_mm"), 0.0)
            for item in labels
            if isinstance(item, dict) and _safe_float_value(item.get("raw_length_mm"), 0.0) > 1.0
        ]
        calculated: dict[str, Any] = {
            "motor3_steps_per_mm": round(new_steps, 6),
            "MAP0077": new_infeed_um,
            "MAP0078": new_drive_um,
            "measured_infeed_mm": measured_infeed,
            "measured_drive_mm": measured_drive,
            "actual_travel_mm": actual_travel,
        }
        actual_label_length = (
            _safe_float_value(body.actual_label_length_mm, 0.0)
            if body.actual_label_length_mm is not None
            else 0.0
        )
        if actual_label_length > 1.0 and raw_lengths:
            raw_avg = sum(raw_lengths) / len(raw_lengths)
            label_comp_tenths = int(round((actual_label_length - raw_avg) * 10.0))
            calculated["raw_label_avg_mm"] = round(raw_avg, 4)
            calculated["MAP0076"] = max(-50, min(50, label_comp_tenths))

        if abs(new_steps / current_steps - 1.0) > 0.05:
            raise RuntimeError(f"ID3-Korrektur >5% gesperrt: alt={current_steps}, neu={new_steps}")
        if abs(new_infeed_um / current_infeed_um - 1.0) > 0.05:
            raise RuntimeError(f"Einlaufencoder-Korrektur >5% gesperrt: alt={current_infeed_um}, neu={new_infeed_um}")
        if abs(new_drive_um / current_drive_um - 1.0) > 0.05:
            raise RuntimeError(f"Auslaufencoder-Korrektur >5% gesperrt: alt={current_drive_um}, neu={new_drive_um}")

        router = Router(cfg2, inbox, outbox, params, logs)
        esp = EspPlcClient(
            cfg2.esp_host,
            cfg2.esp_port,
            timeout_s=cfg2.get_float("esp_connect_timeout_s", 1.5),
        )
        writes: list[dict[str, Any]] = []
        motor_set = esp.exchange_line(f"MOTOR 3 SET steps_per_mm={new_steps:.6f}", read_timeout_s=5.0)
        motor_save = esp.exchange_line("MOTOR 3 SAVE", read_timeout_s=5.0)
        writes.append({"target": "MOTOR3", "set": motor_set, "save": motor_save})
        for key in ("MAP0077", "MAP0078", "MAP0076"):
            if key not in calculated:
                continue
            value = str(int(calculated[key]))
            response = router.handle_microtom_line(f"{key}={value}", correlation=f"motor3-calibration:{key}")
            sync_response = esp.exchange_line(f"SYNC {key}={value}", read_timeout_s=5.0)
            writes.append({"pkey": key, "value": value, "response": response, "sync": sync_response})
        return {"ok": True, "calculated": calculated, "writes": writes, "result": result}

    def is_format_relevant(value: Any) -> bool:
        text = str(value or "").strip().lower()
        return text not in ("", "0", "false", "no", "nein", "n", "none", "null", "-")

    def production_status_keys() -> list[str]:
        return [
            "MAS0001",
            "MAS0002",
            "MAS0003",
            "MAS0008",
            "MAS0009",
            "MAS0026",
            "MAS0027",
            "MAS0028",
            "MAS0029",
            "MAS0030",
        ]

    def get_production_setup_payload() -> dict[str, Any]:
        rows = params.list_params(limit=100000, offset=0)
        format_rows = []
        for row in rows:
            meta = params.get_meta(str(row.get("pkey") or ""))
            if not meta or not is_format_relevant(meta.get("format_relevant")):
                continue
            item = {
                **row,
                "format_relevant": meta.get("format_relevant"),
                "value": row.get("effective_v"),
                "can_write_microtom": params.can_actor_write(str(row.get("pkey") or ""), actor="microtom"),
                "can_read_microtom": params.can_actor_read(str(row.get("pkey") or ""), actor="microtom"),
            }
            format_rows.append(item)

        status = []
        for pkey in production_status_keys():
            meta = params.get_meta(pkey)
            if not meta:
                continue
            status.append(
                {
                    "pkey": pkey,
                    "value": params.get_effective_value(pkey),
                    "name": meta.get("name") or "",
                    "message": meta.get("message") or "",
                    "unit": meta.get("unit") or "",
                }
            )
        return {
            "ok": True,
            "parameters": format_rows,
            "production_status": status,
            "machine": get_machine_overview(),
        }

    def send_format_values(values: dict[str, Any], name: str = "") -> dict[str, Any]:
        cfg2 = Settings.load(cfg_path)
        router = Router(cfg2, inbox, outbox, params, logs)
        clean_values = FormatProfileStore._clean_values(values)
        allowed_items = {item["pkey"]: item for item in get_production_setup_payload()["parameters"]}
        results = []
        for pkey in sorted(clean_values):
            item = allowed_items.get(pkey)
            if not item:
                results.append({"line": f"{pkey}={clean_values[pkey]}", "ok": False, "error": "NAK_NotFormatRelevant"})
                continue
            if not bool(item.get("can_write_microtom")):
                results.append(
                    {
                        "line": f"{pkey}={clean_values[pkey]}",
                        "ok": True,
                        "skipped": True,
                        "error": "SKIP_ReadOnly",
                    }
                )
                continue
            value = clean_values[pkey]
            line = f"{pkey}={value}"
            try:
                response = router.handle_microtom_line(line, correlation=f"machine-setup-format:{name or 'unsaved'}:{pkey}")
                ok = bool(response) and "NAK" not in str(response).upper()
                results.append({"line": line, "response": response, "ok": ok})
            except Exception as exc:
                results.append({"line": line, "ok": False, "error": str(exc)})
        return {
            "ok": all(bool(item.get("ok")) for item in results) if results else True,
            "name": name or "",
            "count": len(results),
            "results": results,
        }

    @app.get("/ui", response_class=HTMLResponse)
    def ui():
        return home()

    @app.get("/health")
    def health():
        return {"ok": True}

    # -----------------------------
    # UI status (public mini status for Home page)
    # -----------------------------
    @app.get("/api/ui/status/public")
    def ui_status_public():
        return {
            "ok": True,
            "outbox_count": outbox.count(),
            "inbox_pending": inbox.count_pending(),
        }

    # -----------------------------
    # UI status
    # -----------------------------
    @app.get("/api/ui/status")
    def ui_status(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        esp_broker_diag: dict[str, Any] = {}
        if not bool(getattr(cfg2, "esp_simulation", False)) and str(getattr(cfg2, "esp_host", "") or "").strip():
            try:
                esp_broker_diag = EspPlcClient(
                    cfg2.esp_host,
                    int(cfg2.esp_port),
                    timeout_s=float(getattr(cfg2, "esp_connect_timeout_s", 1.5) or 1.5),
                ).diagnostics()
            except Exception as exc:
                esp_broker_diag = {"error": repr(exc)}
        return {
            "ok": True,
            "outbox_count": outbox.count(),
            "inbox_pending": inbox.count_pending(),
            "peer_base_url": cfg2.peer_base_url,
            "peer_base_url_secondary": getattr(cfg2, "peer_base_url_secondary", ""),
            "ntp": {
                "server": getattr(cfg2, "ntp_server", ""),
                "sync_interval_min": getattr(cfg2, "ntp_sync_interval_min", 60),
            },
            "devices": {
                "raspi_io": {
                    "model": getattr(cfg2, "raspi_plc_model", "RPIPLC_21"),
                    "simulation": bool(getattr(cfg2, "raspi_io_simulation", True)),
                    "poll_interval_s": float(getattr(cfg2, "raspi_io_poll_interval_s", 1.0) or 1.0),
                },
                "esp": {
                    "host": cfg2.esp_host,
                    "port": cfg2.esp_port,
                    "simulation": cfg2.esp_simulation,
                    "watchdog_host": cfg2.esp_watchdog_host,
                    "poll_interval_s": float(getattr(cfg2, "esp_io_poll_interval_s", 1.0) or 1.0),
                    "broker": esp_broker_diag,
                },
                "moxa1": {
                    "host": getattr(cfg2, "moxa1_host", ""),
                    "port": int(getattr(cfg2, "moxa1_port", 0) or 0),
                    "simulation": bool(getattr(cfg2, "moxa1_simulation", True)),
                    "poll_interval_s": float(getattr(cfg2, "moxa_poll_interval_s", 1.0) or 1.0),
                },
                "moxa2": {
                    "host": getattr(cfg2, "moxa2_host", ""),
                    "port": int(getattr(cfg2, "moxa2_port", 0) or 0),
                    "simulation": bool(getattr(cfg2, "moxa2_simulation", True)),
                    "poll_interval_s": float(getattr(cfg2, "moxa_poll_interval_s", 1.0) or 1.0),
                },
                "moxa3": {
                    "host": getattr(cfg2, "moxa3_host", ""),
                    "port": int(getattr(cfg2, "moxa3_port", 0) or 0),
                    "simulation": bool(getattr(cfg2, "moxa3_simulation", True)),
                    "poll_interval_s": float(getattr(cfg2, "moxa_poll_interval_s", 1.0) or 1.0),
                },
                "vj3350": {
                    "host": cfg2.vj3350_host,
                    "port": cfg2.vj3350_port,
                    "simulation": cfg2.vj3350_simulation,
                },
                "vj6530": {
                    "host": cfg2.vj6530_host,
                    "port": cfg2.vj6530_port,
                    "simulation": cfg2.vj6530_simulation,
                    "poll_interval_s": getattr(cfg2, "vj6530_poll_interval_s", 15.0),
                    "async_enabled": bool(getattr(cfg2, "vj6530_async_enabled", True)),
                },
                "smart_unwinder": {
                    "host": getattr(cfg2, "smart_unwinder_host", ""),
                    "port": getattr(cfg2, "smart_unwinder_port", 0),
                    "simulation": bool(getattr(cfg2, "smart_unwinder_simulation", True)),
                },
                "smart_rewinder": {
                    "host": getattr(cfg2, "smart_rewinder_host", ""),
                    "port": getattr(cfg2, "smart_rewinder_port", 0),
                    "simulation": bool(getattr(cfg2, "smart_rewinder_simulation", True)),
                },
            }
        }

    @app.post("/api/esp/command")
    def api_esp_command(body: EspCommandReq, x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        line = str(body.line or "").strip()
        if not line or "\n" in line or "\r" in line or len(line) > 1024:
            raise HTTPException(status_code=400, detail="Invalid ESP command line")
        if bool(getattr(cfg2, "esp_simulation", False)) or not str(getattr(cfg2, "esp_host", "") or "").strip():
            return {
                "ok": True,
                "simulation": True,
                "line": line,
                "reply": f"SIM_{line}",
                "broker": {},
            }
        try:
            client = EspPlcClient(
                cfg2.esp_host,
                int(cfg2.esp_port),
                timeout_s=float(getattr(cfg2, "esp_connect_timeout_s", 1.5) or 1.5),
            )
            reply = client.exchange_line(
                line,
                read_timeout_s=max(0.2, float(body.read_timeout_s or 2.0)),
                read_limit=max(128, min(65536, int(body.read_limit or 8192))),
                priority=bool(body.priority),
            )
            return {"ok": True, "line": line, "reply": reply, "broker": client.diagnostics()}
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"ESP command failed: {exc!r}") from exc

    @app.post("/api/queues/outbox/clear")
    def clear_outbox_queue(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        deleted = outbox.clear()
        return {"ok": True, "queue": "outbox", "deleted": deleted, "outbox_count": outbox.count()}

    @app.post("/api/queues/inbox/clear")
    def clear_inbox_queue(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        deleted = inbox.clear()
        return {"ok": True, "queue": "inbox", "deleted": deleted, "inbox_pending": inbox.count_pending()}

    # -----------------------------
    # Config API (Databridge + device endpoints)
    # -----------------------------
    @app.get("/api/config")
    def get_config(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        d = cfg2.__dict__.copy()
        d["ui_token"] = "***"
        d["shared_secret"] = "***" if (cfg2.shared_secret or "") else ""
        d["diclient_adapter_key"] = "***" if (getattr(cfg2, "diclient_adapter_key", "") or "") else ""
        return {"ok": True, "config": d}

    @app.post("/api/config")
    def update_config(u: ConfigUpdate, x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)

        for k, v in u.model_dump().items():
            if v is not None:
                setattr(cfg2, k, v)

        # basic normalization for runtime loops
        try:
            cfg2.ntp_sync_interval_min = int(getattr(cfg2, "ntp_sync_interval_min", 60) or 60)
        except Exception:
            cfg2.ntp_sync_interval_min = 60
        cfg2.ntp_sync_interval_min = max(1, min(24 * 60, cfg2.ntp_sync_interval_min))

        try:
            cfg2.vj6530_poll_interval_s = float(getattr(cfg2, "vj6530_poll_interval_s", 15.0) or 15.0)
        except Exception:
            cfg2.vj6530_poll_interval_s = 15.0
        cfg2.vj6530_poll_interval_s = max(0.5, min(300.0, cfg2.vj6530_poll_interval_s))
        try:
            cfg2.esp_io_poll_interval_s = float(getattr(cfg2, "esp_io_poll_interval_s", 1.0) or 1.0)
        except Exception:
            cfg2.esp_io_poll_interval_s = 1.0
        cfg2.esp_io_poll_interval_s = max(0.2, min(60.0, cfg2.esp_io_poll_interval_s))
        try:
            cfg2.raspi_io_poll_interval_s = float(getattr(cfg2, "raspi_io_poll_interval_s", 1.0) or 1.0)
        except Exception:
            cfg2.raspi_io_poll_interval_s = 1.0
        cfg2.raspi_io_poll_interval_s = max(0.2, min(60.0, cfg2.raspi_io_poll_interval_s))
        try:
            cfg2.moxa_poll_interval_s = float(getattr(cfg2, "moxa_poll_interval_s", 1.0) or 1.0)
        except Exception:
            cfg2.moxa_poll_interval_s = 1.0
        cfg2.moxa_poll_interval_s = max(0.2, min(60.0, cfg2.moxa_poll_interval_s))
        try:
            cfg2.machine_audit_keep_hours = int(float(getattr(cfg2, "machine_audit_keep_hours", 72) or 72))
        except Exception:
            cfg2.machine_audit_keep_hours = 72
        cfg2.machine_audit_keep_hours = max(1, min(24 * 3650, cfg2.machine_audit_keep_hours))
        cfg2.raspi_plc_model = str(getattr(cfg2, "raspi_plc_model", "RPIPLC_21") or "RPIPLC_21").strip() or "RPIPLC_21"

        cfg2.save(cfg_path)
        # Restart service to apply
        subprocess.call(["bash", "-lc", "systemctl restart mas004-rpi-databridge.service"])
        return {"ok": True}

    # -----------------------------
    # Network API (eth0/eth1)
    # -----------------------------
    @app.get("/api/system/network")
    def get_network(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        return {"ok": True, "config": {
            "eth0_ip": cfg2.eth0_ip, "eth0_subnet": cfg2.eth0_subnet, "eth0_gateway": cfg2.eth0_gateway,
            "eth0_dns": getattr(cfg2, "eth0_dns", ""),
            "eth1_ip": cfg2.eth1_ip, "eth1_subnet": cfg2.eth1_subnet, "eth1_gateway": cfg2.eth1_gateway,
            "eth1_dns": getattr(cfg2, "eth1_dns", ""),
        }, "status": get_current_ip_info()}

    @app.get("/api/system/time")
    def get_system_time(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        return get_current_time_info()

    @app.post("/api/system/network")
    def set_network(req: NetworkUpdate, x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)

        def parse_dns(raw: str) -> list[str]:
            txt = (raw or "").strip()
            if not txt:
                return []
            out = []
            for part in re.split(r"[,\s;]+", txt):
                s = (part or "").strip()
                if not s:
                    continue
                octets = s.split(".")
                if len(octets) != 4:
                    raise HTTPException(status_code=400, detail=f"Invalid DNS server '{s}'")
                try:
                    vals = [int(x) for x in octets]
                except Exception:
                    raise HTTPException(status_code=400, detail=f"Invalid DNS server '{s}'")
                if any(v < 0 or v > 255 for v in vals):
                    raise HTTPException(status_code=400, detail=f"Invalid DNS server '{s}'")
                if s not in out:
                    out.append(s)
            return out

        dns0 = parse_dns(req.eth0_dns)
        dns1 = parse_dns(req.eth1_dns)

        # Save into config.json
        cfg2.eth0_ip = req.eth0_ip
        cfg2.eth0_subnet = str(req.eth0_prefix)
        cfg2.eth0_gateway = (req.eth0_gateway or "").strip()
        cfg2.eth0_dns = " ".join(dns0)

        cfg2.eth1_ip = req.eth1_ip
        cfg2.eth1_subnet = str(req.eth1_prefix)
        cfg2.eth1_gateway = (req.eth1_gateway or "").strip()
        cfg2.eth1_dns = " ".join(dns1)

        cfg2.save(cfg_path)

        applied = []
        if req.apply_now:
            # dhcpcd kann bei mehrfachen Restart-Zyklen DNS der ersten NIC ueberschreiben.
            # Deshalb bei leerem eth1 DNS fuer den Live-Apply das eth0 DNS mitgeben.
            dns1_runtime = dns1 if dns1 else dns0
            # try to apply immediately
            r0 = apply_static("eth0", IfaceCfg(ip=req.eth0_ip, prefix=req.eth0_prefix, gw=cfg2.eth0_gateway, dns=dns0))
            r1 = apply_static("eth1", IfaceCfg(ip=req.eth1_ip, prefix=req.eth1_prefix, gw=cfg2.eth1_gateway, dns=dns1_runtime))
            applied = [("eth0", r0), ("eth1", r1)]

        return {"ok": True, "applied": applied}

    # -----------------------------
    # Outbox enqueue helper
    # -----------------------------
    @app.post("/api/outbox/enqueue")
    def api_outbox_enqueue(req: OutboxEnqueue, x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)

        if req.url:
            targets = [req.url]
        else:
            targets = peer_urls(cfg2, req.path)
            if not targets:
                raise HTTPException(status_code=400, detail="No peer base URL configured")

        items = []
        for url in targets:
            idem = outbox.enqueue(req.method, url, req.headers, req.body, req.idempotency_key, priority=50)
            items.append({"url": url, "idempotency_key": idem})
        return {
            "ok": True,
            "count": len(items),
            "items": items,
            "idempotency_key": items[0]["idempotency_key"],
        }

    # -----------------------------
    # Test helper API (manual simulation from UI windows)
    # -----------------------------
    @app.post("/api/test/send")
    def api_test_send(req: TestSendReq, x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)

        src = normalize_test_source(req.source)
        hint = req.ptype_hint if req.ptype_hint is not None else default_ptype_hint.get(src, "")
        lines = [normalize_test_line(part, hint) for part in split_test_messages(req.msg)]
        targets = peer_urls(cfg2, "/api/inbox")
        if not targets:
            raise HTTPException(status_code=400, detail="No peer base URL configured")
        headers = {}
        items = []

        for line in lines:
            item_ts = format_local_timestamp(datetime.now().timestamp())
            parsed = re.match(r"^\s*([A-Za-z]{3})([0-9A-Za-z_]+)\s*=\s*(.+?)\s*$", line)
            persisted = None
            persist_msg = None
            pkey = None
            if parsed:
                ptype = parsed.group(1).upper()
                pid = parsed.group(2)
                rhs = parsed.group(3).strip()
                if rhs != "?":
                    if pid.isdigit():
                        pid = normalize_pid(ptype, pid)
                    pkey = f"{ptype}{pid}"
                    persisted, persist_msg = params.apply_device_value(pkey, rhs)
                    if not persisted:
                        logs.log("raspi", "info", f"value not persisted for {pkey}: {persist_msg}")
                    else:
                        event = production_logs.handle_param_change(pkey, rhs)
                        if event and event.get("event") == "start":
                            logs.log("raspi", "info", f"production logging started: {event.get('production_label')}")
                        elif event and event.get("event") == "stop":
                            logs.log("raspi", "info", f"production logging ready: {event.get('production_label')}")
                        if pkey == "MAS0028" and rhs.strip().lower() in ("", "0", "false", "off", "no", "none", "null"):
                            mark_external_purge_clear(db, source=src)
                            deleted = outbox.delete_status_updates("MAS0028")
                            if deleted:
                                logs.log("raspi", "info", f"cleared {deleted} pending stale MAS0028 status callback(s)")

            if src == "raspi":
                logs.log("raspi", "out", f"manual->microtom: {line}")
                idems = []
                for url in targets:
                    idem = outbox.enqueue("POST", url, headers, {"msg": line, "source": "raspi"}, None, priority=20)
                    idems.append({"url": url, "idempotency_key": idem})
                items.append({
                    "source": src,
                    "line": line,
                    "route": "raspi->microtom",
                    "ack": "ACK_QUEUED",
                    "ts_display": item_ts,
                    "idempotency_key": idems[0]["idempotency_key"],
                    "idempotency_keys": idems,
                    "persisted_local": persisted,
                    "persist_msg": persist_msg,
                })
                continue

            logs.log(src, "out", f"manual->raspi: {line}")
            logs.log("raspi", "in", f"{src}: {line}")
            idems = []
            for url in targets:
                idem = outbox.enqueue(
                    "POST",
                    url,
                    headers,
                    {"msg": line, "source": "raspi", "origin": src},
                    None,
                    priority=50,
                )
                idems.append({"url": url, "idempotency_key": idem})
            logs.log("raspi", "out", f"forward to microtom: {line}")
            items.append({
                "source": src,
                "line": line,
                "route": f"{src}->raspi->microtom",
                "ack": "ACK_QUEUED",
                "ts_display": item_ts,
                "idempotency_key": idems[0]["idempotency_key"],
                "idempotency_keys": idems,
                "persisted_local": persisted,
                "persist_msg": persist_msg,
            })

        first = items[0]
        return {
            "ok": True,
            "source": src,
            "count": len(items),
            "items": items,
            # Legacy single-item fields for older UI clients.
            "line": first["line"],
            "route": first["route"],
            "ack": first["ack"],
            "ts_display": first["ts_display"],
            "idempotency_key": first["idempotency_key"],
            "persisted_local": first["persisted_local"],
            "persist_msg": first["persist_msg"],
        }

    # -----------------------------
    # Inbox (receive from Microtom)
    # -----------------------------
    @app.post("/api/inbox")
    async def api_inbox(
        request: Request,
        x_idempotency_key: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_shared_secret(x_shared_secret, cfg2)

        raw_body = await request.body()
        body = None
        if raw_body:
            try:
                body = json.loads(raw_body.decode("utf-8"))
            except Exception:
                txt = raw_body.decode("utf-8", errors="replace").strip()
                body = {"msg": txt} if txt else None

        headers = dict(request.headers)
        idem = x_idempotency_key or headers.get("x-idempotency-key") or str(uuid.uuid4())
        source = request.client.host if request.client else None
        if isinstance(body, dict):
            src = body.get("source")
            if isinstance(src, str) and src.strip():
                source = src.strip()
        inserted = inbox.store(source, headers, body, idem)
        return {"ok": True, "stored": inserted, "idempotency_key": idem}

    @app.get("/api/inbox/next")
    def api_inbox_next(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)

        msg = inbox.next_pending()
        if not msg:
            return {"ok": True, "msg": None}

        return {
            "ok": True,
            "msg": {
                "id": msg.id,
                "received_ts": msg.received_ts,
                "source": msg.source,
                "headers_json": msg.headers_json,
                "body_json": msg.body_json,
                "idempotency_key": msg.idempotency_key,
            },
        }

    @app.post("/api/inbox/{msg_id}/ack")
    def api_inbox_ack(msg_id: int, x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        inbox.ack(msg_id)
        return {"ok": True}

    # =========================
    # ===== PARAMS API ========
    # =========================
    if MULTIPART_AVAILABLE:

        @app.post("/api/params/import")
        async def params_import(file: UploadFile = File(...), x_token: Optional[str] = Header(default=None)):
            cfg2 = Settings.load(cfg_path)
            require_token(x_token, cfg2)

            suffix = os.path.splitext(file.filename or "")[1].lower()
            if suffix not in (".xlsx",):
                raise HTTPException(status_code=400, detail="Bitte eine .xlsx Datei hochladen")

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp_path = tmp.name
                content = await file.read()
                tmp.write(content)

            try:
                res = params.import_xlsx(tmp_path)
                if res.get("ok"):
                    master_path = cfg2.master_params_xlsx_path
                    os.makedirs(os.path.dirname(master_path), exist_ok=True)
                    shutil.copyfile(tmp_path, master_path)
                    motor_restore = reapply_motor_setup_master_to_params(
                        params,
                        cfg2,
                        get_motor_bindings(),
                        workbook_paths=motor_master_workbook_paths(cfg2),
                    )
                    res["motor_setup_master_reapplied"] = motor_restore
                    res["master_workbook"] = get_master_workbook_info()
                return res
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    else:

        @app.post("/api/params/import")
        def params_import_unavailable(x_token: Optional[str] = Header(default=None)):
            cfg2 = Settings.load(cfg_path)
            require_token(x_token, cfg2)
            raise HTTPException(status_code=503, detail="Workbook-Upload nicht verfuegbar (python-multipart fehlt)")

    @app.get("/api/params/master/info")
    def params_master_info(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        return {"ok": True, "master_workbook": get_master_workbook_info()}

    @app.get("/api/params/master/download")
    def params_master_download(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        info = get_master_workbook_info()
        if not info["exists"]:
            raise HTTPException(status_code=404, detail="Master workbook not stored on Raspi")
        filename = os.path.basename(info["path"])
        return FileResponse(
            info["path"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

    @app.get("/api/params/export")
    def params_export(
        x_token: Optional[str] = Header(default=None),
        ptype: Optional[str] = Query(default=None),
        q: Optional[str] = Query(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)

        data = params.export_xlsx_bytes(ptype=ptype, q=q)
        filename = "params_export.xlsx"
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # =========================
    # ===== IO API ============
    # =========================
    if MULTIPART_AVAILABLE:

        @app.post("/api/io/import")
        async def io_import(
            request: Request,
            file: UploadFile = File(...),
        ):
            cfg2 = Settings.load(cfg_path)
            require_machine_setup_session(request, cfg2)

            suffix = os.path.splitext(file.filename or "")[1].lower()
            if suffix not in (".xlsx",):
                raise HTTPException(status_code=400, detail="Bitte eine .xlsx Datei hochladen")

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp_path = tmp.name
                content = await file.read()
                tmp.write(content)

            try:
                res = io_store.import_xlsx(tmp_path)
                if res.get("ok"):
                    master_path = cfg2.master_ios_xlsx_path
                    os.makedirs(os.path.dirname(master_path), exist_ok=True)
                    shutil.copyfile(tmp_path, master_path)
                    res["master_workbook"] = get_io_workbook_info()
                return res
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass

    else:

        @app.post("/api/io/import")
        def io_import_unavailable(request: Request):
            cfg2 = Settings.load(cfg_path)
            require_machine_setup_session(request, cfg2)
            raise HTTPException(status_code=503, detail="Workbook-Upload nicht verfuegbar (python-multipart fehlt)")

    @app.get("/api/io/master/info")
    def io_master_info(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return {"ok": True, "master_workbook": get_io_workbook_info()}

    @app.get("/api/io/master/download")
    def io_master_download(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        info = get_io_workbook_info()
        if not info["exists"]:
            raise HTTPException(status_code=404, detail="IO workbook not stored on Raspi")
        filename = os.path.basename(info["path"])
        return FileResponse(
            info["path"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

    @app.get("/api/io/overview")
    def io_overview(
        request: Request,
        device: Optional[str] = Query(default=None),
        include_reserved: bool = Query(default=True),
        live: bool = Query(default=False),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        payload = get_io_snapshot(live=live)
        if device:
            payload["points"] = [
                item
                for item in (payload.get("points") or [])
                if str(item.get("device_code") or "").strip().lower() == str(device).strip().lower()
            ]
        if not include_reserved:
            payload["points"] = [item for item in (payload.get("points") or []) if not bool(item.get("is_reserved"))]
        return payload

    @app.post("/api/io/{io_key}/write")
    def io_write(request: Request, io_key: str, req: IoWriteReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        runtime = IoRuntime(cfg2, io_store)
        result = runtime.override_output(io_key, bool(int(req.value)), source="machine-setup-ui-legacy-write")
        payload = get_io_snapshot()
        payload["write_result"] = result
        return payload

    @app.post("/api/io/{io_key}/override")
    def io_override(request: Request, io_key: str, req: IoOverrideReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        runtime = IoRuntime(cfg2, io_store)
        result = runtime.override_output(io_key, bool(int(req.value)), source="machine-setup-ui")
        payload = get_io_snapshot()
        payload["override_result"] = result
        return payload

    @app.post("/api/io/{io_key}/release")
    def io_release(request: Request, io_key: str):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        runtime = IoRuntime(cfg2, io_store)
        result = runtime.release_override(io_key)
        payload = get_io_snapshot()
        payload["release_result"] = result
        return payload

    @app.post("/api/io/overrides/release-all")
    def io_release_all(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        runtime = IoRuntime(cfg2, io_store)
        result = runtime.release_all_overrides()
        payload = get_io_snapshot()
        payload["release_result"] = result
        return payload

    @app.get("/api/params/list")
    def params_list(
        x_token: Optional[str] = Header(default=None),
        ptype: Optional[str] = Query(default=None),
        q: Optional[str] = Query(default=None),
        limit: int = Query(default=200),
        offset: int = Query(default=0),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        return {"ok": True, "items": params.list_params(ptype=ptype, q=q, limit=limit, offset=offset)}

    @app.post("/api/params/edit")
    def params_edit(req: ParamEdit, x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        ok, msg = params.update_meta(
            pkey=req.pkey,
            default_v=req.default_v,
            min_v=req.min_v,
            max_v=req.max_v,
            rw=req.rw,
            esp_rw=req.esp_rw,
        )
        if not ok:
            raise HTTPException(status_code=400, detail=msg)
        return {"ok": True, "msg": msg}

    # ========================
    # ===== LOG API ==========
    # ========================
    @app.get("/api/ui/logs/channels")
    def log_channels(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        return {"ok": True, "channels": logs.list_channels()}

    @app.get("/api/ui/logs")
    def get_logs(
        x_token: Optional[str] = Header(default=None),
        channel: str = Query(...),
        limit: int = Query(default=250),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        return {"ok": True, "items": logs.list_logs(channel, limit=limit)}

    @app.post("/api/ui/logs/clear")
    def clear_logs(
        x_token: Optional[str] = Header(default=None),
        channel: str = Query(...),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        return logs.clear_channel(channel)

    @app.get("/api/ui/logs/download")
    def download_log(
        x_token: Optional[str] = Header(default=None),
        channel: str = Query(...),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        data = logs.read_logfile(channel)
        return Response(
            content=data.encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{channel}.log"'},
        )

    @app.get("/api/logfiles/list")
    def list_logfiles(x_token: Optional[str] = Header(default=None)):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        logs.apply_retention(cfg2)
        items = logs.list_daily_files()
        out = []
        for it in items:
            out.append(
                {
                    "name": it.get("name"),
                    "group": it.get("group"),
                    "group_label": it.get("group_label"),
                    "date": it.get("date"),
                    "size_bytes": it.get("size_bytes"),
                    "mtime_ts": it.get("mtime_ts"),
                }
            )
        return {"ok": True, "items": out}

    @app.get("/api/logfiles/download")
    def download_daily_logfile(
        x_token: Optional[str] = Header(default=None),
        name: str = Query(...),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token(x_token, cfg2)
        try:
            data = logs.read_daily_file(name)
        except Exception as e:
            raise HTTPException(status_code=404, detail=str(e))
        safe_name = os.path.basename(name)
        return Response(
            content=data.encode("utf-8"),
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
        )

    @app.get("/api/production/logfiles/list")
    def list_production_logfiles(
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token_or_shared_secret(x_token, x_shared_secret, cfg2)
        return production_logs.ready_manifest()

    @app.get("/api/production/logfiles/download")
    def download_production_logfile(
        name: str = Query(...),
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token_or_shared_secret(x_token, x_shared_secret, cfg2)
        try:
            data = logs.consume_production_file(name)
        except Exception as e:
            raise HTTPException(status_code=404, detail=str(e))
        safe_name = os.path.basename(name)
        return Response(
            content=data,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
        )

    @app.post("/api/production/logfiles/ack")
    def ack_production_logfiles(
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_token_or_shared_secret(x_token, x_shared_secret, cfg2)
        return logs.acknowledge_production_files()

    # =========================
    # ===== Motors API ========
    # =========================
    @app.get("/api/motors/overview")
    def api_motors_overview(
        request: Request,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(reason="motors_overview")
        return get_motor_overview_payload()

    @app.get("/api/motors/poll")
    def api_motor_poll_state(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(reason="motors_poll_state")
        return get_motor_poll_payload(cfg2)

    @app.post("/api/motors/poll")
    def api_motor_poll_set(request: Request, body: MotorPollReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(reason="motors_poll_set")
        client = EspMotorClient(cfg2)
        if not client.available():
            raise HTTPException(status_code=503, detail="ESP motor endpoint missing or simulation enabled")
        if bool(body.enabled):
            raise HTTPException(
                status_code=409,
                detail="Globales ESP-Motorpolling ist deaktiviert. Bitte pro Motor die 1s-Polling-Checkbox im Motors-Setup verwenden.",
            )
        result = client.set_poll(bool(body.enabled))
        if not bool(result.get("ok")):
            raise HTTPException(status_code=502, detail=result.get("reply") or "ESP rejected MOTOR POLL command")
        payload = get_motor_poll_payload(cfg2)
        payload["set_result"] = result
        return payload

    @app.get("/api/motors/{motor_id}")
    def api_motor_status(
        request: Request,
        motor_id: int,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_status")
        if not motor_live_allowed(motor_id, cfg2):
            merged = get_motor_overview_payload()
            for motor in merged.get("motors") or []:
                if int(motor.get("id") or 0) == int(motor_id):
                    return motor
            raise HTTPException(status_code=404, detail=f"Unknown motor id {motor_id}")
        client = get_motor_client()
        if client.available():
            payload = motor_client_call("STATUS", lambda: client.status(motor_id))
            bindings = get_motor_bindings().get(int(motor_id))
            payload["bindings"] = bindings
            return payload
        merged = get_motor_overview_payload()
        for motor in merged.get("motors") or []:
            if int(motor.get("id") or 0) == int(motor_id):
                return motor
        raise HTTPException(status_code=404, detail=f"Unknown motor id {motor_id}")

    @app.post("/api/motors/{motor_id}/move")
    def api_motor_move(
        request: Request,
        motor_id: int,
        body: MotorMoveReq,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_move")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        mode = (body.mode or "").strip().lower()
        guard_motor_position_move(client, motor_id, mode, float(body.value))
        if mode == "relative_steps":
            result = motor_client_call(
                "MOVE_REL_STEPS",
                lambda: client.move_relative_steps(motor_id, int(round(body.value))),
            )
            return motor_status_action_response(client, motor_id, "MOVE_REL_STEPS", result)
        if mode == "relative_mm":
            result = motor_client_call("MOVE_REL_MM", lambda: client.move_relative_mm(motor_id, float(body.value)))
            return motor_status_action_response(client, motor_id, "MOVE_REL_MM", result)
        if mode == "absolute_mm":
            result = motor_client_call("MOVE_ABS_MM", lambda: client.move_absolute_mm(motor_id, float(body.value)))
            return motor_status_action_response(client, motor_id, "MOVE_ABS_MM", result)
        raise HTTPException(status_code=400, detail="Unsupported motor move mode")

    @app.post("/api/motors/{motor_id}/position")
    def api_motor_position(
        request: Request,
        motor_id: int,
        body: MotorPositionReq,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_position")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        result = motor_client_call(
            "SET_POSITION_MM",
            lambda: client.set_current_position_mm(
                motor_id,
                float(body.value),
                allow_machine_setup_write=True,
            ),
        )
        require_motor_ack(result, "SET_POSITION_MM")
        save_result = motor_client_call(
            "SAVE",
            lambda: client.save(motor_id, allow_machine_setup_write=True),
        )
        require_motor_ack(save_result, "SAVE")
        combined = {
            "ok": True,
            "reply": f"{result.get('reply', '')}; {save_result.get('reply', '')}".strip("; "),
        }
        response = motor_action_response(client, motor_id, "SET_POSITION_MM", combined)
        verify_motor_feedback_position(motor_id, response, float(body.value))
        return sync_motor_master_from_response(cfg2, motor_id, "SET_POSITION_MM", response)

    @app.post("/api/motors/{motor_id}/config")
    def api_motor_config(
        request: Request,
        motor_id: int,
        body: MotorConfigReq,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_config")
        payload = body.model_dump(exclude_none=True)
        position_mm = payload.pop("position_mm", None)
        if int(motor_id) != 3:
            if payload.get("min_enabled") is False:
                raise HTTPException(
                    status_code=409,
                    detail=f"Motor {int(motor_id)} Min-Grenze darf bei Positionsachsen nicht deaktiviert werden",
                )
            if payload.get("max_enabled") is False:
                raise HTTPException(
                    status_code=409,
                    detail=f"Motor {int(motor_id)} Max-Grenze darf bei Positionsachsen nicht deaktiviert werden",
                )
            current_cfg: dict[str, Any] = {}
            try:
                client_for_cfg = EspMotorClient(cfg2)
                if client_for_cfg.available():
                    current_cfg = (motor_client_call("CONFIG", lambda: client_for_cfg.config(motor_id)) or {}).get("config") or {}
            except HTTPException:
                raise
            except Exception:
                current_cfg = {}
            min_tenths = payload.get("min_tenths_mm", current_cfg.get("min_tenths_mm"))
            max_tenths = payload.get("max_tenths_mm", current_cfg.get("max_tenths_mm"))
            min_int = _to_int_or_none(min_tenths)
            max_int = _to_int_or_none(max_tenths)
            if min_int is not None and max_int is not None and min_int > max_int:
                raise HTTPException(status_code=409, detail="Min-Grenze darf nicht groesser als Max-Grenze sein")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        result = motor_client_call(
            "SET_CONFIG",
            lambda: client.set_config(motor_id, payload, allow_machine_setup_write=True),
        )
        response = save_and_refresh_motor_master(
            cfg2,
            client,
            motor_id,
            "SET_CONFIG",
            result,
            config_payload=payload,
            refresh_after_save=False,
        )
        if position_mm is None:
            return response

        position_result = motor_client_call(
            "SET_POSITION_MM",
            lambda: client.set_current_position_mm(
                motor_id,
                float(position_mm),
                allow_machine_setup_write=True,
            ),
        )
        position_response = save_and_refresh_motor_master(
            cfg2,
            client,
            motor_id,
            "SET_POSITION_MM",
            position_result,
            expected_position_mm=float(position_mm),
        )
        position_response["config_save"] = response
        return position_response

    @app.post("/api/motors/{motor_id}/save")
    def api_motor_save(
        request: Request,
        motor_id: int,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_save")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        result = motor_client_call(
            "SAVE",
            lambda: client.save(motor_id, allow_machine_setup_write=True),
        )
        return motor_config_action_response(cfg2, client, motor_id, "SAVE", result)

    @app.post("/api/motors/{motor_id}/zero")
    def api_motor_zero(
        request: Request,
        motor_id: int,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_zero")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        result = motor_client_call("ZERO", lambda: client.zero(motor_id, allow_machine_setup_write=True))
        return save_and_refresh_motor_master(cfg2, client, motor_id, "ZERO", result, expected_position_mm=0.0)

    @app.post("/api/motors/{motor_id}/min")
    def api_motor_min(
        request: Request,
        motor_id: int,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_min")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        result = motor_client_call("SET_MIN", lambda: client.set_min(motor_id, allow_machine_setup_write=True))
        return save_and_refresh_motor_master(
            cfg2,
            client,
            motor_id,
            "SET_MIN",
            result,
            refresh_after_save=False,
        )

    @app.post("/api/motors/{motor_id}/max")
    def api_motor_max(
        request: Request,
        motor_id: int,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_max")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        result = motor_client_call("SET_MAX", lambda: client.set_max(motor_id, allow_machine_setup_write=True))
        return save_and_refresh_motor_master(
            cfg2,
            client,
            motor_id,
            "SET_MAX",
            result,
            refresh_after_save=False,
        )

    @app.post("/api/motors/{motor_id}/reset-alarm")
    def api_motor_reset_alarm(
        request: Request,
        motor_id: int,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_reset_alarm")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        result = motor_client_call("RESET_ALARM", lambda: client.reset_alarm(motor_id))
        return motor_action_response(client, motor_id, "RESET_ALARM", result)

    @app.post("/api/motors/{motor_id}/refresh")
    def api_motor_refresh(
        request: Request,
        motor_id: int,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_refresh")
        require_live_motor_or_raise(motor_id, cfg2)
        client = EspMotorClient(cfg2)
        return motor_client_call("REFRESH", lambda: refresh_motor_snapshot(client, motor_id))

    @app.post("/api/motors/{motor_id}/simulation")
    def api_motor_simulation(
        request: Request,
        motor_id: int,
        body: MotorSimulationReq,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        touch_motor_setup_authority(motor_id=motor_id, reason="motor_simulation")
        store = get_motor_state_store()
        ids = sorted(store.set_simulation(motor_id, bool(body.enabled)))
        merged = get_motor_overview_payload()
        motor_payload = next((item for item in (merged.get("motors") or []) if int(item.get("id") or 0) == int(motor_id)), None)
        return {"ok": True, "simulation_ids": ids, "motor": motor_payload}

    @app.get("/api/winders/{role}/state")
    def api_winder_state(
        request: Request,
        role: str,
        x_token: Optional[str] = Header(default=None),
        x_shared_secret: Optional[str] = Header(default=None),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return get_winder_state(role)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    @app.get("/api/machine/overview")
    def api_machine_overview(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return get_machine_overview()

    @app.get("/api/machine/production-visualization")
    def api_machine_production_visualization(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return get_production_visualization_payload()

    @app.post("/api/machine/production-visualization/component")
    def api_machine_production_visualization_component(request: Request, body: ProductionVisualizationComponentReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return set_visualization_component_position(cfg2, body)

    @app.get("/api/machine/mae0048-diagnostics")
    def api_machine_mae0048_diagnostics(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return collect_mae0048_diagnostics(cfg2, db)

    @app.get("/api/machine/motor3-calibration/status")
    def api_machine_motor3_calibration_status(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return motor3_calibration_status(cfg2)

    @app.post("/api/machine/motor3-calibration/prepare")
    def api_machine_motor3_calibration_prepare(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return start_motor3_calibration_process("prepare")
        except RuntimeError as exc:
            raise HTTPException(status_code=409, detail=str(exc))

    @app.post("/api/machine/motor3-calibration/start")
    def api_machine_motor3_calibration_start(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return start_motor3_calibration_process("run")
        except RuntimeError as exc:
            raise HTTPException(status_code=409, detail=str(exc))

    @app.post("/api/machine/motor3-calibration/apply")
    def api_machine_motor3_calibration_apply(request: Request, body: Motor3CalibrationApplyReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return apply_motor3_calibration(cfg2, body)
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    @app.post("/api/machine/led-test")
    def api_machine_led_test(request: Request, body: MachineLedTestReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return execute_led_test_command(cfg2, body)

    @app.post("/api/machine/button")
    def api_machine_button(request: Request, body: MachineButtonReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        clear_motor_setup_manual_lock(db, reason=f"machine_button:{body.button}")
        runtime = MachineRuntime(cfg2, db, params, io_store, logs, outbox)
        try:
            return runtime.press_virtual_button(body.button)
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    @app.get("/api/machine/bypass")
    def api_machine_bypass(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return get_machine_bypass_payload()

    @app.post("/api/machine/bypass")
    def api_machine_bypass_update(request: Request, body: MachineBypassReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        payload = write_machine_bypass_values(body.values)
        if not bool(payload.get("ok")):
            failed = [item for item in payload.get("results", []) if not item.get("ok")]
            raise HTTPException(status_code=400, detail={"message": "Bypass write failed", "failed": failed})
        return payload

    @app.get("/api/machine/audit")
    def api_machine_audit(
        request: Request,
        hours: Optional[int] = Query(default=None),
        limit: int = Query(default=500),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        logs.apply_audit_retention(cfg2)
        keep_hours = logs.audit_keep_hours_from_settings(cfg2)
        view_hours = hours if hours is not None else keep_hours
        return {
            "ok": True,
            "keep_hours": keep_hours,
            "view_hours": view_hours,
            "entries": logs.list_audit_entries(hours=view_hours, limit=limit),
        }

    @app.post("/api/machine/audit/retention")
    def api_machine_audit_retention(request: Request, body: MachineAuditRetentionReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            keep_hours = int(float(body.keep_hours))
        except Exception:
            keep_hours = 72
        cfg2.machine_audit_keep_hours = max(1, min(24 * 3650, keep_hours))
        cfg2.save(cfg_path)
        logs.apply_audit_retention(cfg2)
        return {"ok": True, "keep_hours": cfg2.machine_audit_keep_hours}

    @app.get("/api/machine/audit/download")
    def api_machine_audit_download(
        request: Request,
        hours: Optional[int] = Query(default=None),
        limit: int = Query(default=5000),
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        keep_hours = logs.audit_keep_hours_from_settings(cfg2)
        view_hours = hours if hours is not None else keep_hours
        content = logs.read_audit_log(hours=view_hours, limit=limit)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return Response(
            content,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="mas004_machine_audit_{stamp}.log"'},
        )

    @app.get("/api/production-setup/parameters")
    def api_production_setup_parameters(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return get_production_setup_payload()

    @app.get("/api/production-setup/status")
    def api_production_setup_status(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        payload = get_production_setup_payload()
        return {
            "ok": True,
            "production_status": payload.get("production_status", []),
            "machine": payload.get("machine", {}),
        }

    @app.get("/api/production-setup/profiles")
    def api_production_setup_profiles(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return {"ok": True, "profiles": format_profiles.list_profiles()}

    @app.get("/api/production-setup/profiles/{name}")
    def api_production_setup_profile(request: Request, name: str):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            profile = format_profiles.get_profile(name)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        if not profile:
            raise HTTPException(status_code=404, detail="Formatprofil nicht gefunden")
        return {"ok": True, "profile": profile}

    @app.post("/api/production-setup/profiles")
    def api_production_setup_profile_save(request: Request, body: FormatProfileReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            profile = format_profiles.save_profile(body.name, body.values, note=body.note)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        return {"ok": True, "profile": profile, "profiles": format_profiles.list_profiles()}

    @app.delete("/api/production-setup/profiles/{name}")
    def api_production_setup_profile_delete(request: Request, name: str):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return format_profiles.delete_profile(name)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    @app.post("/api/production-setup/send")
    def api_production_setup_send(request: Request, body: FormatSendReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            name = normalize_profile_name(body.name) if (body.name or "").strip() else ""
        except ValueError:
            name = str(body.name or "").strip()[:80]
        return send_format_values(body.values, name=name)

    @app.get("/api/commissioning/overview")
    def api_commissioning_overview(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return get_commissioning_store(cfg2).overview()

    @app.post("/api/commissioning/run/start")
    def api_commissioning_start(request: Request, body: CommissioningStartReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return get_commissioning_store(cfg2).start_run(body.mode)

    @app.get("/api/commissioning/run/{run_id}")
    def api_commissioning_run(request: Request, run_id: int):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return get_commissioning_store(cfg2).get_run(run_id)

    @app.post("/api/commissioning/run/{run_id}/step/{step_id}")
    def api_commissioning_step_update(
        request: Request,
        run_id: int,
        step_id: str,
        body: CommissioningStepReq,
    ):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return get_commissioning_store(cfg2).update_step(run_id, step_id, body.status, note=body.note)
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    @app.post("/api/commissioning/run/{run_id}/step/{step_id}/check")
    def api_commissioning_step_check(request: Request, run_id: int, step_id: str):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return get_commissioning_store(cfg2).auto_check_step(run_id, step_id)
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    @app.get("/api/backups/overview")
    def api_backups_overview(request: Request):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        return get_backup_manager(cfg2).overview()

    @app.post("/api/backups/identity")
    def api_backups_identity(request: Request, body: BackupIdentityReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        if not str(body.machine_serial_number or "").strip():
            raise HTTPException(status_code=400, detail="machine_serial_number is required")
        mgr = get_backup_manager(cfg2)
        return {"ok": True, "identity": mgr.set_identity(body.machine_serial_number, body.machine_name)}

    @app.post("/api/backups/create")
    def api_backups_create(request: Request, body: BackupCreateReq):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        mgr = get_backup_manager(cfg2)
        if not str(body.name or "").strip():
            raise HTTPException(status_code=400, detail="Backup name is required")
        backup_type = str(body.backup_type or "").strip().lower()
        try:
            if backup_type == "settings":
                result = mgr.create_settings_backup(body.name, note=body.note)
            elif backup_type == "full":
                result = mgr.create_full_backup(body.name, note=body.note)
            else:
                raise RuntimeError("backup_type must be 'settings' or 'full'")
            return {"ok": True, "backup": result}
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    if MULTIPART_AVAILABLE:

        @app.post("/api/backups/import")
        async def api_backups_import(
            request: Request,
            file: UploadFile = File(...),
        ):
            cfg2 = Settings.load(cfg_path)
            require_machine_setup_session(request, cfg2)
            suffix = os.path.splitext(file.filename or "backup.zip")[1] or ".zip"
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                content = await file.read()
                tmp.write(content)
                tmp_path = tmp.name
            try:
                backup = get_backup_manager(cfg2).import_backup(tmp_path, file.filename or "")
                return {"ok": True, "backup": backup}
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass

    else:

        @app.post("/api/backups/import")
        def api_backups_import_unavailable(request: Request):
            cfg2 = Settings.load(cfg_path)
            require_machine_setup_session(request, cfg2)
            raise HTTPException(status_code=503, detail="Backup-Import nicht verfuegbar (python-multipart fehlt)")

    @app.get("/api/backups/{backup_id}/download")
    def api_backups_download(request: Request, backup_id: str):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        backup = get_backup_manager(cfg2).get_backup(backup_id)
        filename = os.path.basename(backup["file_path"])
        return FileResponse(backup["file_path"], media_type="application/zip", filename=filename)

    @app.post("/api/backups/{backup_id}/restore")
    def api_backups_restore(request: Request, backup_id: str):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return get_backup_manager(cfg2).restore_backup(backup_id)
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    @app.delete("/api/backups/{backup_id}")
    def api_backups_delete(request: Request, backup_id: str):
        cfg2 = Settings.load(cfg_path)
        require_machine_setup_session(request, cfg2)
        try:
            return get_backup_manager(cfg2).delete_backup(backup_id)
        except RuntimeError as exc:
            raise HTTPException(status_code=404, detail=str(exc))

    # =========================
    # ===== SIMPLE UI =========
    # =========================
    @app.get("/ui/machine-setup", include_in_schema=False)
    def ui_machine_setup(request: Request):
        cfg2 = Settings.load(cfg_path)
        target = "/ui/machine-setup/motors"
        if has_machine_setup_session(request, cfg2):
            return RedirectResponse(url=target, status_code=303)
        return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)

    @app.get("/ui/machine-setup/commissioning", response_class=HTMLResponse, include_in_schema=False)
    def ui_machine_commissioning(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/commissioning"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("commissioning")
        return build_commissioning_ui_html(nav)

    @app.get("/ui/machine-setup/backups", response_class=HTMLResponse, include_in_schema=False)
    def ui_machine_backups(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/backups"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("backups")
        return build_backup_ui_html(nav)

    @app.get("/ui/machine-setup/production", response_class=HTMLResponse, include_in_schema=False)
    def ui_machine_production(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/production"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("production")
        return build_production_setup_ui_html(nav)

    @app.get("/ui/machine-setup/login", response_class=HTMLResponse, include_in_schema=False)
    def ui_machine_setup_login(request: Request, next: str = Query(default="/ui/machine-setup/motors")):
        cfg2 = Settings.load(cfg_path)
        target = sanitize_machine_setup_target(next)
        if has_machine_setup_session(request, cfg2):
            return RedirectResponse(url=target, status_code=303)
        return HTMLResponse(machine_setup_login_html(target))

    @app.post("/ui/machine-setup/login", include_in_schema=False)
    def ui_machine_setup_login_submit(
        request: Request,
        response: Response,
        body: MachineSetupLoginReq,
    ):
        cfg2 = Settings.load(cfg_path)
        target = sanitize_machine_setup_target(body.next)
        if body.username != MACHINE_SETUP_USER or body.password != MACHINE_SETUP_PASSWORD:
            raise HTTPException(status_code=401, detail="Ungueltiger Benutzer oder Passwort.")
        response.set_cookie(MACHINE_SETUP_COOKIE, build_machine_setup_cookie(cfg2), **machine_setup_cookie_kwargs(request, cfg2))
        return {"ok": True, "redirect": target}

    @app.get("/ui/machine-setup/logout", include_in_schema=False)
    def ui_machine_setup_logout():
        response = RedirectResponse(url="/ui/machine-setup/login", status_code=303)
        response.delete_cookie(MACHINE_SETUP_COOKIE, path="/")
        return response

    @app.get("/ui/process", include_in_schema=False)
    def ui_process_redirect():
        return RedirectResponse(url="/ui/machine-setup/process", status_code=303)

    @app.get("/ui/machine-setup/visualization", response_class=HTMLResponse, include_in_schema=False)
    def ui_machine_production_visualization(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/visualization"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("visualization")
        return build_production_visualization_ui_html(nav)

    @app.get("/ui/machine-setup/mae0048", response_class=HTMLResponse, include_in_schema=False)
    def ui_machine_mae0048_diagnostics(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/mae0048"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("mae0048")
        return build_mae0048_diagnostics_ui_html(nav)

    @app.get("/ui/machine-setup/calibration", response_class=HTMLResponse, include_in_schema=False)
    def ui_machine_motor3_calibration(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/calibration"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("calibration")
        return build_motor3_calibration_ui_html(nav)

    @app.get("/ui/machine-setup/process", response_class=HTMLResponse, include_in_schema=False)
    def ui_machine_process(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/process"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("process")
        return build_machine_control_ui_html(nav)
        return """
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>Machine Process</title>
  <style>
    :root{--bg:#f4f6f9;--card:#fff;--text:#1f2933;--muted:#5f6b7a;--border:#d6dde7;--blue:#005eb8}
    *{box-sizing:border-box} body{margin:0;font-family:Segoe UI,Arial,sans-serif;background:var(--bg);color:var(--text)}
    .wrap{max-width:1500px;margin:0 auto;padding:16px}
    .toolbar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:12px}
    .grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:12px}
    .card{background:#fff;border:1px solid var(--border);border-radius:12px;padding:14px}
    .btn{min-height:38px;padding:8px 12px;border:1px solid #aec4db;border-radius:10px;background:#e8f0f8;color:#17324b;font-weight:600;cursor:pointer}
    .muted{color:var(--muted)} .kv{display:grid;grid-template-columns:170px 1fr;gap:6px 10px}
    .box{display:inline-block;padding:4px 8px;border-radius:999px;border:1px solid var(--border);background:#eef3f8;font-size:12px}
    table{width:100%;border-collapse:collapse} th,td{padding:8px;border-bottom:1px solid var(--border);text-align:left;vertical-align:top;font-size:13px}
    code{white-space:pre-wrap;word-break:break-word}
  </style>
</head>
<body>
<div class="wrap">
__NAV__
  <div class="toolbar">
    <button class="btn" onclick="loadAll()">Reload</button>
    <span id="status" class="muted">loading...</span>
  </div>
  <div class="grid">
    <div class="card"><h3>Maschinenzustand</h3><div class="kv" id="state_kv"></div></div>
    <div class="card"><h3>Sicherheit / Bedienung</h3><div class="kv" id="flags_kv"></div></div>
    <div class="card"><h3>Statusleuchte / Buttons</h3><div class="kv" id="io_kv"></div></div>
  </div>
  <div class="grid" style="margin-top:12px">
    <div class="card">
      <h3>Letzte Ereignisse</h3>
      <table><thead><tr><th>Zeit</th><th>Typ</th><th>Meldung</th></tr></thead><tbody id="events"></tbody></table>
    </div>
    <div class="card">
      <h3>Letzte Labels</h3>
      <table><thead><tr><th>Label</th><th>Status</th><th>Payload</th></tr></thead><tbody id="labels"></tbody></table>
    </div>
  </div>
</div>
<script>
const TOKEN_KEY="mas004_ui_token";
function token(){ try{return localStorage.getItem(TOKEN_KEY)||"";}catch(e){return"";} }
function esc(v){ return String(v ?? "").replaceAll("&","&amp;").replaceAll("<","&lt;").replaceAll(">","&gt;").replaceAll('"',"&quot;"); }
async function api(path,opt={}){
  opt.headers=opt.headers||{};
  const t=token(); if(t) opt.headers["X-Token"]=t;
  const r=await fetch(path,opt); const txt=await r.text();
  let j=null; try{j=JSON.parse(txt);}catch(e){}
  if(!r.ok){ throw new Error((j&&j.detail)?j.detail:(`HTTP ${r.status} ${txt}`)); }
  return j;
}
function kv(rootId, pairs){
  const root=document.getElementById(rootId);
  root.innerHTML = pairs.map(([k,v])=>`<div class="muted">${esc(k)}</div><div>${v}</div>`).join("");
}
async function loadAll(){
  document.getElementById("status").textContent="loading...";
  try{
    const j=await api("/api/machine/overview");
    const info=j.info || {};
    kv("state_kv", [
      ["Current", `<span class="box">${esc(j.current_state)} - ${esc(j.current_state_label)}</span>`],
      ["Requested", `<span class="box">${esc(j.requested_state)} - ${esc(j.requested_state_label)}</span>`],
      ["Production label", esc(j.production_label || "-")],
      ["Last label no", esc(j.last_label_no ?? "-")]
    ]);
    kv("flags_kv", [
      ["Warning active", esc(!!j.warning_active)],
      ["Purge active", esc(!!j.purge_active)],
      ["Allowed actions", `<code>${esc(JSON.stringify(info.allowed_actions || {}, null, 2))}</code>`],
      ["Button mask", `<code>${esc(JSON.stringify(info.button_mask || {}, null, 2))}</code>`],
      ["Critical reasons", `<code>${esc(JSON.stringify(info.critical_reasons || [], null, 2))}</code>`]
    ]);
    kv("io_kv", [
      ["Status lamp", `<code>${esc(JSON.stringify(info.status_lamp || {}, null, 2))}</code>`],
      ["Button inputs", `<code>${esc(JSON.stringify(info.button_inputs || {}, null, 2))}</code>`],
      ["Button LEDs", `<code>${esc(JSON.stringify(info.button_leds || {}, null, 2))}</code>`],
      ["Warnings", `<code>${esc(JSON.stringify(info.warning_keys || [], null, 2))}</code>`],
      ["Errors", `<code>${esc(JSON.stringify(info.error_keys || [], null, 2))}</code>`]
    ]);
    document.getElementById("events").innerHTML = (j.events || []).map(it =>
      `<tr><td>${esc(new Date((it.ts||0)*1000).toLocaleString())}</td><td>${esc(it.event_type || "")}</td><td>${esc(it.message || "")}</td></tr>`
    ).join("");
    document.getElementById("labels").innerHTML = (j.labels || []).map(it =>
      `<tr><td>${esc(it.label_no || "")}</td><td>${esc(`M=${it.material_ok?1:0} P=${it.print_ok?1:0} V=${it.verify_ok?1:0} R=${it.removed?1:0} OK=${it.production_ok?1:0}`)}</td><td><code>${esc(JSON.stringify(it.payload || {}, null, 2))}</code></td></tr>`
    ).join("");
    document.getElementById("status").textContent="ok";
  }catch(err){
    document.getElementById("status").textContent=err.message;
  }
}
loadAll();
setInterval(()=>{ if(!document.hidden) loadAll(); }, 1000);
</script>
</body>
</html>
        """.replace("__NAV__", nav)

    @app.get("/ui/winders/{role}", include_in_schema=False)
    def ui_winder_redirect(role: str):
        try:
            normalized = normalize_winder_role(role)
        except ValueError:
            raise HTTPException(status_code=404, detail="Unknown winder role")
        return RedirectResponse(url=f"/ui/machine-setup/winders/{normalized}", status_code=303)

    @app.get("/ui/machine-setup/winders/{role}", response_class=HTMLResponse, include_in_schema=False)
    def ui_winder(request: Request, role: str):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = sanitize_machine_setup_target(f"/ui/machine-setup/winders/{role}")
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        try:
            normalized = normalize_winder_role(role)
        except ValueError:
            raise HTTPException(status_code=404, detail="Unknown winder role")
        nav = nav_html("machine_setup") + machine_setup_nav_html("unwinder" if normalized == "unwinder" else "rewinder")
        label = "Abwickler" if normalized == "unwinder" else "Aufwickler"
        return build_winder_ui_html(normalized, label, nav)

    @app.get("/ui/io", include_in_schema=False)
    def ui_io_redirect():
        return RedirectResponse(url="/ui/machine-setup/io", status_code=303)

    @app.get("/ui/machine-setup/io", response_class=HTMLResponse, include_in_schema=False)
    def ui_io(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/io"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("io")
        return """
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>Hardware I/O</title>
  <style>
    :root{
      --bg:#f4f6f9; --card:#fff; --text:#1f2933; --muted:#5f6b7a; --border:#d6dde7; --blue:#005eb8;
      --good:#2e7d32; --warn:#ed6c02; --bad:#c62828;
    }
    *{box-sizing:border-box}
    body{margin:0; font-family:Segoe UI,Arial,sans-serif; background:var(--bg); color:var(--text)}
    .wrap{max-width:1700px; margin:0 auto; padding:16px}
    .toolbar,.row,.actions{display:flex; gap:10px; align-items:center; flex-wrap:wrap}
    .toolbar{margin-bottom:12px}
    .btn{min-height:38px; padding:8px 12px; border:1px solid #aec4db; border-radius:10px; background:#e8f0f8; color:#17324b; font-weight:600; cursor:pointer}
    .btn.small{min-height:30px; padding:4px 8px; font-size:12px}
    .grid{display:grid; gap:12px}
    .device-grid{grid-template-columns:repeat(auto-fit,minmax(250px,1fr)); margin-bottom:12px}
    .card{background:#fff; border:1px solid var(--border); border-radius:12px; padding:14px}
    .muted{color:var(--muted)}
    .pill{display:inline-flex; align-items:center; gap:8px; padding:6px 10px; border-radius:999px; border:1px solid var(--border); background:#eef3f8; font-size:12px}
    .ok{color:var(--good)} .warn{color:var(--warn)} .bad{color:var(--bad)}
    table{width:100%; border-collapse:collapse; background:#fff; border:1px solid var(--border); border-radius:12px; overflow:hidden}
    th,td{padding:8px 10px; border-top:1px solid #e7edf6; text-align:left; vertical-align:top}
    th{background:#f7fafc; color:#425466; font-size:12px; text-transform:uppercase; letter-spacing:.04em}
    tr:first-child td{border-top:none}
    .state-1{color:var(--good); font-weight:700}
    .state-0{color:#6b7280}
    .quality-live{color:var(--good); font-weight:700}
    .quality-simulation{color:var(--warn); font-weight:700}
    .quality-offline{color:var(--bad); font-weight:700}
    .quality-override{color:#7a4f00; font-weight:700}
    .quality-default{color:#6b7280}
    .io-high.active{background:#d7f8df; border-color:#54b96b; color:#14532d}
    .io-low.active{background:#ffd9d9; border-color:#e57373; color:#8a1111}
    .io-release.active{background:#fff3bf; border-color:#e0b22f; color:#654400}
    .io-action-set{display:flex; gap:6px; flex-wrap:wrap}
    .checkline{display:inline-flex; gap:8px; align-items:center}
    @media(max-width:900px){ .hide-mobile{display:none} }
  </style>
</head>
<body>
  <div class="wrap">
    __NAV__
    <div class="toolbar">
      <button class="btn" onclick="reloadAll()">Reload</button>
      <button class="btn" onclick="reloadAll(false,true)">Live lesen</button>
      <button class="btn" onclick="downloadWorkbook()">IO Workbook herunterladen</button>
      <label class="btn" for="io_file">IO Workbook importieren</label>
      <input id="io_file" type="file" accept=".xlsx" style="display:none" onchange="uploadWorkbook(this.files[0])"/>
      <label class="checkline"><input type="checkbox" id="show_reserved" checked onchange="renderRows(window.__io || null)"/>Reserven anzeigen</label>
      <span class="muted">Motorpolling: aus; Live-Refresh nur im Motors-Setup pro Motor</span>
      <span id="status" class="muted"></span>
    </div>
    <div class="card" style="margin-bottom:12px">
      <div class="row">
        <span class="pill">Masterdatei</span>
        <span id="wb_path" class="muted">-</span>
        <span id="wb_meta" class="muted">-</span>
      </div>
    </div>
    <div id="devices" class="grid device-grid"></div>
    <div class="card" style="padding:0; overflow:auto">
      <table>
        <thead>
          <tr>
            <th>Device</th>
            <th>Pin</th>
            <th>Richtung</th>
            <th>Funktion</th>
            <th>Wert</th>
            <th>Qualitaet</th>
            <th class="hide-mobile">Zone</th>
            <th>Aktion</th>
          </tr>
        </thead>
        <tbody id="rows"></tbody>
      </table>
    </div>
  </div>
<script>
let refreshHandle = null;
window.__io = null;
window.__activeOverrides = [];
window.__skipOverrideUnloadPrompt = false;

async function api(path, opt={}){
  const r = await fetch(path, opt);
  const txt = await r.text();
  let j = null; try{ j = JSON.parse(txt); }catch(e){}
  if(!r.ok){ throw new Error((j && j.detail) ? j.detail : (`HTTP ${r.status} ${txt}`)); }
  return j;
}

function esc(v){ return String(v ?? "").replaceAll("&","&amp;").replaceAll("<","&lt;").replaceAll(">","&gt;").replaceAll('"',"&quot;"); }
function qualityClass(v){ return `quality-${String(v || "default").toLowerCase()}`; }
function valueClass(v){ return Number(v || 0) ? "state-1" : "state-0"; }
function isPulseOnly(item){ return false; }
function canOverride(item){ return (item.io_dir === "output" || item.io_dir === "gpio") && !isPulseOnly(item); }
function overrideLabel(item){
  if(!item.override_active){ return ""; }
  return ` <span class="pill warn">Override ${Number(item.override_value || 0) ? "High" : "Low"}</span>`;
}

function renderDevices(data){
  const root = document.getElementById("devices");
  const cards = (data.device_catalog || []).map(device => {
    const mode = device.simulation ? "Simulation" : (device.reachable ? "Live" : "Offline");
    const modeCls = device.simulation ? "warn" : (device.reachable ? "ok" : "bad");
    const endpoint = device.port ? `${device.host}:${device.port}` : (device.host || "-");
    return `<div class="card">
      <div class="row" style="justify-content:space-between">
        <strong>${esc(device.device_label)}</strong>
        <span class="pill ${modeCls}">${esc(mode)}</span>
      </div>
      <div class="muted" style="margin-top:6px">${esc(device.device_code)}</div>
      <div class="muted" style="margin-top:6px">Endpoint: ${esc(endpoint)}</div>
      <div class="muted" style="margin-top:6px">Punkte: ${esc(device.total_points)} | Inputs: ${esc(device.inputs)} | Outputs: ${esc(device.outputs)} | GPIO: ${esc(device.gpio_points)}</div>
      <div class="muted" style="margin-top:6px">${esc(device.error || "")}</div>
    </div>`;
  });
  root.innerHTML = cards.join("");
}

function renderRows(data){
  if(!data){ return; }
  const showReserved = document.getElementById("show_reserved").checked;
  const rows = (data.points || []).filter(item => showReserved || !item.is_reserved).map(item => {
    const writable = canOverride(item);
    const overrideActive = !!item.override_active;
    const overrideValue = Number(item.override_value || 0);
    let actions = `<span class="muted">read-only</span>`;
    if(writable){
      actions = `<span class="io-action-set">
        <button class="btn small io-high ${overrideActive && overrideValue === 1 ? "active" : ""}" onclick="overrideIo('${item.io_key}',1)">High</button>
        <button class="btn small io-low ${overrideActive && overrideValue === 0 ? "active" : ""}" onclick="overrideIo('${item.io_key}',0)">Low</button>
        <button class="btn small io-release ${overrideActive ? "active" : ""}" onclick="releaseIo('${item.io_key}')">Release</button>
      </span>`;
    }else if(isPulseOnly(item)){
      actions = `<span class="muted">pulse/no override</span>`;
    }
    return `<tr>
      <td>${esc(item.device_label)}</td>
      <td><strong>${esc(item.pin_label)}</strong></td>
      <td>${esc(item.io_dir)}</td>
      <td>${esc(item.function_text || (item.is_reserved ? "Reserve" : ""))}</td>
      <td class="${valueClass(item.value)}">${esc(item.value)}${overrideLabel(item)}</td>
      <td class="${qualityClass(item.quality)}">${esc(item.quality)}</td>
      <td class="hide-mobile">${esc(item.zone_label || "-")}</td>
      <td>${actions}</td>
    </tr>`;
  });
  document.getElementById("rows").innerHTML = rows.join("") || '<tr><td colspan="8" class="muted">Keine IO-Punkte geladen</td></tr>';
}

function renderWorkbook(data){
  const wb = (data.master_workbook || {});
  const meta = wb.meta || {};
  document.getElementById("wb_path").textContent = wb.path || "-";
  document.getElementById("wb_meta").textContent = `exists=${wb.exists ? "yes" : "no"} | size=${wb.size_bytes || 0} B | channels=${meta.channel_count || 0}`;
}

function scheduleRefresh(data){
  const anyLive = (data.device_catalog || []).some(item => !item.simulation);
  const desiredMs = anyLive ? 2000 : 5000;
  if(refreshHandle){ clearInterval(refreshHandle); refreshHandle = null; }
  refreshHandle = setInterval(() => {
    if(document.hidden) return;
    reloadAll(true).catch(err => { document.getElementById("status").textContent = err.message; });
  }, desiredMs);
}

async function reloadAll(silent=false, live=false){
  if(!silent){ document.getElementById("status").textContent = "loading..."; }
  const data = await api(`/api/io/overview${live ? "?live=1" : ""}`);
  window.__io = data;
  window.__activeOverrides = (data.points || []).filter(item => !!item.override_active);
  renderWorkbook(data);
  renderDevices(data);
  renderRows(data);
  document.getElementById("status").textContent = `${(data.points || []).length} IO-Punkte geladen`;
  scheduleRefresh(data);
}

async function overrideIo(ioKey, value){
  document.getElementById("status").textContent = `override ${ioKey}=${value ? "High" : "Low"}...`;
  await api(`/api/io/${encodeURIComponent(ioKey)}/override`, {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({value})
  });
  await reloadAll(true);
}

async function releaseIo(ioKey){
  document.getElementById("status").textContent = `release ${ioKey}...`;
  await api(`/api/io/${encodeURIComponent(ioKey)}/release`, {method:"POST"});
  await reloadAll(true);
}

async function releaseAllOverrides(){
  document.getElementById("status").textContent = "gebe alle IO-Overrides frei...";
  await api("/api/io/overrides/release-all", {method:"POST"});
  await reloadAll(true);
}

async function handleNavigationWithOverrides(href){
  const count = (window.__activeOverrides || []).length;
  if(count > 0){
    const release = confirm(`${count} IO-Ausgang/Ausgaenge sind noch uebersteuert. OK = alle freigeben, Abbrechen = uebersteuert lassen.`);
    if(release){
      await releaseAllOverrides();
    }
  }
  window.__skipOverrideUnloadPrompt = true;
  window.location.href = href;
}

async function uploadWorkbook(file){
  if(!file) return;
  const fd = new FormData();
  fd.append("file", file);
  document.getElementById("status").textContent = "importiere IO Workbook...";
  await api("/api/io/import", {method:"POST", body: fd});
  document.getElementById("io_file").value = "";
  await reloadAll(true);
}

function downloadWorkbook(){
  window.location.href = "/api/io/master/download";
}

reloadAll().catch(err => { document.getElementById("status").textContent = err.message; });
document.addEventListener("click", (ev) => {
  const link = ev.target.closest ? ev.target.closest("a[href]") : null;
  if(!link || link.target || link.href.startsWith("javascript:")) return;
  if((window.__activeOverrides || []).length < 1) return;
  ev.preventDefault();
  handleNavigationWithOverrides(link.href).catch(err => { document.getElementById("status").textContent = err.message; });
});
window.addEventListener("beforeunload", (ev) => {
  if(!window.__skipOverrideUnloadPrompt && (window.__activeOverrides || []).length > 0){
    ev.preventDefault();
    ev.returnValue = "Es sind noch IO-Overrides aktiv. Beim Verlassen bleiben sie aktiv, wenn sie nicht freigegeben werden.";
  }
});
document.addEventListener("visibilitychange", () => {
  if(!document.hidden){
    reloadAll(true).catch(err => { document.getElementById("status").textContent = err.message; });
  }
});
</script>
</body>
</html>
        """.replace("__NAV__", nav)

    @app.get("/ui/params", response_class=HTMLResponse)
    def ui_params():
        nav = nav_html("params")
        return """
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>Params UI</title>
  <style>
    body{font-family:Segoe UI,Arial,sans-serif; margin:0; background:#f4f6f9; color:#1f2933}
    .wrap{max-width:1500px; margin:0 auto; padding:16px}
    .topnav{display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px}
    .navbtn{padding:8px 12px; border:1px solid #c2d2e4; border-radius:8px; background:#e8f0f8; color:#1f2933; text-decoration:none}
    .navbtn.active{background:#005eb8; color:#fff; border-color:#005eb8}
    .card{background:#fff; border:1px solid #d6dde7; border-radius:10px; padding:14px}
    .row{display:flex; gap:12px; align-items:flex-end; flex-wrap:wrap; margin-bottom:10px}
    .field{display:flex; flex-direction:column; gap:4px; min-width:140px; flex:0 1 180px}
    .field.grow{flex:1 1 320px}
    .field.small{flex:0 1 140px}
    .actions{display:flex; gap:8px; align-items:center; flex-wrap:wrap}
    .field label{font-size:12px; color:#5f6b7a; font-weight:600}
    input{padding:9px 10px; margin:0; border:1px solid #c8d6e5; border-radius:10px; background:#fff; min-height:38px; box-sizing:border-box}
    input[type=file]{padding:7px 9px; background:#f5f8fc}
    table{border-collapse:collapse; width:100%}
    th,td{border:1px solid #dbe2ea; padding:6px; font-size:13px}
    th{background:#f3f6fa; position:sticky; top:0}
    .btn{padding:9px 12px; border-radius:10px; border:1px solid #b9cde3; background:#e8f0f8; color:#17324b; font-weight:600; cursor:pointer}
    .btn:hover{background:#dce8f5}
    .btn:active{background:#cfe0f1}
    .muted{color:#666}
    .pill{padding:4px 8px; border:1px solid #b6c5d6; border-radius:999px; font-size:12px; background:#eef3f8}
  </style>
</head>
<body>
  <div class="wrap">
  __NAV__
  <div class="card">
  <h2>Parameter UI</h2>

  <div class="row">
    <div class="field grow">
      <label>Suche</label>
      <input id="q" placeholder="pkey / name / message"/>
    </div>
    <div class="field small">
      <label>ParamType</label>
      <input id="ptype" placeholder="z.B. TTP"/>
    </div>
    <div class="field grow">
      <label>Excel Import (.xlsx)</label>
      <input type="file" id="file" accept=".xlsx"/>
    </div>
    <div class="actions">
      <button class="btn" onclick="load()">Reload</button>
      <button class="btn" onclick="exportXlsx()">Export XLSX</button>
      <button class="btn" onclick="importXlsx()">Import XLSX</button>
      <button class="btn" onclick="downloadMaster()">Download Master XLSX</button>
      <span id="status" class="muted"></span>
    </div>
  </div>

  <div class="row">
    <span class="pill" id="masterInfo">Master workbook: loading...</span>
  </div>

  <h3>Liste</h3>
  <table>
    <thead>
      <tr>
        <th>pkey</th><th>min</th><th>max</th><th>default</th><th>rw</th><th>esp_rw</th>
        <th>current</th><th>effective</th><th>name</th><th>message</th><th>KI</th><th>edit</th>
      </tr>
    </thead>
    <tbody id="tbody"></tbody>
  </table>
  </div>
  </div>

<script>
const LS_KEY = "mas004_ui_token";

function lsGet(k){
  try { return localStorage.getItem(k) || ""; } catch(e){ return ""; }
}

function cookieGet(name){
  const m = document.cookie.match(new RegExp('(?:^|; )' + name.replace(/([.$?*|{}()\\[\\]\\\\\\/\\+^])/g, '\\\\$1') + '=([^;]*)'));
  return m ? decodeURIComponent(m[1]) : "";
}

function getToken(){
  return lsGet(LS_KEY) || cookieGet(LS_KEY) || "";
}

async function api(path, opt={}){
  opt.headers = opt.headers || {};
  const t = getToken();
  if(t) opt.headers["X-Token"] = t;

  const r = await fetch(path, opt);
  const txt = await r.text();
  let j=null; try{ j=JSON.parse(txt); }catch(e){}

  if(!r.ok){
    throw new Error((j && j.detail) ? j.detail : ("HTTP "+r.status+" "+txt));
  }
  return j;
}

async function load(){
  const q = document.getElementById("q").value.trim();
  const ptype = document.getElementById("ptype").value.trim();
  document.getElementById("status").textContent = "loading...";
  const url = `/api/params/list?limit=400&offset=0` + (q?`&q=${encodeURIComponent(q)}`:"") + (ptype?`&ptype=${encodeURIComponent(ptype)}`:"");
  const j = await api(url);
  const tb = document.getElementById("tbody");
  tb.innerHTML = "";
  for(const it of j.items){
    const tr = document.createElement("tr");
    tr.innerHTML = `
      <td>${it.pkey}</td>
      <td>${it.min_v ?? ""}</td>
      <td>${it.max_v ?? ""}</td>
      <td>${it.default_v ?? ""}</td>
      <td>${it.rw ?? ""}</td>
      <td>${it.esp_rw ?? ""}</td>
      <td>${it.current_v ?? ""}</td>
      <td>${it.effective_v ?? ""}</td>
      <td>${it.name ?? ""}</td>
      <td>${it.message ?? ""}</td>
      <td>${it.ai_instructions ?? ""}</td>
      <td><button class="btn" onclick="edit('${it.pkey}','${it.min_v ?? ""}','${it.max_v ?? ""}','${it.default_v ?? ""}','${it.rw ?? ""}','${it.esp_rw ?? ""}')">edit</button></td>
    `;
    tb.appendChild(tr);
  }
  await refreshMasterInfo();
  document.getElementById("status").textContent = `ok: ${j.items.length} items`;
}

async function edit(pkey, minv, maxv, defv, rw, espRw){
  const nmin = prompt(`min_v fuer ${pkey}`, minv);
  if(nmin === null) return;
  const nmax = prompt(`max_v fuer ${pkey}`, maxv);
  if(nmax === null) return;
  const ndef = prompt(`default_v fuer ${pkey}`, defv);
  if(ndef === null) return;
  const nrw = prompt(`rw fuer ${pkey} (R / W / R/W)`, rw);
  if(nrw === null) return;
  const nespRw = prompt(`esp_rw fuer ${pkey} (R / W / N)`, espRw);
  if(nespRw === null) return;

  const payload = {
    pkey: pkey,
    min_v: (nmin.trim()===""? null : Number(nmin)),
    max_v: (nmax.trim()===""? null : Number(nmax)),
    default_v: (ndef.trim()===""? null : ndef),
    rw: (nrw.trim()===""? null : nrw),
    esp_rw: (nespRw.trim()===""? null : nespRw)
  };

  document.getElementById("status").textContent = "saving...";
  await api("/api/params/edit", {
    method: "POST",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify(payload)
  });
  await load();
}

async function importXlsx(){
  const f = document.getElementById("file").files[0];
  if(!f){ alert("Bitte .xlsx auswaehlen"); return; }
  document.getElementById("status").textContent = "importing...";
  const fd = new FormData();
  fd.append("file", f);
  const t = getToken();
  const r = await fetch("/api/params/import", {method:"POST", body: fd, headers: t?{"X-Token":t}:{}} );
  const txt = await r.text();
  if(!r.ok){ alert("Import Fehler: " + txt); return; }
  document.getElementById("status").textContent = "import ok";
  await load();
}

async function refreshMasterInfo(){
  try{
    const j = await api("/api/params/master/info");
    const m = j.master_workbook || {};
    document.getElementById("masterInfo").textContent =
      m.exists
        ? `Master workbook: ${m.path} | ${m.mtime_iso || "-"} | ${m.size_bytes || 0} bytes`
        : `Master workbook: nicht auf Raspi gespeichert (${m.path || "-"})`;
  }catch(e){
    document.getElementById("masterInfo").textContent = `Master workbook: Fehler - ${e.message}`;
  }
}

function downloadMaster(){
  (async ()=>{
    const t = getToken();
    const r = await fetch("/api/params/master/download", {headers: t?{"X-Token":t}:{}} );
    if(!r.ok){ alert("Master Download Fehler: " + await r.text()); return; }
    const blob = await r.blob();
    const a = document.createElement("a");
    a.href = URL.createObjectURL(blob);
    a.download = "Parameterliste_master.xlsx";
    a.click();
    URL.revokeObjectURL(a.href);
  })();
}

function exportXlsx(){
  const q = document.getElementById("q").value.trim();
  const ptype = document.getElementById("ptype").value.trim();
  let url = "/api/params/export" + (q||ptype ? "?" : "");
  if(q) url += "q=" + encodeURIComponent(q) + "&";
  if(ptype) url += "ptype=" + encodeURIComponent(ptype) + "&";
  url = url.replace(/[&?]$/, "");

  (async ()=>{
    const t = getToken();
    const r = await fetch(url, {headers: t?{"X-Token":t}:{}} );
    if(!r.ok){ alert("Export Fehler: " + await r.text()); return; }
    const blob = await r.blob();
    const a = document.createElement("a");
    a.href = URL.createObjectURL(blob);
    a.download = "params_export.xlsx";
    a.click();
    URL.revokeObjectURL(a.href);
  })();
}

load();
</script>
</body>
</html>
        """.replace("__NAV__", nav)

    @app.get("/ui/motors", include_in_schema=False)
    def ui_motors_redirect():
        return RedirectResponse(url="/ui/machine-setup/motors", status_code=303)

    @app.get("/ui/machine-setup/motors", response_class=HTMLResponse, include_in_schema=False)
    def ui_motors(request: Request):
        cfg2 = Settings.load(cfg_path)
        if not has_machine_setup_session(request, cfg2):
            target = "/ui/machine-setup/motors"
            return RedirectResponse(url=f"/ui/machine-setup/login?next={target}", status_code=303)
        nav = nav_html("machine_setup") + machine_setup_nav_html("motors")
        return """
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>Motor Setup</title>
  <style>
    :root{
      --bg:#f4f6f9; --card:#fff; --text:#1f2933; --muted:#5f6b7a; --border:#d6dde7; --blue:#005eb8;
      --good:#2e7d32; --warn:#ed6c02; --bad:#c62828;
    }
    *{box-sizing:border-box}
    body{margin:0; font-family:Segoe UI,Arial,sans-serif; background:var(--bg); color:var(--text)}
    .wrap{max-width:1600px; margin:0 auto; padding:16px}
    .topnav{display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px}
    .navbtn{padding:8px 12px; border:1px solid #c2d2e4; border-radius:8px; background:#e8f0f8; color:#1f2933; text-decoration:none}
    .navbtn.active{background:#005eb8; color:#fff; border-color:#005eb8}
    .toolbar,.row,.actions{display:flex; gap:10px; align-items:center; flex-wrap:wrap}
    .toolbar{margin-bottom:12px}
    .muted{color:var(--muted)}
    .status-panel{display:flex; flex-direction:column; gap:4px; min-width:min(760px,100%)}
    #status{font-weight:600}
    #status-log{max-height:86px; overflow:auto; font-size:12px; line-height:1.35; color:#334155; background:#f8fafc; border:1px solid var(--border); border-radius:8px; padding:6px 8px; min-width:min(760px,100%)}
    #status-log div{white-space:nowrap; overflow:hidden; text-overflow:ellipsis}
    .btn{min-height:38px; padding:8px 12px; border:1px solid #aec4db; border-radius:10px; background:#e8f0f8; color:#17324b; font-weight:600; cursor:pointer}
    .btn:hover{background:#dce8f5}
    .btn:disabled{opacity:.55; cursor:not-allowed}
    .cards{display:grid; grid-template-columns:repeat(auto-fit,minmax(470px,1fr)); gap:12px}
    .card{background:var(--card); border:1px solid var(--border); border-radius:12px; padding:14px}
    .card h3{margin:0 0 6px 0}
    .meta{display:flex; gap:8px; flex-wrap:wrap; margin-bottom:10px}
    .pill{display:inline-flex; align-items:center; gap:8px; padding:6px 10px; border-radius:999px; border:1px solid var(--border); background:#eef3f8; font-size:12px}
    .ok{color:var(--good)} .warn{color:var(--warn)} .bad{color:var(--bad)}
    .grid{display:grid; gap:10px; grid-template-columns:repeat(3,minmax(0,1fr))}
    .field{display:flex; flex-direction:column; gap:4px}
    .field label{font-size:12px; color:var(--muted); font-weight:600}
    input,select{width:100%; min-height:38px; padding:9px 10px; border:1px solid var(--border); border-radius:10px; background:#fff}
    input:focus,select:focus{outline:none; border-color:var(--blue); box-shadow:0 0 0 3px rgba(0,94,184,.12)}
    .span-2{grid-column:span 2}
    .span-3{grid-column:span 3}
    .status-grid{display:grid; gap:8px; grid-template-columns:repeat(2,minmax(0,1fr)); margin-bottom:10px}
    .status-box{background:#f8fafc; border:1px solid var(--border); border-radius:10px; padding:10px}
    .status-box strong{display:block; margin-bottom:6px}
    .status-box code{font-size:12px; white-space:pre-wrap; word-break:break-word}
    .section-title{margin:12px 0 8px 0; font-size:14px; font-weight:700}
    @media(max-width:980px){ .grid,.status-grid{grid-template-columns:1fr} .span-2,.span-3{grid-column:auto} }
  </style>
</head>
<body>
  <div class="wrap">
    __NAV__
    <div class="toolbar">
      <button class="btn" onclick="reloadAll()">Reload</button>
      <button class="btn" onclick="refreshAllLive()">Live-Status aktualisieren</button>
      <div class="status-panel">
        <span id="status" class="muted">loading...</span>
        <div id="status-log"></div>
      </div>
    </div>
    <div class="cards" id="cards"></div>
  </div>
<script>
const TOKEN_KEY = "mas004_ui_token";
const dirtyFields = new Set();
const renderedIds = new Set();
const motorPollHandles = new Map();
const motorActionInFlight = new Set();
const motorRefreshInFlight = new Set();
let autoRefreshHandle = null;
let currentAutoRefreshMs = null;
let statusHoldUntilMs = 0;

function token(){ try { return localStorage.getItem(TOKEN_KEY) || ""; } catch(e){ return ""; } }
function fieldKey(id, name){ return `${id}:${name}`; }
function esc(v){ return String(v ?? "").replaceAll("&","&amp;").replaceAll("<","&lt;").replaceAll(">","&gt;").replaceAll('"',"&quot;"); }
function numOrNull(v){ const t = String(v ?? "").trim(); return t === "" ? null : Number(t); }
function boolFromInput(id){ return document.getElementById(id).value === "1"; }

async function api(path, opt={}){
  opt.headers = opt.headers || {};
  const t = token();
  if(t) opt.headers["X-Token"] = t;
  const r = await fetch(path, opt);
  const txt = await r.text();
  let j = null; try { j = JSON.parse(txt); } catch(e) {}
  if(!r.ok){
    const detail = j && j.detail !== undefined ? j.detail : (`HTTP ${r.status} ${txt}`);
    throw new Error(typeof detail === "string" ? detail : JSON.stringify(detail));
  }
  return j;
}

function bindDirty(input, motorId, field){
  const key = fieldKey(motorId, field);
  ["input","change"].forEach(evt => input.addEventListener(evt, () => dirtyFields.add(key)));
}

function setInputValueIfClean(motorId, field, value){
  const el = document.getElementById(`m-${motorId}-${field}`);
  if(!el) return;
  const key = fieldKey(motorId, field);
  if(dirtyFields.has(key) || document.activeElement === el){ return; }
  el.value = value ?? "";
}

function statusKind(v){
  if(v === true || v === 1 || v === "1"){ return "ok"; }
  if(v === false || v === 0 || v === "0"){ return "muted"; }
  return "warn";
}

function tenthsToMmText(value){
  const n = Number(value);
  if(!Number.isFinite(n)){ return ""; }
  return String((n / 10).toFixed(3)).replace(/0+$/,"").replace(/\\.$/,"");
}

function setMotorSimulationUi(id, enabled){
  if(enabled){
    stopMotorPolling(id);
    const pollEl = document.getElementById(`m-${id}-poll1s`);
    if(pollEl){ pollEl.checked = false; }
  }
  document.querySelectorAll(`[data-motor-id="${id}"][data-live-only="1"]`).forEach(el => {
    el.disabled = !!enabled;
  });
}

function setStatus(text, opts = {}){
  const now = Date.now();
  if(opts.soft && now < statusHoldUntilMs){ return; }
  const value = text || "";
  const statusEl = document.getElementById("status");
  if(statusEl){ statusEl.textContent = value; }
  const holdMs = Number(opts.holdMs || 0);
  if(holdMs > 0){ statusHoldUntilMs = now + holdMs; }
  if(value && opts.log !== false){
    const log = document.getElementById("status-log");
    if(log){
      const row = document.createElement("div");
      row.textContent = `${new Date().toLocaleTimeString()} ${value}`;
      log.prepend(row);
      while(log.childElementCount > 8){ log.removeChild(log.lastElementChild); }
    }
  }
}

function setMotorControlsBusy(id, busy){
  const simEl = document.getElementById(`m-${id}-simulation`);
  const simulated = !!(simEl && simEl.checked);
  document.querySelectorAll(`button[data-motor-id="${id}"][data-live-only="1"]`).forEach(el => {
    el.disabled = !!busy || simulated;
  });
}

function renderCard(motor){
  const bindings = motor.bindings || {};
  const related = (bindings.related_params || []).map(it => esc(it.pkey)).join(", ");
  const cfg = motor.config || {};
  const state = motor.state || {};
  const card = document.createElement("div");
  card.className = "card";
  card.id = `motor-card-${motor.id}`;
    card.innerHTML = `
      <h3>${esc(motor.name || ("Motor " + motor.id))}</h3>
      <div class="muted" style="margin-bottom:8px">${esc(motor.description || "")}</div>
      <div class="meta">
        <span class="pill">ID ${esc(motor.id)}</span>
      <span class="pill">${esc(motor.controller || "")}</span>
      <span class="pill">${esc(motor.motor_type || "")}</span>
      <span class="pill">${motor.positional ? "Positionierachse" : "Transportachse"}</span>
      <label class="pill"><input type="checkbox" id="m-${motor.id}-simulation" onchange="toggleSimulation(${motor.id}, this.checked)"/>Simulation</label>
      <label class="pill"><input type="checkbox" id="m-${motor.id}-poll1s" data-motor-id="${motor.id}" data-live-only="1" onchange="toggleMotorPolling(${motor.id}, this.checked)"/>1s Polling</label>
    </div>
    <div class="status-grid">
      <div class="status-box">
        <strong>Live</strong>
        <div>Link: <span id="live-${motor.id}-link" class="${statusKind(state.link_ok)}">${esc(state.link_ok)}</span></div>
        <div>Ready: <span id="live-${motor.id}-ready" class="${statusKind(state.ready)}">${esc(state.ready)}</span></div>
        <div>Move: <span id="live-${motor.id}-move" class="${statusKind(state.move)}">${esc(state.move)}</span></div>
        <div>InPos: <span id="live-${motor.id}-inpos" class="${statusKind(state.in_pos)}">${esc(state.in_pos)}</span></div>
        <div>Alarm: <span id="live-${motor.id}-alarm" class="${statusKind(!state.alarm ? 1 : 0)}">${esc(state.alarm_code ?? state.alarm ?? "")}</span></div>
      </div>
      <div class="status-box">
        <strong>Position / IO</strong>
        <div>Ist: <span id="live-${motor.id}-pos">${esc(state.feedback_tenths_mm ?? "")}</span> 1/10mm</div>
        <div>Command: <span id="live-${motor.id}-cmd">${esc(state.command_tenths_mm ?? "")}</span> 1/10mm</div>
        <div>Input raw: <span id="live-${motor.id}-inraw">${esc(state.input_raw_hex ?? "")}</span></div>
        <div>Output raw: <span id="live-${motor.id}-outraw">${esc(state.output_raw_hex ?? "")}</span></div>
      </div>
    </div>
    <div class="section-title">Kalibrierung / Parameter</div>
    <div class="grid">
      <div class="field"><label>Test-Schritte</label><input id="m-${motor.id}-test_steps" data-motor-id="${motor.id}" data-live-only="1" value="1000"/></div>
      <div class="field"><label>Manuell bewegen [mm]</label><input id="m-${motor.id}-manual_mm" data-motor-id="${motor.id}" data-live-only="1" value="0"/></div>
      <div class="field"><label>Istposition setzen [mm]</label><input id="m-${motor.id}-position_mm" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Absolut fahren nach [mm]</label><input id="m-${motor.id}-absolute_mm" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Steps / mm</label><input id="m-${motor.id}-steps_per_mm" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Geschwindigkeit [mm/s]</label><input id="m-${motor.id}-speed_mm_s" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Acceleration [mm/s2]</label><input id="m-${motor.id}-accel_mm_s2" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Deceleration [mm/s2]</label><input id="m-${motor.id}-decel_mm_s2" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Fahrstrom [%]</label><input id="m-${motor.id}-current_pct" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Haltestrom [%]</label><input id="m-${motor.id}-hold_current_pct" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Drehrichtung</label><select id="m-${motor.id}-invert_direction" data-motor-id="${motor.id}" data-live-only="1"><option value="0">Normal</option><option value="1">Invertiert</option></select></div>
      <div class="field"><label>Min [1/10mm]</label><input id="m-${motor.id}-min_tenths_mm" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Max [1/10mm]</label><input id="m-${motor.id}-max_tenths_mm" data-motor-id="${motor.id}" data-live-only="1"/></div>
      <div class="field"><label>Min aktiv</label><select id="m-${motor.id}-min_enabled" data-motor-id="${motor.id}" data-live-only="1"><option value="0">Nein</option><option value="1">Ja</option></select></div>
      <div class="field"><label>Max aktiv</label><select id="m-${motor.id}-max_enabled" data-motor-id="${motor.id}" data-live-only="1"><option value="0">Nein</option><option value="1">Ja</option></select></div>
      <div class="field span-3"><label>Zuordnung aus Excel</label><input disabled value="${esc(related)}"/></div>
    </div>
    <div class="actions" style="margin-top:10px">
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="moveSteps(${motor.id})">Schritte fahren</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="defineResolution(${motor.id})">Auflösung definieren</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="manualMove(${motor.id})">Move mm</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="setAbsolutePosition(${motor.id})">Istposition übernehmen</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="moveAbsolute(${motor.id})">Absolut fahren</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="setZero(${motor.id})">Nullpunkt setzen</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="setMin(${motor.id})">Min setzen</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="setMax(${motor.id})">Max setzen</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="refreshOne(${motor.id})">Status aktualisieren</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="resetAlarm(${motor.id})">Alarm Reset</button>
      <button class="btn" data-motor-id="${motor.id}" data-live-only="1" onclick="saveConfig(${motor.id})">Parameter speichern</button>
    </div>
    <div class="section-title">Meldungen</div>
    <div class="status-box"><code id="live-${motor.id}-msg">${esc(state.last_error || motor.last_reply || "-")}</code></div>
  `;
  document.getElementById("cards").appendChild(card);

  ["steps_per_mm","speed_mm_s","accel_mm_s2","decel_mm_s2","current_pct","hold_current_pct","invert_direction","min_tenths_mm","max_tenths_mm","min_enabled","max_enabled","test_steps","manual_mm","position_mm","absolute_mm"].forEach(field => {
    const el = document.getElementById(`m-${motor.id}-${field}`);
    if(el) bindDirty(el, motor.id, field);
  });
  renderedIds.add(motor.id);
  updateCard(motor);
}

function updateLiveText(id, suffix, value, cls){
  const el = document.getElementById(`live-${id}-${suffix}`);
  if(!el) return;
  el.textContent = value ?? "";
  if(cls !== undefined){ el.className = cls; }
}

function updateCard(motor){
  const cfg = motor.config || {};
  const state = motor.state || {};
  const sim = !!motor.simulation;
  setInputValueIfClean(motor.id, "steps_per_mm", cfg.steps_per_mm ?? "");
  setInputValueIfClean(motor.id, "speed_mm_s", cfg.speed_mm_s ?? "");
  setInputValueIfClean(motor.id, "accel_mm_s2", cfg.accel_mm_s2 ?? "");
  setInputValueIfClean(motor.id, "decel_mm_s2", cfg.decel_mm_s2 ?? "");
  setInputValueIfClean(motor.id, "current_pct", cfg.current_pct ?? "");
  setInputValueIfClean(motor.id, "hold_current_pct", cfg.hold_current_pct ?? "");
  setInputValueIfClean(motor.id, "invert_direction", cfg.invert_direction ? "1" : "0");
  setInputValueIfClean(motor.id, "min_tenths_mm", cfg.min_tenths_mm ?? "");
  setInputValueIfClean(motor.id, "max_tenths_mm", cfg.max_tenths_mm ?? "");
  setInputValueIfClean(motor.id, "min_enabled", cfg.min_enabled ? "1" : "0");
  setInputValueIfClean(motor.id, "max_enabled", cfg.max_enabled ? "1" : "0");
  setInputValueIfClean(motor.id, "position_mm", tenthsToMmText(state.feedback_tenths_mm));
  setInputValueIfClean(motor.id, "absolute_mm", tenthsToMmText(state.target_tenths_mm ?? state.command_tenths_mm ?? state.feedback_tenths_mm));

  updateLiveText(motor.id, "link", String(state.link_ok ?? ""), statusKind(state.link_ok));
  updateLiveText(motor.id, "ready", String(state.ready ?? ""), statusKind(state.ready));
  updateLiveText(motor.id, "move", String(state.move ?? ""), statusKind(state.move));
  updateLiveText(motor.id, "inpos", String(state.in_pos ?? ""), statusKind(state.in_pos));
  updateLiveText(motor.id, "alarm", String(state.alarm_code ?? state.alarm ?? ""), statusKind(!state.alarm ? 1 : 0));
  updateLiveText(motor.id, "pos", state.feedback_tenths_mm ?? "");
  updateLiveText(motor.id, "cmd", state.command_tenths_mm ?? "");
  updateLiveText(motor.id, "inraw", state.input_raw_hex ?? "");
  updateLiveText(motor.id, "outraw", state.output_raw_hex ?? "");
  updateLiveText(motor.id, "msg", state.last_error || motor.last_reply || "-");
  const simEl = document.getElementById(`m-${motor.id}-simulation`);
  if(simEl){ simEl.checked = sim; }
  setMotorSimulationUi(motor.id, sim);
}

async function toggleSimulation(id, enabled){
  setStatus(`simulation motor ${id}...`);
  await post(`/api/motors/${id}/simulation`, {enabled: !!enabled});
  await reloadAll();
}

async function post(path, body){
  const result = await api(path, {method:"POST", headers:{"Content-Type":"application/json"}, body: body ? JSON.stringify(body) : null});
  if(result && result.ok === false){
    throw new Error(result.detail || result.reply || result.error || "Command rejected");
  }
  return result;
}

function applyMotorResult(id, result){
  if(!result){ return; }
  const motor = result.motor || (result.refresh && result.refresh.motor);
  if(motor){ updateCard(motor); }
  if(result.reply){ updateLiveText(id, "msg", result.reply); }
  if(result.refresh_error){ updateLiveText(id, "msg", `ACK, Refresh: ${result.refresh_error}`); }
}

async function runMotorAction(id, label, action){
  if(motorActionInFlight.has(id)){
    setStatus(`Motor ${id}: Aktion laeuft bereits`, {holdMs:5000});
    return null;
  }
  motorActionInFlight.add(id);
  setMotorControlsBusy(id, true);
  setStatus(`Motor ${id}: ${label}...`, {holdMs:3000});
  try{
    const result = await action();
    applyMotorResult(id, result);
    const detail = result && result.refresh_error ? `, Refresh: ${result.refresh_error}` : "";
    setStatus(`Motor ${id}: ${label} OK${detail}`, {holdMs:10000});
    return result || {ok:true};
  }catch(e){
    const msg = e && e.message ? e.message : String(e);
    updateLiveText(id, "msg", msg);
    setStatus(`Motor ${id}: ${label} fehlgeschlagen: ${msg}`, {holdMs:15000});
    return null;
  }finally{
    motorActionInFlight.delete(id);
    setMotorControlsBusy(id, false);
  }
}

function configPayload(id){
  const payload = {
    steps_per_mm: numOrNull(document.getElementById(`m-${id}-steps_per_mm`).value),
    speed_mm_s: numOrNull(document.getElementById(`m-${id}-speed_mm_s`).value),
    accel_mm_s2: numOrNull(document.getElementById(`m-${id}-accel_mm_s2`).value),
    decel_mm_s2: numOrNull(document.getElementById(`m-${id}-decel_mm_s2`).value),
    current_pct: numOrNull(document.getElementById(`m-${id}-current_pct`).value),
    hold_current_pct: numOrNull(document.getElementById(`m-${id}-hold_current_pct`).value),
    invert_direction: boolFromInput(`m-${id}-invert_direction`),
    min_tenths_mm: numOrNull(document.getElementById(`m-${id}-min_tenths_mm`).value),
    max_tenths_mm: numOrNull(document.getElementById(`m-${id}-max_tenths_mm`).value),
    min_enabled: boolFromInput(`m-${id}-min_enabled`),
    max_enabled: boolFromInput(`m-${id}-max_enabled`)
  };
  return payload;
}

async function saveConfig(id){
  const saveResult = await runMotorAction(id, "Parameter speichern", () => post(`/api/motors/${id}/config`, configPayload(id)));
  if(!saveResult){ return; }
  ["steps_per_mm","speed_mm_s","accel_mm_s2","decel_mm_s2","current_pct","hold_current_pct","invert_direction","min_tenths_mm","max_tenths_mm","min_enabled","max_enabled","position_mm"].forEach(f => dirtyFields.delete(fieldKey(id, f)));
  setStatus(`Motor ${id}: Parameter gespeichert`, {holdMs:10000});
}

function askDirectionCorrect(id){
  return new Promise(resolve => {
    const existing = document.getElementById("direction-modal");
    if(existing){ existing.remove(); }
    const overlay = document.createElement("div");
    overlay.id = "direction-modal";
    overlay.style.cssText = "position:fixed; inset:0; z-index:9999; background:rgba(15,23,42,.36); display:flex; align-items:center; justify-content:center; padding:18px;";
    overlay.innerHTML = `
      <div style="background:#fff; border-radius:18px; width:min(460px, 96vw); padding:22px; box-shadow:0 28px 80px rgba(15,23,42,.28); border:1px solid #d8e1ee;">
        <h3 style="margin:0 0 10px 0;">Motor ${esc(id)}: Richtung pruefen</h3>
        <p style="margin:0 0 18px 0; color:#334155; line-height:1.45;">War die Bewegungsrichtung mechanisch korrekt?</p>
        <div style="display:flex; gap:10px; flex-wrap:wrap; justify-content:flex-end;">
          <button class="btn" data-answer="yes">Ja, korrekt</button>
          <button class="btn" data-answer="no" style="background:#fee2e2; border-color:#fecaca; color:#7f1d1d;">Nein, Richtung drehen</button>
          <button class="btn" data-answer="cancel" style="background:#f1f5f9;">Abbrechen</button>
        </div>
      </div>`;
    function finish(value){
      overlay.remove();
      resolve(value);
    }
    overlay.addEventListener("click", ev => {
      if(ev.target === overlay){ finish(null); }
      const btn = ev.target.closest ? ev.target.closest("button[data-answer]") : null;
      if(!btn){ return; }
      const answer = btn.getAttribute("data-answer");
      if(answer === "yes"){ finish(true); }
      else if(answer === "no"){ finish(false); }
      else { finish(null); }
    });
    document.body.appendChild(overlay);
  });
}

async function moveSteps(id){
  const value = numOrNull(document.getElementById(`m-${id}-test_steps`).value);
  if(value === null){ alert("Bitte Schrittzahl eingeben."); return; }
  await runMotorAction(id, `Schritte fahren ${value}`, () => post(`/api/motors/${id}/move`, {mode:"relative_steps", value:value}));
}

async function defineResolution(id){
  const stepValue = numOrNull(document.getElementById(`m-${id}-test_steps`).value);
  if(stepValue === null || stepValue === 0){ alert("Bitte eine Schrittzahl ungleich 0 eingeben."); return; }
  const moveResult = await runMotorAction(id, `Schritte fahren ${stepValue}`, () => post(`/api/motors/${id}/move`, {mode:"relative_steps", value:stepValue}));
  if(!moveResult){ return; }
  const mmRaw = prompt("Wie viele mm hat sich der Motor bewegt?");
  if(mmRaw === null) return;
  const measured = Number(mmRaw);
  if(!Number.isFinite(measured) || measured <= 0){ alert("Ungueltiger mm-Wert."); return; }
  const directionOk = await askDirectionCorrect(id);
  if(directionOk === null){ setStatus(`Motor ${id}: Aufloesungskalibrierung abgebrochen`); return; }
  const stepsPerMm = Math.abs(stepValue) / measured;
  const spmm = document.getElementById(`m-${id}-steps_per_mm`);
  spmm.value = String(stepsPerMm.toFixed(6)).replace(/0+$/,"").replace(/\\.$/,"");
  dirtyFields.add(fieldKey(id, "steps_per_mm"));
  if(!directionOk){
    const dir = document.getElementById(`m-${id}-invert_direction`);
    dir.value = dir.value === "1" ? "0" : "1";
    dirtyFields.add(fieldKey(id, "invert_direction"));
  }
  setStatus(`Motor ${id}: neue Aufloesung berechnet, speichere persistent...`);
  await saveConfig(id);
}

async function manualMove(id){
  const value = numOrNull(document.getElementById(`m-${id}-manual_mm`).value);
  if(value === null){ alert("Bitte mm-Wert eingeben."); return; }
  await runMotorAction(id, `Move ${value} mm`, () => post(`/api/motors/${id}/move`, {mode:"relative_mm", value:value}));
}

async function setAbsolutePosition(id){
  const value = numOrNull(document.getElementById(`m-${id}-position_mm`).value);
  if(value === null){ alert("Bitte absolute Istposition in mm eingeben."); return; }
  const result = await runMotorAction(id, `Istposition = ${value} mm übernehmen`, () => post(`/api/motors/${id}/position`, {value:value}));
  if(result){ dirtyFields.delete(fieldKey(id, "position_mm")); }
}

async function moveAbsolute(id){
  const value = numOrNull(document.getElementById(`m-${id}-absolute_mm`).value);
  if(value === null){ alert("Bitte absolute Zielposition in mm eingeben."); return; }
  await runMotorAction(id, `Absolut fahren ${value} mm`, () => post(`/api/motors/${id}/move`, {mode:"absolute_mm", value:value}));
}

async function setZero(id){ await runMotorAction(id, "Nullpunkt setzen", () => post(`/api/motors/${id}/zero`, {})); }
async function setMin(id){ await runMotorAction(id, "Min setzen", () => post(`/api/motors/${id}/min`, {})); }
async function setMax(id){ await runMotorAction(id, "Max setzen", () => post(`/api/motors/${id}/max`, {})); }
async function resetAlarm(id){ await runMotorAction(id, "Alarm Reset", () => post(`/api/motors/${id}/reset-alarm`, {})); }
async function refreshOne(id, opts = {}){
  const silent = !!opts.silent;
  if(motorActionInFlight.has(id)){
    if(!silent){ setStatus(`Motor ${id}: Aktion laeuft, Refresh uebersprungen`, {holdMs:5000}); }
    return null;
  }
  if(motorRefreshInFlight.has(id)){
    if(!silent){ setStatus(`Motor ${id}: Refresh laeuft bereits`, {holdMs:5000}); }
    return null;
  }
  motorRefreshInFlight.add(id);
  if(!silent){ setStatus(`Motor ${id}: Status aktualisieren...`, {holdMs:3000}); }
  try{
    const result = await post(`/api/motors/${id}/refresh`, {});
    applyMotorResult(id, result);
    if(!silent){
      const suffix = result && result.refresh_in_progress ? " (alter Wert, Refresh laeuft)" : (result && result.cached ? " (Cache)" : "");
      setStatus(`Motor ${id}: Status aktualisiert${suffix}`, {holdMs:8000});
    }
    return result;
  }catch(e){
    const msg = e && e.message ? e.message : String(e);
    updateLiveText(id, "msg", msg);
    if(!silent){ setStatus(`Motor ${id}: Status fehlgeschlagen: ${msg}`, {holdMs:12000}); }
    return null;
  }finally{
    motorRefreshInFlight.delete(id);
  }
}

async function refreshAllLive(){
  const ids = Array.from(renderedIds).sort((a,b) => Number(a) - Number(b));
  for(const id of ids){
    const simEl = document.getElementById(`m-${id}-simulation`);
    if(simEl && simEl.checked){ continue; }
    if(motorActionInFlight.has(id) || motorRefreshInFlight.has(id)){ continue; }
    setStatus(`refresh motor ${id}/${ids.length}...`);
    try { await refreshOne(id, {silent:true}); }
    catch(e){ console.warn(e); }
  }
  await reloadAll({silent:true});
}

function stopMotorPolling(id){
  const handle = motorPollHandles.get(id);
  if(handle){ clearInterval(handle); }
  motorPollHandles.delete(id);
}

function toggleMotorPolling(id, enabled){
  stopMotorPolling(id);
  if(!enabled){
    setStatus(`Motor ${id}: 1s Polling aus`);
    return;
  }
  setStatus(`Motor ${id}: 1s Polling ein`);
  refreshOne(id, {silent:true});
  const handle = setInterval(() => {
    if(document.hidden){ return; }
    if(motorActionInFlight.has(id) || motorRefreshInFlight.has(id)){ return; }
    refreshOne(id, {silent:true}).catch(err => { console.warn(err); });
  }, 1000);
  motorPollHandles.set(id, handle);
}

function scheduleAutoRefresh(ms){
  if(autoRefreshHandle){
    clearInterval(autoRefreshHandle);
    autoRefreshHandle = null;
  }
  currentAutoRefreshMs = ms;
  if(!ms || ms <= 0){ return; }
  autoRefreshHandle = setInterval(() => {
    reloadAll({silent:true}).catch(err => { setStatus(err.message); });
  }, ms);
}

async function reloadAll(opts = {}){
  const silent = !!opts.silent;
  if(!silent){ setStatus("loading..."); }
  const data = await api("/api/motors/overview");
  for(const motor of (data.motors || [])){
    if(!renderedIds.has(motor.id)){ renderCard(motor); }
    else { updateCard(motor); }
  }
  const allSimulated = (data.motors || []).length > 0 && (data.motors || []).every(m => !!m.simulation);
  const suffix = data.message ? ` | ${data.message}` : (allSimulated ? " | nur Simulation" : "");
  setStatus(`${(data.motors || []).length} Motoren geladen${suffix}`, {soft:silent, log:!silent});
  const desiredMs = 0;
  if(currentAutoRefreshMs !== desiredMs){
    scheduleAutoRefresh(desiredMs);
  }
}

window.addEventListener("beforeunload", () => {
  for(const id of Array.from(motorPollHandles.keys())){
    stopMotorPolling(id);
  }
});

reloadAll().catch(err => { setStatus(err.message); });
</script>
</body>
</html>
        """.replace("__NAV__", nav)

    # -----------------------------
    # Settings UI
    # -----------------------------
    @app.get("/ui/settings", response_class=HTMLResponse)
    def ui_settings():
        nav = nav_html("settings")
        return """
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>System Settings</title>
  <style>

  :root{
    --blue:#005eb8;
    --red:#c62828;
    --bg:#f4f6f9;
    --card:#ffffff;
    --text:#1f2933;
    --muted:#5f6b7a;
    --border:#d6dde7;
    --radius:10px;
    --shadow:none;
  }
  *, *::before, *::after{box-sizing:border-box;}
  body{
    margin:0;
    font-family:Segoe UI,Arial,sans-serif;
    background:var(--bg);
    color:var(--text);
  }
  .wrap{max-width:1500px; margin:0 auto; padding:16px}
  .grid{display:grid; gap:12px; align-items:end; justify-content:start; margin-bottom:10px;}
  .token-grid{grid-template-columns:minmax(320px,760px) auto;}
  .cols-4{grid-template-columns:220px 220px 110px 220px;}
  .cols-5{grid-template-columns:220px 220px 110px 220px 320px;}
  .cols-3a{grid-template-columns:460px 460px 220px 200px;}
  .cols-3b{grid-template-columns:160px 160px 260px;}
  .cols-ntp{grid-template-columns:420px 220px;}
  .cols-device{grid-template-columns:230px 120px 280px 240px auto;}
  .cols-log{grid-template-columns:180px 180px 180px 180px;}
  .field{display:flex; flex-direction:column; gap:4px; min-width:0;}
  .field label{font-size:12px; color:var(--muted); font-weight:600;}
  .field.empty{visibility:hidden;}
  .actions{display:flex; gap:8px; align-items:center; flex-wrap:wrap; margin-top:2px;}
  .checkline{
    display:inline-flex;
    align-items:center;
    gap:8px;
    min-height:38px;
    padding:0 4px;
    color:var(--text);
    font-size:14px;
  }
  input,select,textarea{
    width:100%;
    padding:10px 12px;
    border:1px solid var(--border);
    border-radius:12px;
    background:#fff;
    font-size:14px;
    outline:none;
  }
  .checkline input{
    width:18px;
    height:18px;
    margin:0;
    border-radius:4px;
  }
  textarea{min-height:110px; font-family:ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;}
  input:focus,select:focus,textarea:focus{border-color:var(--blue); box-shadow:0 0 0 3px rgba(0,94,184,.15);}
  button{
    min-height:38px;
    padding:8px 14px;
    border-radius:10px;
    border:1px solid #a8bfd8;
    background:#e8f0f8;
    color:#17324b;
    font-weight:600;
    cursor:pointer;
    box-shadow:inset 0 1px 0 rgba(255,255,255,.65);
  }
  button:hover{background:#d9e7f5; border-color:#96b2cf;}
  button:active{background:#cfe0f1; border-color:#8aa9c9;}
  button:focus-visible{outline:none; box-shadow:0 0 0 3px rgba(0,94,184,.2);}
  .pill{
    display:inline-flex;
    align-items:center;
    gap:8px;
    padding:8px 10px;
    border:1px solid var(--border);
    border-radius:999px;
    font-size:13px;
    background:#eef3f8;
  }
  .muted{color:var(--muted);}
  fieldset{
    background:var(--card);
    border:1px solid var(--border);
    border-radius:var(--radius);
    box-shadow:var(--shadow);
    padding:14px;
    margin:12px 0;
  }
  legend{padding:0 6px; font-weight:600;}
  pre{
    background:#f8fafc;
    border:1px solid var(--border);
    border-radius:8px;
    padding:10px;
    overflow:auto;
    white-space:pre-wrap;
    word-break:break-word;
    max-width:100%;
  }
  @media(max-width:1360px){
    .token-grid,.cols-4,.cols-5,.cols-3a,.cols-3b,.cols-ntp,.cols-device,.cols-log{
      grid-template-columns:repeat(2,minmax(220px,1fr));
    }
  }
  @media(max-width:1100px){
    .token-grid,.cols-4,.cols-5,.cols-3a,.cols-3b,.cols-ntp,.cols-device,.cols-log{grid-template-columns:1fr;}
  }
  .topnav{display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px}
  .navbtn{padding:8px 12px; border:1px solid #c2d2e4; border-radius:8px; background:#e8f0f8; color:#1f2933; text-decoration:none}
  .navbtn.active{background:#005eb8; color:#fff; border-color:#005eb8}

</style>
</head>
<body>
  <div class="wrap">
    __NAV__
  <h2>System Settings</h2>
  <p class="muted">
    Token wird im Browser gespeichert (localStorage). Aenderungen an Network koennen dich aussperren - daher "Apply now" bewusst setzen.
    <br/>Hinweis: Subnet-Maske (z.B. 255.255.255.0) und Prefix (z.B. /24) sind identisch - nur andere Schreibweise.
  </p>

  <div class="grid token-grid">
    <div class="field">
      <label>UI Token</label>
      <input id="token" placeholder="MAS004-..."/>
    </div>
    <div class="actions">
      <button onclick="saveToken()">Save</button>
      <span id="tokstate" class="pill"></span>
    </div>
  </div>

  <fieldset>
    <legend>Raspi Network (eth0/eth1)</legend>

    <div class="grid cols-5">
      <div class="field"><label>eth0 IP</label><input id="eth0_ip"/></div>
      <div class="field"><label>Subnet</label><input id="eth0_mask" placeholder="255.255.255.0" oninput="maskChanged('eth0')"/></div>
      <div class="field"><label>Prefix</label><input id="eth0_pre" placeholder="24" oninput="prefixChanged('eth0')"/></div>
      <div class="field"><label>GW</label><input id="eth0_gw"/></div>
      <div class="field"><label>DNS (eth0)</label><input id="eth0_dns" placeholder="z.B. 10.28.193.4, 10.27.30.201"/></div>
    </div>

    <div class="grid cols-5">
      <div class="field"><label>eth1 IP</label><input id="eth1_ip"/></div>
      <div class="field"><label>Subnet</label><input id="eth1_mask" placeholder="255.255.255.0" oninput="maskChanged('eth1')"/></div>
      <div class="field"><label>Prefix</label><input id="eth1_pre" placeholder="24" oninput="prefixChanged('eth1')"/></div>
      <div class="field"><label>GW</label><input id="eth1_gw"/></div>
      <div class="field"><label>DNS (eth1)</label><input id="eth1_dns" placeholder="optional"/></div>
    </div>
    <div class="muted">Hinweis: Fuer Produktions-/Firmennetz normalerweise nur `eth0` mit Gateway und DNS setzen. `eth1` Gateway leer lassen, falls nur Maschinen-LAN.</div>

    <div class="actions">
      <label class="checkline"><input type="checkbox" id="apply_now"/>Apply now (live setzen)</label>
      <button onclick="saveNetwork()">Save Network</button>
      <button onclick="reloadAll()">Reload</button>
      <span id="net_status" class="muted"></span>
    </div>

    <h4>Status</h4>
    <pre id="netinfo"></pre>
  </fieldset>

  <fieldset>
    <legend>Databridge / Microtom</legend>
    <div class="grid cols-3a">
      <div class="field"><label>peer_base_url</label><input id="peer_base_url"/></div>
      <div class="field"><label>peer_base_url_secondary (optional parallel)</label><input id="peer_base_url_secondary" placeholder="z.B. https://192.168.5.2:9090"/></div>
      <div class="field"><label>peer_watchdog_host</label><input id="peer_watchdog_host"/></div>
      <div class="field"><label>peer_health_path</label><input id="peer_health_path"/></div>
    </div>
    <div class="muted">Wenn gesetzt, werden ausgehende Raspi-&gt;Microtom Nachrichten parallel an beide Endpunkte gesendet.</div>
    <div class="grid cols-3b">
      <div class="field"><label>http_timeout_s</label><input id="http_timeout_s"/></div>
      <div class="field"><label>tls_verify</label><input id="tls_verify" placeholder="true/false"/></div>
      <div class="field"><label>eth0_source_ip</label><input id="eth0_source_ip"/></div>
    </div>
    <div class="grid cols-ntp">
      <div class="field"><label>ntp_server</label><input id="ntp_server" placeholder="z.B. 10.27.30.201 oder pool.ntp.org"/></div>
      <div class="field"><label>ntp_sync_interval_min</label><input id="ntp_sync_interval_min" type="number" min="1" max="1440"/></div>
    </div>
    <div class="muted">NTP: Raspi synchronisiert die Zeit zyklisch gegen den konfigurierten Server.</div>
    <div class="grid cols-5" style="margin-top:8px;">
      <div class="field"><label>System local time</label><input id="sys_local_time" readonly/></div>
      <div class="field"><label>Time zone</label><input id="sys_time_zone" readonly/></div>
      <div class="field"><label>Clock synchronized</label><input id="sys_clock_sync" readonly/></div>
      <div class="field"><label>OS NTP service</label><input id="sys_ntp_service" readonly/></div>
      <div class="field"><label>OS time source</label><input id="sys_timesync_server" readonly/></div>
    </div>
    <div class="muted" id="sys_time_hint">Systemzeit/Zeitzone des Raspi. Zeitzone wird vom Betriebssystem vorgegeben, nicht vom Databridge-NTP-Feld.</div>
    <div class="grid">
      <div class="field">
        <label>shared_secret</label>
        <input id="shared_secret" placeholder="(leer = aus)"/>
        <div id="shared_secret_state" class="muted">(leer = aus)</div>
      </div>
      <div class="field">
        <label>X-DIClient-Adapter-Key</label>
        <input id="diclient_adapter_key" placeholder="(leer = kein Header)"/>
        <div id="diclient_adapter_key_state" class="muted">(leer = kein Header)</div>
      </div>
    </div>
    <div class="actions">
      <label class="checkline"><input type="checkbox" id="clear_shared_secret"/>shared_secret loeschen (auf leer setzen)</label>
      <label class="checkline"><input type="checkbox" id="clear_diclient_adapter_key"/>DIClient-Key loeschen</label>
    </div>
    <div class="actions">
      <button onclick="saveBridge()">Save Bridge + Restart</button>
      <span id="bridge_status" class="muted"></span>
    </div>
  </fieldset>

  <fieldset>
    <legend>Device Endpoints (ETH0 / ETH1)</legend>
    <div class="muted">ETH0 ist das Hauptnetz fuer Microtom, Laser, TTO und die beiden Wickler. ETH1 ist das Maschinen-LAN fuer ESP32-PLC58 und die drei Moxa ioLogik E1213. Das alte Port-Forwarding wurde entfernt; alle Geraete werden direkt ueber ihre Ziel-IP angesprochen.</div>
    <div class="grid cols-device">
      <div class="field"><label>Raspi PLC model</label><input id="raspi_plc_model" placeholder="RPIPLC_21"/></div>
      <div class="field"><label>Raspi IO poll interval (s)</label><input id="raspi_io_poll_interval_s" type="number" min="0.2" max="60" step="0.1"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="lokale PLC21 I/Os"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled/></div>
      <label class="checkline"><input type="checkbox" id="raspi_io_simulation"/>Raspi IO Simulation</label>
    </div>
    <div class="grid cols-device">
      <div class="field"><label>ESP host</label><input id="esp_host"/></div>
      <div class="field"><label>ESP port</label><input id="esp_port"/></div>
      <div class="field"><label>ESP watchdog host</label><input id="esp_watchdog_host" placeholder="leer = esp_host"/></div>
      <div class="field"><label>ESP IO poll interval (s)</label><input id="esp_io_poll_interval_s" type="number" min="0.2" max="60" step="0.1"/></div>
      <label class="checkline"><input type="checkbox" id="esp_simulation"/>Simulation</label>
    </div>
    <div class="actions">
      <label class="checkline"><input type="checkbox" id="light_curtain_auto_reset_enabled"/>Lichtgitter Auto-Reset alle 5s aktiv</label>
      <span class="muted">Nur wenn ausschliesslich das Lichtgitter ausgeloest hat und Not-Aus OK ist; pulst ESP Q0.2 ohne Purge/Statuswechsel.</span>
    </div>
    <div class="grid cols-device">
      <div class="field"><label>Moxa #1 host</label><input id="moxa1_host"/></div>
      <div class="field"><label>Moxa #1 port</label><input id="moxa1_port"/></div>
      <div class="field"><label>Moxa poll interval (s)</label><input id="moxa_poll_interval_s" type="number" min="0.2" max="60" step="0.1"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="Modbus/TCP direkt, default 502"/></div>
      <label class="checkline"><input type="checkbox" id="moxa1_simulation"/>Simulation</label>
    </div>
    <div class="grid cols-device">
      <div class="field"><label>Moxa #2 host</label><input id="moxa2_host"/></div>
      <div class="field"><label>Moxa #2 port</label><input id="moxa2_port"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="Modbus/TCP direkt, default 502"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled/></div>
      <label class="checkline"><input type="checkbox" id="moxa2_simulation"/>Simulation</label>
    </div>
    <div class="grid cols-device">
      <div class="field"><label>Moxa #3 host</label><input id="moxa3_host"/></div>
      <div class="field"><label>Moxa #3 port</label><input id="moxa3_port"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="Modbus/TCP direkt, default 502"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled/></div>
      <label class="checkline"><input type="checkbox" id="moxa3_simulation"/>Simulation</label>
    </div>
    <div class="grid cols-device">
      <div class="field"><label>VJ3350 host</label><input id="vj3350_host"/></div>
      <div class="field"><label>VJ3350 port</label><input id="vj3350_port"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="direkt auf ETH0"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled/></div>
      <label class="checkline"><input type="checkbox" id="vj3350_simulation"/>Simulation</label>
    </div>
    <div class="grid cols-device">
      <div class="field"><label>VJ6530 host</label><input id="vj6530_host"/></div>
      <div class="field"><label>VJ6530 port</label><input id="vj6530_port"/></div>
      <div class="field"><label>VJ6530 poll interval (s)</label><input id="vj6530_poll_interval_s" type="number" min="0.5" max="300" step="0.5"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="direkt auf ETH0"/></div>
      <label class="checkline"><input type="checkbox" id="vj6530_simulation"/>Simulation</label>
    </div>
    <div class="actions">
      <label class="checkline"><input type="checkbox" id="vj6530_async_enabled"/>VJ6530 async ZBC Events aktiv</label>
      <span class="muted">Async ist der Primaerpfad fuer Online/Offline/Warning/Fault/Buzy/Print-Events. Polling bleibt als Fallback.</span>
    </div>
    <div class="grid cols-device">
      <div class="field"><label>Abwickler host</label><input id="smart_unwinder_host"/></div>
      <div class="field"><label>Abwickler port</label><input id="smart_unwinder_port"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="direkt auf ETH0"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="oeffnet /api/state bzw. /"/></div>
      <label class="checkline"><input type="checkbox" id="smart_unwinder_simulation"/>Simulation</label>
    </div>
    <div class="grid cols-device">
      <div class="field"><label>Aufwickler host</label><input id="smart_rewinder_host"/></div>
      <div class="field"><label>Aufwickler port</label><input id="smart_rewinder_port"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="direkt auf ETH0"/></div>
      <div class="field empty"><label>&nbsp;</label><input disabled placeholder="oeffnet /api/state bzw. /"/></div>
      <label class="checkline"><input type="checkbox" id="smart_rewinder_simulation"/>Simulation</label>
    </div>
    <div class="actions">
      <button onclick="saveDevices()">Save Devices + Restart</button>
      <span id="dev_status" class="muted"></span>
    </div>
  </fieldset>

  <fieldset>
    <legend>Daily Log Files</legend>
    <div class="grid cols-log">
      <div class="field"><label>Keep days (All)</label><input id="logs_keep_days_all" type="number" min="1" max="3650"/></div>
      <div class="field"><label>Keep days (ESP32)</label><input id="logs_keep_days_esp" type="number" min="1" max="3650"/></div>
      <div class="field"><label>Keep days (TTO)</label><input id="logs_keep_days_tto" type="number" min="1" max="3650"/></div>
      <div class="field"><label>Keep days (Laser)</label><input id="logs_keep_days_laser" type="number" min="1" max="3650"/></div>
      <div class="field"><label>Machine audit keep hours</label><input id="machine_audit_keep_hours" type="number" min="1" max="87600"/></div>
    </div>
    <div class="actions">
      <button onclick="saveLogSettings()">Save Log Settings + Restart</button>
      <button onclick="loadDailyLogFiles()">Reload Log File List</button>
      <span id="logcfg_status" class="muted"></span>
    </div>
    <div style="overflow:auto; margin-top:8px;">
      <table style="width:100%; border-collapse:collapse;">
        <thead>
          <tr>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Datei</th>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Typ</th>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Datum</th>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Groesse</th>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Aktion</th>
          </tr>
        </thead>
        <tbody id="daily_log_files"></tbody>
      </table>
    </div>
  </fieldset>

  <fieldset>
    <legend>Production Log Files</legend>
    <div class="muted">Diese Dateien werden pro Produktion erzeugt. Ein Download entfernt die jeweilige Produktionsdatei direkt vom Raspi. Wenn die letzte Datei heruntergeladen wurde, wird `MAS0030=0` automatisch gesetzt und an Microtom gemeldet.</div>
    <div class="grid cols-3b">
      <div class="field"><label>Production label</label><input id="prod_label" readonly/></div>
      <div class="field"><label>Recording active</label><input id="prod_active" readonly/></div>
      <div class="field"><label>Files ready</label><input id="prod_ready" readonly/></div>
    </div>
    <div class="actions">
      <button onclick="loadProductionLogFiles()">Reload Production Log File List</button>
      <span id="prodlog_status" class="muted"></span>
    </div>
    <div style="overflow:auto; margin-top:8px;">
      <table style="width:100%; border-collapse:collapse;">
        <thead>
          <tr>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Datei</th>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Typ</th>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Groesse</th>
            <th style="text-align:left; border-bottom:1px solid #d6dde7; padding:6px;">Aktion</th>
          </tr>
        </thead>
        <tbody id="production_log_files"></tbody>
      </table>
    </div>
  </fieldset>

  <fieldset>
    <legend>Queue Maintenance</legend>
    <div class="grid cols-3b">
      <div class="field"><label>Outbox count</label><input id="queue_outbox" readonly/></div>
      <div class="field"><label>Inbox pending</label><input id="queue_inbox" readonly/></div>
      <div class="field"><label>Last action</label><input id="queue_action" readonly placeholder="none"/></div>
    </div>
    <div class="actions">
      <button onclick="refreshQueueStatus()">Reload Queue Status</button>
      <button onclick="clearQueue('outbox')">Clear Outbox</button>
      <button onclick="clearQueue('inbox')">Clear Inbox</button>
      <span class="muted">Clear Inbox leert die komplette Inbox-Tabelle, nicht nur pending.</span>
    </div>
  </fieldset>
  </div>

<script>
function getToken(){ return localStorage.getItem("mas004_ui_token") || ""; }
function saveToken(){
  localStorage.setItem("mas004_ui_token", document.getElementById("token").value.trim());
  showTok();
}
function showTok(){
  const t = getToken();
  document.getElementById("token").value = t;
  document.getElementById("tokstate").textContent = t ? "token ok" : "no token";
}
async function api(path, opt={}){
  opt.headers = opt.headers || {};
  const t = getToken();
  if(t) opt.headers["X-Token"] = t;
  const r = await fetch(path, opt);
  const txt = await r.text();
  let j=null; try{ j=JSON.parse(txt); }catch(e){}
  if(!r.ok){
    throw new Error((j && j.detail) ? j.detail : ("HTTP "+r.status+" "+txt));
  }
  return j;
}

// ---------- Mask <-> Prefix ----------
function prefixToMask(prefix){
  const p = Number(prefix);
  if(!Number.isInteger(p) || p < 0 || p > 32) return null;
  let mask = 0 >>> 0;
  if(p === 0) mask = 0;
  else mask = (0xFFFFFFFF << (32 - p)) >>> 0;
  const a = (mask >>> 24) & 255;
  const b = (mask >>> 16) & 255;
  const c = (mask >>> 8) & 255;
  const d = mask & 255;
  return `${a}.${b}.${c}.${d}`;
}

function maskToPrefix(maskStr){
  const parts = (maskStr||"").trim().split(".");
  if(parts.length !== 4) return null;
  const nums = parts.map(x => Number(x));
  if(nums.some(n => !Number.isInteger(n) || n < 0 || n > 255)) return null;

  let m = ((nums[0]<<24)>>>0) | ((nums[1]<<16)>>>0) | ((nums[2]<<8)>>>0) | (nums[3]>>>0);

  // contiguous ones then zeros check
  let seenZero = false;
  let prefix = 0;
  for(let i=31;i>=0;i--){
    const bit = (m >>> i) & 1;
    if(bit === 1){
      if(seenZero) return null;
      prefix++;
    }else{
      seenZero = true;
    }
  }
  return prefix;
}

function setBad(el, bad){
  if(bad) el.classList.add("bad");
  else el.classList.remove("bad");
}

function maskChanged(iface){
  const maskEl = document.getElementById(`${iface}_mask`);
  const preEl  = document.getElementById(`${iface}_pre`);
  const p = maskToPrefix(maskEl.value);
  if(p === null){
    setBad(maskEl, true);
  }else{
    setBad(maskEl, false);
    preEl.value = String(p);
    setBad(preEl, false);
  }
}

function prefixChanged(iface){
  const maskEl = document.getElementById(`${iface}_mask`);
  const preEl  = document.getElementById(`${iface}_pre`);
  const m = prefixToMask(preEl.value);
  if(m === null){
    setBad(preEl, true);
  }else{
    setBad(preEl, false);
    maskEl.value = m;
    setBad(maskEl, false);
  }
}

function effectivePrefix(iface){
  // bevorzugt: aus Maske berechnen (wenn gueltig)
  const mask = document.getElementById(`${iface}_mask`).value.trim();
  if(mask){
    const p = maskToPrefix(mask);
    if(p !== null) return p;
  }
  // fallback: Prefix-Feld
  const pre = Number(document.getElementById(`${iface}_pre`).value.trim());
  if(Number.isInteger(pre) && pre >= 0 && pre <= 32) return pre;
  return null;
}

async function reloadAll(){
  showTok();

  const st = await api("/api/ui/status");
  document.getElementById("queue_outbox").value = String(st.outbox_count ?? 0);
  document.getElementById("queue_inbox").value = String(st.inbox_pending ?? 0);
  document.getElementById("queue_action").value = "reloaded";

  // config
  const cfg = await api("/api/config");
  const c = cfg.config;
  document.getElementById("peer_base_url").value = c.peer_base_url || "";
  document.getElementById("peer_base_url_secondary").value = c.peer_base_url_secondary || "";
  document.getElementById("peer_watchdog_host").value = c.peer_watchdog_host || "";
  document.getElementById("peer_health_path").value = c.peer_health_path || "";
  document.getElementById("http_timeout_s").value = c.http_timeout_s ?? "";
  document.getElementById("tls_verify").value = String(c.tls_verify ?? false);
  document.getElementById("eth0_source_ip").value = c.eth0_source_ip || "";
  document.getElementById("ntp_server").value = c.ntp_server || "";
  document.getElementById("ntp_sync_interval_min").value = c.ntp_sync_interval_min ?? 60;
  const secEl = document.getElementById("shared_secret");
  const secStateEl = document.getElementById("shared_secret_state");
  const hasMaskedSecret = c.shared_secret === "***";
  if(hasMaskedSecret){
    secEl.value = "********";
    secEl.placeholder = "gesetzt (versteckt)";
    secStateEl.textContent = "gesetzt (versteckt)";
  }else{
    secEl.value = (c.shared_secret || "").trim();
    secEl.placeholder = "(leer = aus)";
    secStateEl.textContent = secEl.value ? "gesetzt" : "(leer = aus)";
  }
  document.getElementById("clear_shared_secret").checked = false;
  const adapterKeyEl = document.getElementById("diclient_adapter_key");
  const adapterKeyStateEl = document.getElementById("diclient_adapter_key_state");
  const hasMaskedAdapterKey = c.diclient_adapter_key === "***";
  if(hasMaskedAdapterKey){
    adapterKeyEl.value = "********";
    adapterKeyEl.placeholder = "gesetzt (versteckt)";
    adapterKeyStateEl.textContent = "gesetzt (versteckt)";
  }else{
    adapterKeyEl.value = (c.diclient_adapter_key || "").trim();
    adapterKeyEl.placeholder = "(leer = kein Header)";
    adapterKeyStateEl.textContent = adapterKeyEl.value ? "gesetzt" : "(leer = kein Header)";
  }
  document.getElementById("clear_diclient_adapter_key").checked = false;

  document.getElementById("raspi_plc_model").value = c.raspi_plc_model || "RPIPLC_21";
  document.getElementById("raspi_io_poll_interval_s").value = c.raspi_io_poll_interval_s ?? 1.0;
  document.getElementById("raspi_io_simulation").checked = !!(c.raspi_io_simulation ?? true);
  document.getElementById("esp_host").value = c.esp_host || "";
  document.getElementById("esp_port").value = c.esp_port ?? "";
  document.getElementById("esp_watchdog_host").value = c.esp_watchdog_host || "";
  document.getElementById("esp_io_poll_interval_s").value = c.esp_io_poll_interval_s ?? 1.0;
  document.getElementById("esp_simulation").checked = !!c.esp_simulation;
  document.getElementById("light_curtain_auto_reset_enabled").checked = !!(c.light_curtain_auto_reset_enabled ?? true);
  document.getElementById("moxa1_host").value = c.moxa1_host || "";
  document.getElementById("moxa1_port").value = c.moxa1_port ?? 502;
  document.getElementById("moxa1_simulation").checked = !!(c.moxa1_simulation ?? true);
  document.getElementById("moxa2_host").value = c.moxa2_host || "";
  document.getElementById("moxa2_port").value = c.moxa2_port ?? 502;
  document.getElementById("moxa2_simulation").checked = !!(c.moxa2_simulation ?? true);
  document.getElementById("moxa3_host").value = c.moxa3_host || "";
  document.getElementById("moxa3_port").value = c.moxa3_port ?? 502;
  document.getElementById("moxa3_simulation").checked = !!(c.moxa3_simulation ?? true);
  document.getElementById("moxa_poll_interval_s").value = c.moxa_poll_interval_s ?? 1.0;
  document.getElementById("vj3350_host").value = c.vj3350_host || "";
  document.getElementById("vj3350_port").value = c.vj3350_port ?? "";
  document.getElementById("vj3350_simulation").checked = !!c.vj3350_simulation;
  document.getElementById("vj6530_host").value = c.vj6530_host || "";
  document.getElementById("vj6530_port").value = c.vj6530_port ?? "";
  document.getElementById("vj6530_poll_interval_s").value = c.vj6530_poll_interval_s ?? 15.0;
  document.getElementById("vj6530_simulation").checked = !!c.vj6530_simulation;
  document.getElementById("vj6530_async_enabled").checked = !!(c.vj6530_async_enabled ?? true);
  document.getElementById("smart_unwinder_host").value = c.smart_unwinder_host || "";
  document.getElementById("smart_unwinder_port").value = c.smart_unwinder_port ?? "";
  document.getElementById("smart_unwinder_simulation").checked = !!(c.smart_unwinder_simulation ?? true);
  document.getElementById("smart_rewinder_host").value = c.smart_rewinder_host || "";
  document.getElementById("smart_rewinder_port").value = c.smart_rewinder_port ?? "";
  document.getElementById("smart_rewinder_simulation").checked = !!(c.smart_rewinder_simulation ?? true);
  document.getElementById("logs_keep_days_all").value = c.logs_keep_days_all ?? 30;
  document.getElementById("logs_keep_days_esp").value = c.logs_keep_days_esp ?? 30;
  document.getElementById("logs_keep_days_tto").value = c.logs_keep_days_tto ?? 30;
  document.getElementById("logs_keep_days_laser").value = c.logs_keep_days_laser ?? 30;
  document.getElementById("machine_audit_keep_hours").value = c.machine_audit_keep_hours ?? 72;

  // network
  const net = await api("/api/system/network");
  const n = net.config;

  document.getElementById("eth0_ip").value = n.eth0_ip || "";
  document.getElementById("eth0_pre").value = n.eth0_subnet || "";     // bei dir ist das "Subnet" intern Prefix-String
  prefixChanged("eth0");                                               // fuellt Mask automatisch
  document.getElementById("eth0_gw").value = n.eth0_gateway || "";
  document.getElementById("eth0_dns").value = n.eth0_dns || "";

  document.getElementById("eth1_ip").value = n.eth1_ip || "";
  document.getElementById("eth1_pre").value = n.eth1_subnet || "";
  prefixChanged("eth1");
  document.getElementById("eth1_gw").value = n.eth1_gateway || "";
  document.getElementById("eth1_dns").value = n.eth1_dns || "";

  document.getElementById("netinfo").textContent = JSON.stringify(net.status, null, 2);
  const ti = await api("/api/system/time");
  document.getElementById("sys_local_time").value = ti.local_time || "";
  document.getElementById("sys_time_zone").value = ti.time_zone || "";
  document.getElementById("sys_clock_sync").value = ti.system_clock_synchronized || "";
  document.getElementById("sys_ntp_service").value = ti.ntp_service || "";
  document.getElementById("sys_timesync_server").value = ti.timesync_server || ti.timesync_address || "";
  const tiHint = document.getElementById("sys_time_hint");
  if(ti.ok === false && ti.error){
    tiHint.textContent = `Systemzeit-Status konnte nicht gelesen werden: ${ti.error}`;
  }else{
    tiHint.textContent = "Systemzeit/Zeitzone des Raspi. Zeitzone wird vom Betriebssystem vorgegeben, nicht vom Databridge-NTP-Feld.";
  }
  await loadDailyLogFiles();
  await loadProductionLogFiles();
}

async function refreshQueueStatus(){
  const st = await api("/api/ui/status");
  document.getElementById("queue_outbox").value = String(st.outbox_count ?? 0);
  document.getElementById("queue_inbox").value = String(st.inbox_pending ?? 0);
  document.getElementById("queue_action").value = "status reloaded";
}

async function clearQueue(which){
  const target = (which || "").trim().toLowerCase();
  if(target !== "outbox" && target !== "inbox"){
    return;
  }
  if(!confirm(`Really clear ${target}?`)) return;
  const j = await api(`/api/queues/${target}/clear`, {method:"POST"});
  await refreshQueueStatus();
  document.getElementById("queue_action").value = `${target} cleared: ${j.deleted ?? 0}`;
}

async function saveNetwork(){
  document.getElementById("net_status").textContent = "saving...";

  const p0 = effectivePrefix("eth0");
  const p1 = effectivePrefix("eth1");
  if(p0 === null || p1 === null){
    alert("Subnet/Prefix ungueltig. Bitte Maske (z.B. 255.255.255.0) oder Prefix (0..32) korrekt setzen.");
    document.getElementById("net_status").textContent = "ERROR";
    return;
  }

  const payload = {
    eth0_ip: document.getElementById("eth0_ip").value.trim(),
    eth0_prefix: p0,
    eth0_gateway: document.getElementById("eth0_gw").value.trim(),
    eth0_dns: document.getElementById("eth0_dns").value.trim(),
    eth1_ip: document.getElementById("eth1_ip").value.trim(),
    eth1_prefix: p1,
    eth1_gateway: document.getElementById("eth1_gw").value.trim(),
    eth1_dns: document.getElementById("eth1_dns").value.trim(),
    apply_now: document.getElementById("apply_now").checked
  };

  const j = await api("/api/system/network", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify(payload)
  });

  document.getElementById("net_status").textContent = "ok";
  if(j.applied && j.applied.length){
    alert("Applied:\\n" + JSON.stringify(j.applied, null, 2));
  }
  await reloadAll();
}

async function saveBridge(){
  document.getElementById("bridge_status").textContent = "saving...";
  const secEl = document.getElementById("shared_secret");
  const secretRaw = secEl.value.trim();
  const clearSecret = document.getElementById("clear_shared_secret").checked;
  const adapterKeyEl = document.getElementById("diclient_adapter_key");
  const adapterKeyRaw = adapterKeyEl.value.trim();
  const clearAdapterKey = document.getElementById("clear_diclient_adapter_key").checked;
  const ntpIntervalRaw = Number(document.getElementById("ntp_sync_interval_min").value.trim());
  const ntpInterval = Number.isFinite(ntpIntervalRaw) ? ntpIntervalRaw : 60;
  let sharedSecretValue = null; // null => unveraendert lassen
  if(clearSecret){
    sharedSecretValue = "";
  }else if(secretRaw && secretRaw !== "********"){
    sharedSecretValue = secretRaw;
  }
  let adapterKeyValue = null; // null => unveraendert lassen
  if(clearAdapterKey){
    adapterKeyValue = "";
  }else if(adapterKeyRaw && adapterKeyRaw !== "********"){
    adapterKeyValue = adapterKeyRaw;
  }

  const payload = {
    peer_base_url: document.getElementById("peer_base_url").value.trim(),
    peer_base_url_secondary: document.getElementById("peer_base_url_secondary").value.trim(),
    peer_watchdog_host: document.getElementById("peer_watchdog_host").value.trim(),
    peer_health_path: document.getElementById("peer_health_path").value.trim(),
    http_timeout_s: Number(document.getElementById("http_timeout_s").value.trim()),
    tls_verify: (document.getElementById("tls_verify").value.trim().toLowerCase()==="true"),
    eth0_source_ip: document.getElementById("eth0_source_ip").value.trim(),
    ntp_server: document.getElementById("ntp_server").value.trim(),
    ntp_sync_interval_min: ntpInterval,
    shared_secret: sharedSecretValue,
    diclient_adapter_key: adapterKeyValue
  };
  await api("/api/config", {method:"POST", headers:{"Content-Type":"application/json"}, body: JSON.stringify(payload)});
  document.getElementById("bridge_status").textContent = "saved (service restarted)";
}

async function saveDevices(){
  document.getElementById("dev_status").textContent = "saving...";
  const payload = {
    raspi_plc_model: document.getElementById("raspi_plc_model").value.trim(),
    raspi_io_poll_interval_s: Number(document.getElementById("raspi_io_poll_interval_s").value.trim()),
    raspi_io_simulation: document.getElementById("raspi_io_simulation").checked,
    esp_host: document.getElementById("esp_host").value.trim(),
    esp_port: Number(document.getElementById("esp_port").value.trim()),
    esp_watchdog_host: document.getElementById("esp_watchdog_host").value.trim(),
    esp_io_poll_interval_s: Number(document.getElementById("esp_io_poll_interval_s").value.trim()),
    esp_simulation: document.getElementById("esp_simulation").checked,
    light_curtain_auto_reset_enabled: document.getElementById("light_curtain_auto_reset_enabled").checked,
    moxa1_host: document.getElementById("moxa1_host").value.trim(),
    moxa1_port: Number(document.getElementById("moxa1_port").value.trim()),
    moxa1_simulation: document.getElementById("moxa1_simulation").checked,
    moxa2_host: document.getElementById("moxa2_host").value.trim(),
    moxa2_port: Number(document.getElementById("moxa2_port").value.trim()),
    moxa2_simulation: document.getElementById("moxa2_simulation").checked,
    moxa3_host: document.getElementById("moxa3_host").value.trim(),
    moxa3_port: Number(document.getElementById("moxa3_port").value.trim()),
    moxa3_simulation: document.getElementById("moxa3_simulation").checked,
    moxa_poll_interval_s: Number(document.getElementById("moxa_poll_interval_s").value.trim()),
    vj3350_host: document.getElementById("vj3350_host").value.trim(),
    vj3350_port: Number(document.getElementById("vj3350_port").value.trim()),
    vj3350_simulation: document.getElementById("vj3350_simulation").checked,
    vj6530_host: document.getElementById("vj6530_host").value.trim(),
    vj6530_port: Number(document.getElementById("vj6530_port").value.trim()),
    vj6530_poll_interval_s: Number(document.getElementById("vj6530_poll_interval_s").value.trim()),
    vj6530_simulation: document.getElementById("vj6530_simulation").checked,
    vj6530_async_enabled: document.getElementById("vj6530_async_enabled").checked,
    smart_unwinder_host: document.getElementById("smart_unwinder_host").value.trim(),
    smart_unwinder_port: Number(document.getElementById("smart_unwinder_port").value.trim()),
    smart_unwinder_simulation: document.getElementById("smart_unwinder_simulation").checked,
    smart_rewinder_host: document.getElementById("smart_rewinder_host").value.trim(),
    smart_rewinder_port: Number(document.getElementById("smart_rewinder_port").value.trim()),
    smart_rewinder_simulation: document.getElementById("smart_rewinder_simulation").checked
  };
  await api("/api/config", {method:"POST", headers:{"Content-Type":"application/json"}, body: JSON.stringify(payload)});
  document.getElementById("dev_status").textContent = "saved (service restarted)";
}

function toNum(id, fallback){
  const v = Number(document.getElementById(id).value.trim());
  if(!Number.isFinite(v)) return fallback;
  return Math.max(1, Math.min(3650, Math.round(v)));
}

function toHours(id, fallback){
  const v = Number(document.getElementById(id).value.trim());
  if(!Number.isFinite(v)) return fallback;
  return Math.max(1, Math.min(87600, Math.round(v)));
}

function fmtBytes(n){
  const v = Number(n || 0);
  if(v < 1024) return `${v} B`;
  if(v < 1024*1024) return `${(v/1024).toFixed(1)} KB`;
  return `${(v/(1024*1024)).toFixed(2)} MB`;
}

async function saveLogSettings(){
  document.getElementById("logcfg_status").textContent = "saving...";
  const payload = {
    logs_keep_days_all: toNum("logs_keep_days_all", 30),
    logs_keep_days_esp: toNum("logs_keep_days_esp", 30),
    logs_keep_days_tto: toNum("logs_keep_days_tto", 30),
    logs_keep_days_laser: toNum("logs_keep_days_laser", 30),
    machine_audit_keep_hours: toHours("machine_audit_keep_hours", 72)
  };
  await api("/api/config", {method:"POST", headers:{"Content-Type":"application/json"}, body: JSON.stringify(payload)});
  document.getElementById("logcfg_status").textContent = "saved (service restarted)";
  await loadDailyLogFiles();
}

async function loadDailyLogFiles(){
  const tbody = document.getElementById("daily_log_files");
  tbody.innerHTML = '<tr><td colspan="5" style="padding:6px;">loading...</td></tr>';
  try{
    const j = await api("/api/logfiles/list");
    const items = j.items || [];
    if(!items.length){
      tbody.innerHTML = '<tr><td colspan="5" style="padding:6px;">keine Dateien</td></tr>';
      return;
    }
    const rows = items.map(it => {
      const name = it.name || "";
      const grp = it.group_label || it.group || "";
      const dt = it.date || "";
      const sz = fmtBytes(it.size_bytes || 0);
      const btn = `<button onclick="downloadDailyLog('${name.replace(/'/g, "\\'")}')">Download</button>`;
      return `<tr>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${name}</td>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${grp}</td>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${dt}</td>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${sz}</td>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${btn}</td>
      </tr>`;
    });
    tbody.innerHTML = rows.join("");
  }catch(e){
    tbody.innerHTML = `<tr><td colspan="5" style="padding:6px; color:#c62828;">ERROR: ${e.message}</td></tr>`;
  }
}

async function downloadDailyLog(name){
  try{
    const t = getToken();
    const r = await fetch("/api/logfiles/download?name=" + encodeURIComponent(name), {headers: t ? {"X-Token": t} : {}});
    if(!r.ok){
      const txt = await r.text();
      throw new Error(txt || ("HTTP " + r.status));
    }
    const blob = await r.blob();
    const a = document.createElement("a");
    a.href = URL.createObjectURL(blob);
    a.download = name;
    a.click();
    URL.revokeObjectURL(a.href);
  }catch(e){
    alert("Download failed: " + e.message);
  }
}

async function loadProductionLogFiles(){
  const tbody = document.getElementById("production_log_files");
  const status = document.getElementById("prodlog_status");
  tbody.innerHTML = '<tr><td colspan="4" style="padding:6px;">loading...</td></tr>';
  status.textContent = "";
  try{
    const t = getToken();
    const r = await fetch("/api/production/logfiles/list", {headers: t ? {"X-Token": t} : {}});
    const txt = await r.text();
    let j = null; try{ j = JSON.parse(txt); }catch(e){}
    if(!r.ok){
      throw new Error((j && j.detail) ? j.detail : ("HTTP " + r.status + " " + txt));
    }
    document.getElementById("prod_label").value = j.production_label || "";
    document.getElementById("prod_active").value = j.active ? "yes" : "no";
    document.getElementById("prod_ready").value = j.ready ? "yes" : "no";
    const items = j.files || [];
    if(!items.length){
      tbody.innerHTML = '<tr><td colspan="4" style="padding:6px;">keine Produktionsdateien bereit</td></tr>';
      return;
    }
    const rows = items.map(it => {
      const name = it.name || "";
      const grp = it.group_label || it.group || "";
      const sz = fmtBytes(it.size_bytes || 0);
      const btn = `<button onclick="downloadProductionLog('${name.replace(/'/g, "\\'")}')">Download + Delete</button>`;
      return `<tr>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${name}</td>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${grp}</td>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${sz}</td>
        <td style="padding:6px; border-top:1px solid #e7edf6;">${btn}</td>
      </tr>`;
    });
    tbody.innerHTML = rows.join("");
  }catch(e){
    tbody.innerHTML = `<tr><td colspan="4" style="padding:6px; color:#c62828;">ERROR: ${e.message}</td></tr>`;
  }
}

async function downloadProductionLog(name){
  try{
    const t = getToken();
    const r = await fetch("/api/production/logfiles/download?name=" + encodeURIComponent(name), {headers: t ? {"X-Token": t} : {}});
    if(!r.ok){
      const txt = await r.text();
      throw new Error(txt || ("HTTP " + r.status));
    }
    const blob = await r.blob();
    const a = document.createElement("a");
    a.href = URL.createObjectURL(blob);
    a.download = name;
    a.click();
    URL.revokeObjectURL(a.href);
    document.getElementById("prodlog_status").textContent = `downloaded + deleted: ${name}`;
    await loadProductionLogFiles();
  }catch(e){
    alert("Production download failed: " + e.message);
  }
}

showTok();
reloadAll();
</script>
</body></html>
        """.replace("__NAV__", nav)

    # -----------------------------
    # Test UI
    # -----------------------------
    @app.get("/ui/test", response_class=HTMLResponse)
    def ui_test():
        nav = nav_html("test")
        return """
<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <title>MAS-004</title>
  <style>
    :root{
      --bg:#f4f6f9;
      --card:#ffffff;
      --line:#d6dde7;
      --text:#1f2933;
      --muted:#5f6b7a;
      --blue:#005eb8;
      --green:#0f9d58;
      --red:#c62828;
    }
    body{margin:0; font-family:Segoe UI,Arial,sans-serif; background:var(--bg); color:var(--text)}
    .wrap{max-width:1500px; margin:0 auto; padding:16px}
    .row{display:flex; gap:10px; align-items:center; flex-wrap:wrap}
    .topnav{display:flex; gap:8px; flex-wrap:wrap; margin-bottom:12px}
    .navbtn{padding:8px 12px; border:1px solid var(--line); border-radius:8px; background:#fff; color:#1f2933; text-decoration:none}
    .navbtn.active{background:#005eb8; color:#fff; border-color:#005eb8}
    .grid{display:grid; gap:12px; grid-template-columns:repeat(2,minmax(0,1fr)); margin-top:12px}
    .card{background:var(--card); border:1px solid var(--line); border-radius:10px; padding:12px}
    .card h3{margin:0 0 8px 0}
    input,button{padding:8px 10px; border-radius:8px; border:1px solid var(--line)}
    input{background:#fff}
    button{cursor:pointer}
    button.primary{background:var(--blue); color:#fff; border-color:var(--blue)}
    button.danger{background:#fff; color:var(--red); border-color:var(--red)}
    .pill{padding:4px 8px; border:1px solid var(--line); border-radius:999px; font-size:12px}
    .ok{color:var(--green)}
    .err{color:var(--red)}
    pre{
      margin:8px 0 0 0;
      background:#f8fafc;
      border:1px solid var(--line);
      border-radius:8px;
      padding:10px;
      white-space:pre-wrap;
      max-height:220px;
      overflow:auto;
      font-size:12px;
      line-height:1.35;
    }
    .muted{color:var(--muted); font-size:12px}
    @media (max-width:1100px){ .grid{grid-template-columns:1fr;} }
  </style>
</head>
<body>
  <div class="wrap">
    __NAV__

    <div class="grid">
      <section class="card">
        <h3>RASPI-PLC</h3>
        <div class="muted">Manual input goes directly to Microtom. Multi-send: separate with comma, semicolon or new line.</div>
        <div class="row" style="margin-top:8px">
          <label>ParamType hint</label>
          <input id="hint_raspi" style="width:90px" placeholder="optional" value=""/>
          <input id="cmd_raspi" style="flex:1; min-width:260px" placeholder="e.g. TTP00002=23, TTP00003=10 or MAP0001=500"/>
          <button class="primary" onclick="sendFrom('raspi')">Send</button>
          <button onclick="clearOutput('raspi')">Clear Output</button>
          <span id="st_raspi" class="pill"></span>
        </div>
        <pre id="out_raspi"></pre>
        <div class="row" style="margin-top:8px">
          <button onclick="loadLogs('raspi')">Reload Log</button>
          <button onclick="downloadLog('raspi')">Download Log</button>
          <button class="danger" onclick="clearLog('raspi')">Clear Log</button>
          <span id="logst_raspi" class="pill"></span>
        </div>
        <pre id="log_raspi"></pre>
      </section>

      <section class="card">
        <h3>ESP-PLC</h3>
        <div class="muted">Manual input goes ESP-PLC -> RASPI -> Microtom. Multi-send supported.</div>
        <div class="row" style="margin-top:8px">
          <label>ParamType hint</label>
          <input id="hint_esp_plc" style="width:90px" value="MAS"/>
          <input id="cmd_esp_plc" style="flex:1; min-width:260px" placeholder="e.g. 0026=20, 0027=11 or MAP0001=500"/>
          <button class="primary" onclick="sendFrom('esp-plc')">Send</button>
          <button onclick="clearOutput('esp-plc')">Clear Output</button>
          <span id="st_esp_plc" class="pill"></span>
        </div>
        <pre id="out_esp_plc"></pre>
        <div class="row" style="margin-top:8px">
          <button onclick="loadLogs('esp-plc')">Reload Log</button>
          <button onclick="downloadLog('esp-plc')">Download Log</button>
          <button class="danger" onclick="clearLog('esp-plc')">Clear Log</button>
          <span id="logst_esp_plc" class="pill"></span>
        </div>
        <pre id="log_esp_plc"></pre>
      </section>

      <section class="card">
        <h3>VJ3350 (Laser)</h3>
        <div class="muted">Manual input goes VJ3350 -> RASPI -> Microtom. Multi-send supported.</div>
        <div class="row" style="margin-top:8px">
          <label>ParamType hint</label>
          <input id="hint_vj3350" style="width:90px" value="LSE"/>
          <input id="cmd_vj3350" style="flex:1; min-width:260px" placeholder="e.g. 1000=1; 1001=0 or LSW1000=1"/>
          <button class="primary" onclick="sendFrom('vj3350')">Send</button>
          <button onclick="clearOutput('vj3350')">Clear Output</button>
          <span id="st_vj3350" class="pill"></span>
        </div>
        <pre id="out_vj3350"></pre>
        <div class="row" style="margin-top:8px">
          <button onclick="loadLogs('vj3350')">Reload Log</button>
          <button onclick="downloadLog('vj3350')">Download Log</button>
          <button class="danger" onclick="clearLog('vj3350')">Clear Log</button>
          <span id="logst_vj3350" class="pill"></span>
        </div>
        <pre id="log_vj3350"></pre>
      </section>

      <section class="card">
        <h3>VJ6530 (TTO)</h3>
        <div class="muted">Manual input goes VJ6530 -> RASPI -> Microtom. Multi-send supported.</div>
        <div class="row" style="margin-top:8px">
          <label>ParamType hint</label>
          <input id="hint_vj6530" style="width:90px" value="TTE"/>
          <input id="cmd_vj6530" style="flex:1; min-width:260px" placeholder="e.g. TTP00002=23, TTP00003=10"/>
          <button class="primary" onclick="sendFrom('vj6530')">Send</button>
          <button onclick="clearOutput('vj6530')">Clear Output</button>
          <span id="st_vj6530" class="pill"></span>
        </div>
        <pre id="out_vj6530"></pre>
        <div class="row" style="margin-top:8px">
          <button onclick="loadLogs('vj6530')">Reload Log</button>
          <button onclick="downloadLog('vj6530')">Download Log</button>
          <button class="danger" onclick="clearLog('vj6530')">Clear Log</button>
          <span id="logst_vj6530" class="pill"></span>
        </div>
        <pre id="log_vj6530"></pre>
      </section>
    </div>
  </div>

<script>
const TOKEN_KEY = "mas004_ui_token";
const SOURCES = ["raspi","esp-plc","vj3350","vj6530"];
const AUTO_LOG_MS = 2000;
let autoLogTimer = null;

function sid(source){ return String(source||"").replace(/-/g, "_"); }
function el(id){ return document.getElementById(id); }

function cookieGet(name){
  const m = document.cookie.match(new RegExp('(?:^|; )' + name.replace(/[-.$?*|{}()\\[\\]\\\\\\/\\+^]/g,'\\\\$&') + '=([^;]*)'));
  return m ? decodeURIComponent(m[1]) : "";
}
function getToken(){
  try{
    return localStorage.getItem(TOKEN_KEY) || cookieGet(TOKEN_KEY) || "";
  }catch(e){
    return cookieGet(TOKEN_KEY) || "";
  }
}
async function api(path, opt={}){
  opt.headers = opt.headers || {};
  const t = getToken();
  if(t) opt.headers["X-Token"] = t;
  const r = await fetch(path, opt);
  const txt = await r.text();
  let j = null;
  try{ j = JSON.parse(txt); }catch(e){}
  if(!r.ok){
    throw new Error((j && j.detail) ? j.detail : ("HTTP " + r.status + " " + txt));
  }
  return j;
}
function displayTs(value){ return String(value || "").trim() || "time-n/a"; }
function setStatus(source, msg, isErr=false){
  const node = el(`st_${sid(source)}`);
  node.textContent = msg || "";
  node.className = "pill " + (isErr ? "err" : "ok");
}
function setLogStatus(source, msg, isErr=false){
  const node = el(`logst_${sid(source)}`);
  node.textContent = msg || "";
  node.className = "pill " + (isErr ? "err" : "ok");
}
function appendOutput(source, line){
  const node = el(`out_${sid(source)}`);
  const existing = node.textContent || "";
  node.textContent = line + "\\n" + existing;
  node.scrollTop = 0;
}
function clearOutput(source){
  el(`out_${sid(source)}`).textContent = "";
}
function formatLogs(items){
  return items.map(it => {
    const t = displayTs(it.ts_display);
    const dir = String(it.direction || "").toUpperCase();
    return `[${t}] ${dir} ${it.message || ""}`;
  }).join("\\n");
}
async function sendFrom(source){
  const s = sid(source);
  const cmdEl = el(`cmd_${s}`);
  const hintEl = el(`hint_${s}`);
  const msg = (cmdEl.value || "").trim();
  if(!msg){
    setStatus(source, "empty", true);
    return;
  }
  setStatus(source, "sending...");
  try{
    const payload = {
      source: source,
      msg: msg,
      ptype_hint: (hintEl && hintEl.value) ? hintEl.value.trim() : ""
    };
    const j = await api("/api/test/send", {
      method: "POST",
      headers: {"Content-Type":"application/json"},
      body: JSON.stringify(payload)
    });
    const items = Array.isArray(j.items) && j.items.length ? j.items : [j];
    for(const it of items){
      const line = it.line || msg;
      const route = it.route || (source === "raspi" ? "raspi->microtom" : `${source}->raspi->microtom`);
      const ack = it.ack || "ACK_QUEUED";
      const idem = it.idempotency_key || "-";
      const t = displayTs(it.ts_display);
      appendOutput(source, `[${t}] ${route}: ${line} (${ack}, idem=${idem})`);
      if(source !== "raspi"){
        appendOutput("raspi", `[${t}] incoming from ${source}: ${line}`);
      }
    }
    setStatus(source, items.length > 1 ? `ok (${items.length})` : "ok");
    await Promise.all([loadLogs(source), loadLogs("raspi")]);
  }catch(e){
    setStatus(source, "ERROR: " + e.message, true);
  }
}
async function loadLogs(source, silent=false){
  if(!silent) setLogStatus(source, "loading...");
  try{
    const j = await api(`/api/ui/logs?channel=${encodeURIComponent(source)}&limit=350`);
    el(`log_${sid(source)}`).textContent = formatLogs(j.items || []);
    if(!silent) setLogStatus(source, "ok");
  }catch(e){
    setLogStatus(source, "ERROR: " + e.message, true);
  }
}
async function clearLog(source){
  if(!confirm("Clear log: " + source + " ?")) return;
  try{
    await api(`/api/ui/logs/clear?channel=${encodeURIComponent(source)}`, {method:"POST"});
    await loadLogs(source);
  }catch(e){
    setLogStatus(source, "ERROR: " + e.message, true);
  }
}
async function downloadLog(source){
  const t = getToken();
  const r = await fetch(`/api/ui/logs/download?channel=${encodeURIComponent(source)}`, {headers: t ? {"X-Token":t} : {}});
  if(!r.ok){
    alert(await r.text());
    return;
  }
  const blob = await r.blob();
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = source + ".log";
  a.click();
  URL.revokeObjectURL(a.href);
}
async function reloadAll(silent=false){
  const jobs = SOURCES.map(src => loadLogs(src, silent));
  await Promise.all(jobs);
}

function startAutoLogRefresh(){
  if(autoLogTimer) return;
  autoLogTimer = setInterval(() => {
    if(document.hidden) return;
    reloadAll(true);
  }, AUTO_LOG_MS);
}

reloadAll();
startAutoLogRefresh();
document.addEventListener("visibilitychange", () => {
  if(!document.hidden){
    reloadAll(true);
  }
});
</script>
</body></html>
""".replace("__NAV__", nav)

    return app
