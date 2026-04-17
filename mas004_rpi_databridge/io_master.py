from __future__ import annotations

import hashlib
import json
import os
import re
from typing import Any, Dict, List, Optional

import openpyxl

from mas004_rpi_databridge.db import DB, now_ts


_DEVICE_SHEETS = {
    "raspberry plc 21 pinout": ("raspi_plc21", "Raspberry PLC 21"),
    "esp32 plc 58 pinout": ("esp32_plc58", "ESP32 PLC 58"),
    "iologik e1211 pinout modul 1": ("moxa_e1211_1", "Moxa ioLogik E1211 #1"),
    "iologik e1211 pinout modul 2": ("moxa_e1211_2", "Moxa ioLogik E1211 #2"),
}


def _to_clean_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _normalize_sheet_name(sheet_name: str) -> str:
    return re.sub(r"\s+", " ", (sheet_name or "").strip().lower())


def _sheet_descriptor(sheet_name: str) -> tuple[str, str]:
    key = _normalize_sheet_name(sheet_name)
    if key not in _DEVICE_SHEETS:
        raise RuntimeError(f"Unsupported IO sheet '{sheet_name}'")
    return _DEVICE_SHEETS[key]


def _normalize_pin_label(pin_label: str) -> str:
    raw = _to_clean_str(pin_label).upper().replace(" ", "")
    raw = raw.replace(",", ".")
    return raw


def _detect_io_dir(pin_label: str) -> str:
    pin = _normalize_pin_label(pin_label)
    if pin.startswith(("I", "DI")):
        return "input"
    if pin.startswith(("Q", "DO", "R")):
        return "output"
    if pin.startswith("GPIO"):
        return "gpio"
    return "unknown"


def _channel_no(pin_label: str) -> Optional[int]:
    pin = _normalize_pin_label(pin_label)
    match = re.search(r"(\d+)$", pin)
    if not match:
        return None
    try:
        return int(match.group(1))
    except Exception:
        return None


def _safe_io_key(device_code: str, pin_label: str) -> str:
    pin = re.sub(r"[^A-Z0-9]+", "_", _normalize_pin_label(pin_label)).strip("_")
    return f"{device_code}__{pin}"


def _is_reserved(function_text: str) -> bool:
    txt = _to_clean_str(function_text).lower()
    return not txt or "reserve" in txt


def _sha256_file(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


class IoStore:
    def __init__(self, db: DB):
        self.db = db

    def import_xlsx(self, file_path: str) -> Dict[str, Any]:
        if not os.path.exists(file_path):
            raise RuntimeError(f"IO workbook not found: {file_path}")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        rows: List[Dict[str, Any]] = []
        sheet_counts: Dict[str, int] = {}

        for sheet_name in workbook.sheetnames:
            device_code, device_label = _sheet_descriptor(sheet_name)
            ws = workbook[sheet_name]
            zone_label = ""

            for row_idx in range(1, ws.max_row + 1):
                pin_raw = _to_clean_str(ws.cell(row_idx, 1).value)
                function_text = _to_clean_str(ws.cell(row_idx, 2).value)
                if not pin_raw and not function_text:
                    continue

                lower_pin = pin_raw.lower()
                if lower_pin in {"plc pinout", "plc pinout:", "pinout"}:
                    continue
                if lower_pin.startswith("zone "):
                    zone_label = pin_raw
                    continue
                if lower_pin in {"function", "description"}:
                    continue
                if not pin_raw:
                    continue

                pin_label = _normalize_pin_label(pin_raw)
                entry = {
                    "io_key": _safe_io_key(device_code, pin_label),
                    "device_code": device_code,
                    "device_label": device_label,
                    "sheet_name": sheet_name,
                    "zone_label": zone_label,
                    "pin_label": pin_label,
                    "io_dir": _detect_io_dir(pin_label),
                    "channel_no": _channel_no(pin_label),
                    "function_text": function_text,
                    "is_reserved": 1 if _is_reserved(function_text) else 0,
                    "is_active": 1,
                    "source_row": row_idx,
                    "updated_ts": now_ts(),
                }
                rows.append(entry)
                sheet_counts[sheet_name] = sheet_counts.get(sheet_name, 0) + 1

        if not rows:
            raise RuntimeError("No IO rows found in workbook")

        imported_ts = now_ts()
        source_filename = os.path.basename(file_path)
        source_sha256 = _sha256_file(file_path)
        file_mtime_ts = float(os.path.getmtime(file_path))
        notes_json = json.dumps({"sheet_counts": sheet_counts}, ensure_ascii=True, sort_keys=True)

        with self.db._conn() as conn:
            conn.execute("DELETE FROM io_points")
            for row in rows:
                conn.execute(
                    """INSERT INTO io_points(
                       io_key, device_code, device_label, sheet_name, zone_label, pin_label, io_dir,
                       channel_no, function_text, is_reserved, is_active, source_row, updated_ts
                       ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        row["io_key"],
                        row["device_code"],
                        row["device_label"],
                        row["sheet_name"],
                        row["zone_label"],
                        row["pin_label"],
                        row["io_dir"],
                        row["channel_no"],
                        row["function_text"],
                        row["is_reserved"],
                        row["is_active"],
                        row["source_row"],
                        row["updated_ts"],
                    ),
                )
                conn.execute(
                    """INSERT INTO io_values(io_key, value, quality, source, updated_ts)
                       VALUES(?,?,?,?,?)
                       ON CONFLICT(io_key) DO NOTHING""",
                    (row["io_key"], "0", "default", "import", imported_ts),
                )

            conn.execute("DELETE FROM io_values WHERE io_key NOT IN (SELECT io_key FROM io_points)")
            conn.execute(
                """INSERT INTO io_master_meta(
                   singleton_id, source_path, source_filename, source_sha256, imported_ts, file_mtime_ts,
                   channel_count, notes_json
                   ) VALUES(1,?,?,?,?,?,?,?)
                   ON CONFLICT(singleton_id) DO UPDATE SET
                     source_path=excluded.source_path,
                     source_filename=excluded.source_filename,
                     source_sha256=excluded.source_sha256,
                     imported_ts=excluded.imported_ts,
                     file_mtime_ts=excluded.file_mtime_ts,
                     channel_count=excluded.channel_count,
                     notes_json=excluded.notes_json""",
                (
                    file_path,
                    source_filename,
                    source_sha256,
                    imported_ts,
                    file_mtime_ts,
                    len(rows),
                    notes_json,
                ),
            )

        return {
            "ok": True,
            "channels": len(rows),
            "devices": len({row["device_code"] for row in rows}),
            "sheet_counts": sheet_counts,
            "master_workbook": self.master_info(),
        }

    def count_points(self) -> int:
        with self.db._conn() as conn:
            row = conn.execute("SELECT COUNT(*) FROM io_points").fetchone()
        return int(row[0] or 0)

    def master_info(self) -> Dict[str, Any]:
        with self.db._conn() as conn:
            row = conn.execute(
                """SELECT source_path, source_filename, source_sha256, imported_ts, file_mtime_ts,
                          channel_count, notes_json
                   FROM io_master_meta WHERE singleton_id=1"""
            ).fetchone()
        if not row:
            return {
                "path": "",
                "filename": "",
                "sha256": "",
                "imported_ts": 0.0,
                "file_mtime_ts": 0.0,
                "channel_count": 0,
                "notes": {},
                "exists": False,
            }
        notes = {}
        try:
            notes = json.loads(row[6] or "{}")
        except Exception:
            notes = {}
        return {
            "path": row[0],
            "filename": row[1],
            "sha256": row[2],
            "imported_ts": float(row[3] or 0.0),
            "file_mtime_ts": float(row[4] or 0.0),
            "channel_count": int(row[5] or 0),
            "notes": notes,
            "exists": bool(row[0] and os.path.exists(str(row[0]))),
        }

    def get_point(self, io_key: str) -> Optional[Dict[str, Any]]:
        with self.db._conn() as conn:
            row = conn.execute(
                """SELECT p.io_key,p.device_code,p.device_label,p.sheet_name,p.zone_label,p.pin_label,p.io_dir,
                          p.channel_no,p.function_text,p.is_reserved,p.is_active,p.source_row,
                          v.value,v.quality,v.source,v.updated_ts
                   FROM io_points p
                   LEFT JOIN io_values v ON v.io_key = p.io_key
                   WHERE p.io_key=?""",
                (io_key,),
            ).fetchone()
        return self._row_to_point(row) if row else None

    def list_points(self, device_code: Optional[str] = None, include_reserved: bool = True) -> List[Dict[str, Any]]:
        args: List[Any] = []
        where: List[str] = []
        if device_code:
            where.append("p.device_code=?")
            args.append(device_code)
        if not include_reserved:
            where.append("p.is_reserved=0")
        where_sql = f"WHERE {' AND '.join(where)}" if where else ""
        with self.db._conn() as conn:
            rows = conn.execute(
                f"""SELECT p.io_key,p.device_code,p.device_label,p.sheet_name,p.zone_label,p.pin_label,p.io_dir,
                           p.channel_no,p.function_text,p.is_reserved,p.is_active,p.source_row,
                           v.value,v.quality,v.source,v.updated_ts
                    FROM io_points p
                    LEFT JOIN io_values v ON v.io_key = p.io_key
                    {where_sql}
                    ORDER BY p.device_code ASC, p.source_row ASC""",
                args,
            ).fetchall()
        return [self._row_to_point(row) for row in rows]

    def list_devices(self) -> List[Dict[str, Any]]:
        with self.db._conn() as conn:
            rows = conn.execute(
                """SELECT p.device_code,
                          p.device_label,
                          COUNT(*) AS total_points,
                          SUM(CASE WHEN p.io_dir='input' THEN 1 ELSE 0 END) AS inputs,
                          SUM(CASE WHEN p.io_dir='output' THEN 1 ELSE 0 END) AS outputs,
                          SUM(CASE WHEN p.io_dir='gpio' THEN 1 ELSE 0 END) AS gpio_points,
                          SUM(CASE WHEN p.is_reserved=1 THEN 1 ELSE 0 END) AS reserved_points
                   FROM io_points p
                   GROUP BY p.device_code, p.device_label
                   ORDER BY p.device_code ASC"""
            ).fetchall()
        return [
            {
                "device_code": row[0],
                "device_label": row[1],
                "total_points": int(row[2] or 0),
                "inputs": int(row[3] or 0),
                "outputs": int(row[4] or 0),
                "gpio_points": int(row[5] or 0),
                "reserved_points": int(row[6] or 0),
            }
            for row in rows
        ]

    def upsert_value(self, io_key: str, value: Any, quality: str, source: str) -> bool:
        text_value = str(int(value)) if isinstance(value, bool) else str(value)
        quality_txt = _to_clean_str(quality) or "unknown"
        source_txt = _to_clean_str(source)
        ts = now_ts()
        with self.db._conn() as conn:
            current = conn.execute(
                "SELECT value, quality, source FROM io_values WHERE io_key=?",
                (io_key,),
            ).fetchone()
            if current and current[0] == text_value and current[1] == quality_txt and (current[2] or "") == source_txt:
                return False
            conn.execute(
                """INSERT INTO io_values(io_key, value, quality, source, updated_ts)
                   VALUES(?,?,?,?,?)
                   ON CONFLICT(io_key) DO UPDATE SET
                     value=excluded.value,
                     quality=excluded.quality,
                     source=excluded.source,
                     updated_ts=excluded.updated_ts""",
                (io_key, text_value, quality_txt, source_txt, ts),
            )
        return True

    def _row_to_point(self, row) -> Dict[str, Any]:
        return {
            "io_key": row[0],
            "device_code": row[1],
            "device_label": row[2],
            "sheet_name": row[3],
            "zone_label": row[4] or "",
            "pin_label": row[5],
            "io_dir": row[6],
            "channel_no": row[7],
            "function_text": row[8] or "",
            "is_reserved": bool(row[9]),
            "is_active": bool(row[10]),
            "source_row": int(row[11] or 0),
            "value": row[12] if row[12] is not None else "0",
            "quality": row[13] or "unknown",
            "source": row[14] or "",
            "updated_ts": float(row[15] or 0.0),
        }
