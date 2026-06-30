from __future__ import annotations

import json
import os
import re
import time
from typing import Any, Optional

import openpyxl

from mas004_rpi_databridge.config import Settings
from mas004_rpi_databridge.params import ParamStore, SHEET_NAME, motor_param_master_write_context


DEFAULT_MOTOR_PARAM_BINDINGS: dict[int, dict[str, str]] = {
    1: {"setpoint": "MAP0056", "actual": "MAS0011"},
    2: {"setpoint": "MAP0057", "actual": "MAS0012"},
    3: {"setpoint": "MAP0058", "actual": "MAS0031"},
    4: {"setpoint": "MAP0059", "actual": "MAS0032"},
    5: {"setpoint": "MAP0060", "actual": "MAS0017"},
    6: {"setpoint": "MAP0061", "actual": "MAS0013"},
    7: {"setpoint": "MAP0062", "actual": "MAS0014"},
    8: {"setpoint": "MAP0063", "actual": "MAS0016"},
    9: {"setpoint": "MAP0064", "actual": "MAS0015"},
}
POSITIONAL_MOTOR_IDS = frozenset({1, 2, 4, 5, 6, 7, 8, 9})

MOTOR_CONFIG_MASTER_KEYS = (
    "steps_per_mm",
    "speed_mm_s",
    "accel_mm_s2",
    "decel_mm_s2",
    "current_pct",
    "hold_current_pct",
    "invert_direction",
    "zero_offset_steps",
    "min_tenths_mm",
    "max_tenths_mm",
    "min_enabled",
    "max_enabled",
)

POSITIONAL_MOTOR_IDS = frozenset({1, 2, 4, 5, 6, 7, 8, 9})


def _to_int_or_none(value: Any) -> Optional[int]:
    try:
        return int(round(float(str(value).strip())))
    except Exception:
        return None


def _to_float_or_none(value: Any) -> Optional[float]:
    try:
        return float(str(value).strip())
    except Exception:
        return None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _cell_text(value).lower()).strip("_")


def _col_any(header_map: dict[str, int], *names: str) -> Optional[int]:
    for name in names:
        idx = header_map.get(_norm_header(name))
        if idx is not None:
            return idx
    return None


def _row_pkey(ptype: Any, pid: Any) -> str:
    ptype_text = _cell_text(ptype).upper()
    pid_text = _cell_text(pid)
    if pid_text and pid_text.isdigit() and len(pid_text) < 4:
        pid_text = pid_text.zfill(4)
    return f"{ptype_text}{pid_text}"


def _format_default(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and not value.is_integer():
        return f"{value:.6f}".rstrip("0").rstrip(".")
    return str(int(round(float(value))))


def _motor_binding(bindings: dict[int, dict[str, Any]] | dict[str, Any], motor_id: int) -> dict[str, Any]:
    if not isinstance(bindings, dict):
        return {}
    direct = bindings.get(int(motor_id))
    if isinstance(direct, dict):
        return direct
    as_text = bindings.get(str(int(motor_id)))
    return as_text if isinstance(as_text, dict) else {}


def _pkey_for_role(motor_id: int, binding: dict[str, Any], role: str) -> Optional[str]:
    item = binding.get(role)
    if not isinstance(item, dict):
        return DEFAULT_MOTOR_PARAM_BINDINGS.get(int(motor_id), {}).get(role)
    pkey = _cell_text(item.get("pkey"))
    return pkey or DEFAULT_MOTOR_PARAM_BINDINGS.get(int(motor_id), {}).get(role)


def _unique_existing(paths: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in paths:
        raw_text = str(raw or "").strip()
        if not raw_text:
            continue
        path = os.path.abspath(raw_text)
        if path in seen or not os.path.isfile(path):
            continue
        seen.add(path)
        out.append(path)
    return out


def load_motor_setup_master(
    params: ParamStore,
    motor_ids: Any = None,
) -> dict[int, dict[str, Any]]:
    wanted = {int(mid) for mid in motor_ids} if motor_ids is not None else None
    snapshots: dict[int, dict[str, Any]] = {}
    try:
        with params.db._conn() as c:
            rows = c.execute(
                "SELECT motor_id, config_json, state_json, updated_ts FROM motor_setup_master"
            ).fetchall()
    except Exception:
        return snapshots
    for motor_id, config_json, state_json, updated_ts in rows:
        mid = int(motor_id)
        if wanted is not None and mid not in wanted:
            continue
        try:
            config = json.loads(config_json or "{}")
        except Exception:
            config = {}
        try:
            state = json.loads(state_json or "{}")
        except Exception:
            state = {}
        snapshots[mid] = {
            "motor_id": mid,
            "config": config if isinstance(config, dict) else {},
            "state": state if isinstance(state, dict) else {},
            "updated_ts": float(updated_ts or 0.0),
        }
    return snapshots


def motor_master_config_payload(
    snapshot: dict[str, Any],
    *,
    include_zero_offset_steps: bool = True,
) -> dict[str, Any]:
    config = snapshot.get("config") if isinstance(snapshot, dict) else {}
    if not isinstance(config, dict):
        return {}
    payload = {key: config[key] for key in MOTOR_CONFIG_MASTER_KEYS if key in config}
    if not include_zero_offset_steps:
        payload.pop("zero_offset_steps", None)
    return payload


def _config_values_equal(expected: Any, actual: Any) -> bool:
    if isinstance(expected, bool):
        if isinstance(actual, bool):
            return expected is actual
        text = str(actual).strip().lower()
        if text in ("1", "true", "yes", "on"):
            return expected is True
        if text in ("0", "false", "no", "off"):
            return expected is False
        return False
    if isinstance(expected, int) and not isinstance(expected, bool):
        return _to_int_or_none(actual) == int(expected)
    if isinstance(expected, float):
        value = _to_float_or_none(actual)
        return value is not None and abs(value - float(expected)) <= 0.0005
    return str(expected) == str(actual)


def motor_master_config_drift(
    expected_payload: dict[str, Any],
    current_config: dict[str, Any],
) -> list[str]:
    if not isinstance(expected_payload, dict) or not expected_payload:
        return []
    if not isinstance(current_config, dict):
        return sorted(expected_payload)
    drift: list[str] = []
    for key, expected in expected_payload.items():
        if key not in current_config or not _config_values_equal(expected, current_config.get(key)):
            drift.append(key)
    return sorted(drift)


def motor_master_position_mm(snapshot: dict[str, Any]) -> Optional[float]:
    state = snapshot.get("state") if isinstance(snapshot, dict) else {}
    if not isinstance(state, dict):
        return None
    feedback = _to_int_or_none(state.get("feedback_tenths_mm"))
    if feedback is None:
        return None
    return float(feedback) / 10.0


def apply_motor_setup_master_config_to_client(
    params: ParamStore,
    client: Any,
    motor_id: int,
    *,
    restore_position: bool = True,
    allow_machine_setup_write: bool = False,
) -> dict[str, Any]:
    motor_id = int(motor_id)
    if motor_id in POSITIONAL_MOTOR_IDS:
        return {
            "ok": True,
            "restored": False,
            "position_restored": False,
            "config_changed": False,
            "drift_keys": [],
            "blocked": True,
            "reason": "motor_setup_master_restore_disabled_for_position_axis",
            "policy": (
                "Positionsachsen ID1/2/4-9 duerfen Nullpunkt/Min/Max/Skalierung "
                "nie aus DB/Excel/Master-Snapshots auf den ESP zurueckgeschrieben bekommen. "
                "Einzige Schreibquelle ist /ui/machine-setup/motors."
            ),
        }
    snapshot = load_motor_setup_master(params, [motor_id]).get(motor_id)
    payload = motor_master_config_payload(
        snapshot or {},
        include_zero_offset_steps=bool(motor_id != 3),
    )
    if motor_id == 3:
        # ID3 is calibrated online against the real infeed encoder.  A stale
        # Motor-Setup/Excel snapshot must never push an old transport-axis
        # scale back into the ESP after the calibration saved it in NVS.
        payload.pop("steps_per_mm", None)
    if not payload:
        return {"ok": True, "restored": False, "reason": "no_motor_setup_master_snapshot"}

    current_config: dict[str, Any] = {}
    current_config_error = ""
    if hasattr(client, "config"):
        try:
            current_payload = client.config(motor_id)
            raw_config = current_payload.get("config") if isinstance(current_payload, dict) else {}
            current_config = raw_config if isinstance(raw_config, dict) else {}
        except Exception as exc:
            current_config_error = str(exc)
            current_config = {}

    drift_keys = (
        motor_master_config_drift(payload, current_config)
        if current_config or not current_config_error
        else sorted(payload)
    )
    set_result: dict[str, Any] = {"ok": True, "reply": "MASTER_CONFIG_ALREADY_CURRENT"}
    save_result: dict[str, Any] | None = None
    if drift_keys:
        try:
            set_result = client.set_config(
                int(motor_id),
                payload,
                allow_machine_setup_write=allow_machine_setup_write,
            )
        except TypeError:
            set_result = client.set_config(int(motor_id), payload)
        if not isinstance(set_result, dict) or not bool(set_result.get("ok", True)):
            return {
                "ok": False,
                "restored": False,
                "action": "SET_CONFIG",
                "drift_keys": drift_keys,
                "reply": (set_result or {}).get("reply") if isinstance(set_result, dict) else str(set_result),
            }
        try:
            save_result = client.save(
                int(motor_id),
                allow_machine_setup_write=allow_machine_setup_write,
            )
        except TypeError:
            save_result = client.save(int(motor_id))
        if not isinstance(save_result, dict) or not bool(save_result.get("ok", True)):
            return {
                "ok": False,
                "restored": False,
                "action": "SAVE",
                "drift_keys": drift_keys,
                "reply": (save_result or {}).get("reply") if isinstance(save_result, dict) else str(save_result),
            }
    position_mm = motor_master_position_mm(snapshot or {})
    replies = [set_result.get("reply", "")]
    if save_result is not None:
        replies.append(save_result.get("reply", ""))
    return {
        "ok": True,
        "restored": True,
        "payload": payload,
        "position_mm": position_mm,
        "position_restored": False,
        "config_changed": bool(drift_keys),
        "drift_keys": drift_keys,
        "current_config_error": current_config_error,
        "reply": "; ".join(str(item).strip() for item in replies if str(item or "").strip()),
    }


def reapply_motor_setup_master_to_params(
    params: ParamStore,
    cfg: Settings,
    bindings: dict[int, dict[str, Any]] | dict[str, Any],
    *,
    workbook_paths: Optional[list[str]] = None,
    motor_ids: Any = None,
) -> dict[str, Any]:
    """
    Restore Motor-Setup authority after importing older parameter workbooks.

    The commissioning UI is the master for motor parameters.  If a user imports
    an Excel file with stale MAP/MAS defaults or limits, we immediately overlay
    the stored Motor-Setup snapshot again so the workbook/DB import cannot
    silently roll back commissioned axes.
    """

    snapshots = load_motor_setup_master(params, motor_ids)
    results: list[dict[str, Any]] = []
    for motor_id in sorted(snapshots):
        snapshot = snapshots[motor_id]
        motor = {"config": snapshot.get("config") or {}, "state": snapshot.get("state") or {}}
        results.append(
            sync_motor_master_values(
                params,
                cfg,
                int(motor_id),
                motor,
                bindings,
                workbook_paths=workbook_paths,
                sync_position_default=True,
                allow_protected_position_param_write=False,
            )
        )
    return {
        "ok": all(bool(item.get("ok")) for item in results) if results else True,
        "count": len(results),
        "results": results,
    }


def _update_workbook(path: str, updates: dict[str, dict[str, Any]]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    headers = {_norm_header(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    col_type = _col_any(headers, "Params_Type.", "Params_Type.:", "Params Type", "Params_Type")
    col_id = _col_any(headers, "Param ID.", "Param. ID.:", "Param ID", "Param_ID", "Param. ID")
    col_min = _col_any(headers, "Min.", "Min.:", "Min")
    col_max = _col_any(headers, "Max.", "Max.:", "Max")
    col_default = _col_any(headers, "Default Value", "Default Value:", "Default")
    if not col_type or not col_id:
        return {"path": path, "updated": 0, "error": "missing parameter id columns"}

    changed = 0
    for row in range(2, ws.max_row + 1):
        pkey = _row_pkey(ws.cell(row, col_type).value, ws.cell(row, col_id).value)
        update = updates.get(pkey)
        if not update:
            continue
        if col_min and update.get("min_v") is not None:
            ws.cell(row, col_min).value = update["min_v"]
            changed += 1
        if col_max and update.get("max_v") is not None:
            ws.cell(row, col_max).value = update["max_v"]
            changed += 1
        if col_default and update.get("default_v") is not None:
            ws.cell(row, col_default).value = update["default_v"]
            changed += 1

    if changed:
        wb.save(path)
    return {"path": path, "updated": changed}


def sync_motor_master_values(
    params: ParamStore,
    cfg: Settings,
    motor_id: int,
    motor: dict[str, Any],
    bindings: dict[int, dict[str, Any]] | dict[str, Any],
    *,
    workbook_paths: Optional[list[str]] = None,
    sync_position_default: bool = True,
    allow_protected_position_param_write: bool = False,
) -> dict[str, Any]:
    """
    Persist the live Motor-Setup values as the project master.

    The motor setup page is the authoritative commissioning source for ID1-9.
    A successful save therefore updates the runtime DB and the master Excel
    files so an older import cannot silently reintroduce stale soft limits.
    """

    motor_id = int(motor_id)
    if motor_id in POSITIONAL_MOTOR_IDS and not allow_protected_position_param_write:
        return {
            "ok": True,
            "motor_id": motor_id,
            "updated_pkeys": [],
            "db_errors": {},
            "db_snapshot": "",
            "db_snapshot_error": "",
            "workbooks": [],
            "protected_position_param_write_blocked": True,
            "reason": "motor_setup_only_for_position_axis_params",
        }
    binding = _motor_binding(bindings, motor_id)
    state = motor.get("state") if isinstance(motor.get("state"), dict) else {}
    config = motor.get("config") if isinstance(motor.get("config"), dict) else {}
    previous_state: dict[str, Any] = {}
    if not sync_position_default:
        previous_snapshot = load_motor_setup_master(params, [motor_id]).get(motor_id) or {}
        previous_state = previous_snapshot.get("state") if isinstance(previous_snapshot.get("state"), dict) else {}
    feedback = _to_int_or_none(state.get("feedback_tenths_mm"))
    target = _to_int_or_none(state.get("target_tenths_mm"))
    command = _to_int_or_none(state.get("command_tenths_mm"))
    position_default_skipped = False
    valid_live_position = not (motor_id in POSITIONAL_MOTOR_IDS) or (
        bool(state.get("link_ok")) and feedback is not None
    )
    if motor_id in POSITIONAL_MOTOR_IDS and sync_position_default and not valid_live_position:
        # After ESP reboot or with motor polling disabled, STATUS? can contain
        # placeholder feedback=0/link_ok=false. That must never become the new
        # commissioned Motor-Setup default for position axes.
        sync_position_default = False
        position_default_skipped = True
    # For commissioned axis defaults the measured actual position is the
    # source of truth. ESP target/command can still contain stale order data
    # from a previous move and must not overwrite the Motor-Setup baseline.
    default_tenths = feedback if feedback is not None else (target if target is not None else command)
    actual_tenths = feedback if feedback is not None else default_tenths

    min_enabled = bool(config.get("min_enabled", motor_id != 3))
    max_enabled = bool(config.get("max_enabled", motor_id != 3))
    min_v = float(_to_int_or_none(config.get("min_tenths_mm"))) if min_enabled and _to_int_or_none(config.get("min_tenths_mm")) is not None else None
    max_v = float(_to_int_or_none(config.get("max_tenths_mm"))) if max_enabled and _to_int_or_none(config.get("max_tenths_mm")) is not None else None

    # Motor 3 is the transport axis. Its production speed/ramp remains recipe-
    # driven, but its actual position can still be mirrored as a status value.
    if motor_id == 3:
        min_v = None
        max_v = None

    updates: dict[str, dict[str, Any]] = {}
    setpoint_pkey = _pkey_for_role(motor_id, binding, "setpoint")
    actual_pkey = _pkey_for_role(motor_id, binding, "actual")
    if setpoint_pkey:
        updates[setpoint_pkey] = {
            "default_v": _format_default(default_tenths) if sync_position_default else None,
            "min_v": min_v,
            "max_v": max_v,
        }
    if actual_pkey:
        updates[actual_pkey] = {
            "default_v": _format_default(actual_tenths) if sync_position_default else None,
            "min_v": min_v,
            "max_v": max_v,
        }

    db_updates: dict[str, str] = {}
    db_errors: dict[str, str] = {}

    def write_param_updates():
        for pkey, update in updates.items():
            ok, msg = params.update_meta(
                pkey,
                default_v=update.get("default_v"),
                min_v=update.get("min_v"),
                max_v=update.get("max_v"),
            )
            if not ok:
                db_errors[pkey] = msg
                continue
            if update.get("default_v") is not None:
                params.apply_device_value(pkey, str(update["default_v"]), promote_default=True)
            db_updates[pkey] = "OK"

    if allow_protected_position_param_write:
        with motor_param_master_write_context(f"motor_setup_sync:{motor_id}"):
            write_param_updates()
    else:
        write_param_updates()

    db_snapshot_error = ""
    try:
        snapshot_state = state
        if not sync_position_default and previous_state:
            snapshot_state = dict(previous_state)
        with params.db._conn() as c:
            c.execute(
                """CREATE TABLE IF NOT EXISTS motor_setup_master (
                     motor_id INTEGER PRIMARY KEY,
                     config_json TEXT NOT NULL DEFAULT '{}',
                     state_json TEXT NOT NULL DEFAULT '{}',
                     updated_ts REAL NOT NULL
                   )"""
            )
            c.execute(
                """INSERT INTO motor_setup_master(motor_id,config_json,state_json,updated_ts)
                   VALUES(?,?,?,?)
                   ON CONFLICT(motor_id) DO UPDATE SET
                     config_json=excluded.config_json,
                     state_json=excluded.state_json,
                     updated_ts=excluded.updated_ts""",
                (
                    motor_id,
                    json.dumps(config, ensure_ascii=False, sort_keys=True),
                    json.dumps(snapshot_state, ensure_ascii=False, sort_keys=True),
                    time.time(),
                ),
            )
    except Exception as exc:
        db_snapshot_error = str(exc)

    paths = list(workbook_paths or [])
    if cfg.master_params_xlsx_path:
        paths.append(cfg.master_params_xlsx_path)
    workbook_results = [_update_workbook(path, updates) for path in _unique_existing(paths)]
    return {
        "ok": not db_errors and not db_snapshot_error,
        "motor_id": motor_id,
        "updated_pkeys": sorted(db_updates.keys()),
        "db_errors": db_errors,
        "db_snapshot": "OK" if not db_snapshot_error else "",
        "db_snapshot_error": db_snapshot_error,
        "workbooks": workbook_results,
        "position_default_skipped": position_default_skipped,
    }
