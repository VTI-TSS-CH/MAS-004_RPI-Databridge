import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from mas004_rpi_databridge.config import Settings
from mas004_rpi_databridge.db import DB
from mas004_rpi_databridge.io_master import IoStore
from mas004_rpi_databridge import io_runtime as io_runtime_module
from mas004_rpi_databridge.io_runtime import IoRuntime, _MoxaEndpointCooldown
from mas004_rpi_databridge.moxa_iologik import MoxaProtocolError


class IoMasterImportTests(unittest.TestCase):
    def build_workbook(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Raspberry PLC 21 PINOUT"
        ws.append(["PLC Pinout", "Function"])
        ws.append(["Raspberry PLC21+ Pinout Digital I/Os", ""])
        ws.append(["I0.6", "USV Status OK"])
        ws.append(["I0.7", "Taster Start/Pause / Reset"])
        ws.append(["Q0.0", "LED rot Taster Start/Pause"])

        ws = wb.create_sheet("ESP32 PLC 58 PINOUT")
        ws.append(["PLC Pinout", "Function"])
        ws.append(["ESP32 PLC58 Pinout Digital I/Os", ""])
        ws.append(["Zone A", ""])
        ws.append(["GPIO0", "Reserve GPIO0 statischer Test; LED-Streifen nutzt externen UDP-Controller"])
        ws.append(["I0.0", "TTO Drucker Status"])
        ws.append(["I0.7", "Not-Aus OK"])
        ws.append(["I0.8", "Lichtgitter OK"])

        ws = wb.create_sheet("IOLOGIK E1213 PINOUT Modul 1")
        ws.append(["PLC Pinout", "Function"])
        ws.append(["ioLogik E1213 Pinout Digital I/Os", ""])
        ws.append(["DO0", "Reserve 1 Motor Schutzblech Laser"])
        ws.append(["DIO0", "Reserve 1 Motor Etikettenanschlag rechts"])

        ws = wb.create_sheet("IOLOGIK E1213 PINOUT Modul 2")
        ws.append(["PLC Pinout", "Function"])
        ws.append(["DO0", "Start Motor X-Achse (IN0)"])
        ws.append(["DIO0", "Reserve 1 Motor Z-Achse"])

        ws = wb.create_sheet("IOLOGIK E1213 PINOUT Modul 3")
        ws.append(["PLC Pinout", "Function"])
        ws.append(["DIO0", "LED Maschinenstatus rot"])

        wb.save(path)

    def test_import_xlsx_reads_known_devices_and_reserved_flags(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            result = store.import_xlsx(str(workbook_path))

            self.assertTrue(result["ok"])
            self.assertEqual(12, result["channels"])
            self.assertEqual(5, result["devices"])

            points = store.list_points()
            by_key = {item["io_key"]: item for item in points}
            self.assertEqual("raspi_plc21", by_key["raspi_plc21__I0_6"]["device_code"])
            self.assertEqual("gpio", by_key["esp32_plc58__GPIO0"]["io_dir"])
            self.assertFalse(by_key["moxa_e1213_2__DO0"]["is_reserved"])
            self.assertFalse(by_key["moxa_e1213_3__DIO0"]["is_reserved"])
            self.assertEqual("output", by_key["moxa_e1213_3__DIO0"]["io_dir"])
            self.assertEqual(12, store.master_info()["channel_count"])

    def test_esp_simulation_keeps_safety_ok_inputs_high(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            store.upsert_value("esp32_plc58__I0_7", "0", "simulation", "test")
            store.upsert_value("esp32_plc58__I0_8", "0", "simulation", "test")
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    esp_simulation=True,
                    raspi_io_simulation=True,
                    moxa1_simulation=True,
                    moxa2_simulation=True,
                    moxa3_simulation=True,
                ),
                store,
            )

            runtime.refresh()

            self.assertEqual("1", store.get_point("esp32_plc58__I0_7")["value"])
            self.assertEqual("1", store.get_point("esp32_plc58__I0_8")["value"])

    def test_raspi_plc21_i07_to_i012_inputs_are_read_via_analog_threshold(self):
        class FakeRpiplc:
            INPUT = 0
            OUTPUT = 1
            HIGH = 1
            LOW = 0

            def __init__(self):
                self.modes = []
                self.digital_reads = []
                self.analog_reads = []

            def pin_mode(self, pin, mode):
                self.modes.append((pin, mode))

            def digital_read(self, pin):
                self.digital_reads.append(pin)
                return 1 if pin == "I0.6" else 0

            def analog_read(self, pin):
                self.analog_reads.append(pin)
                return 2300 if pin == "I0.7" else 0

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    raspi_io_simulation=False,
                    raspi_analog_input_high_threshold=1000.0,
                ),
                store,
            )
            fake = FakeRpiplc()

            with patch("mas004_rpi_databridge.io_runtime._ensure_rpiplc", return_value=(fake, "")):
                result = runtime.refresh()

            self.assertTrue(result["ok"])
            self.assertEqual("1", store.get_point("raspi_plc21__I0_6")["value"])
            self.assertEqual("1", store.get_point("raspi_plc21__I0_7")["value"])
            self.assertIn("I0.6", fake.digital_reads)
            self.assertIn("I0.7", fake.analog_reads)
            self.assertNotIn(("I0.7", fake.INPUT), fake.modes)
            self.assertNotIn("Q0.0", fake.digital_reads)
            self.assertNotIn(("Q0.0", fake.OUTPUT), fake.modes)

    def test_io_override_blocks_normal_runtime_writes_until_release(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    esp_simulation=True,
                    raspi_io_simulation=True,
                    moxa1_simulation=True,
                    moxa2_simulation=True,
                    moxa3_simulation=True,
                ),
                store,
            )

            io_key = "raspi_plc21__Q0_0"
            override_result = runtime.override_output(io_key, True, source="test-ui")
            self.assertTrue(override_result["override_active"])
            self.assertEqual("1", store.get_point(io_key)["override_value"])

            runtime_result = runtime.write_output(io_key, False)
            self.assertTrue(runtime_result["overridden"])
            point = store.get_point(io_key)
            self.assertEqual("1", point["value"])
            self.assertTrue(point["override_active"])

            force_result = runtime.write_output(io_key, False, force=True, source="safety-led")
            self.assertTrue(force_result["overridden"])
            point = store.get_point(io_key)
            self.assertEqual("1", point["value"])
            self.assertEqual("override", point["quality"])
            self.assertTrue(point["override_active"])

            store.upsert_value(io_key, "0", "live", "poller")
            point = store.get_point(io_key)
            self.assertEqual("1", point["value"])
            self.assertEqual("override", point["quality"])
            self.assertTrue(point["override_active"])

            release_result = runtime.release_override(io_key)
            self.assertTrue(release_result["released"])
            runtime.write_output(io_key, False)
            point = store.get_point(io_key)
            self.assertEqual("0", point["value"])
            self.assertFalse(point["override_active"])

    def test_esp_override_blocks_force_writer_and_reasserts_override_value(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    esp_host="127.0.0.1",
                    esp_port=3010,
                    esp_simulation=False,
                    raspi_io_simulation=True,
                    moxa1_simulation=True,
                    moxa2_simulation=True,
                    moxa3_simulation=True,
                ),
                store,
            )
            commands: list[str] = []

            class FakeEspClient:
                def __init__(self, host, port, timeout_s):
                    pass

                def exchange_line(self, line, **_kwargs):
                    commands.append(str(line))
                    return "ACK"

            with patch("mas004_rpi_databridge.io_runtime.EspPlcClient", FakeEspClient):
                runtime.override_output("esp32_plc58__GPIO0", True, source="test-ui")
                runtime.write_output("esp32_plc58__GPIO0", False, force=True, source="safety-reset")

            self.assertNotIn("IO SET GPIO0=0", commands)
            self.assertGreaterEqual(commands.count("IO SET GPIO0=1"), 2)
            point = store.get_point("esp32_plc58__GPIO0")
            self.assertEqual("1", point["value"])
            self.assertEqual("override", point["quality"])
            self.assertTrue(point["override_active"])

    def test_esp_refresh_reasserts_physical_output_when_override_snapshot_differs(self):
        io_runtime_module._ESP_IO_COOLDOWN_UNTIL = 0.0
        io_runtime_module._ESP_IO_COOLDOWN_ERROR = ""
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    esp_host="127.0.0.1",
                    esp_port=3010,
                    esp_simulation=False,
                    raspi_io_simulation=True,
                    moxa1_simulation=True,
                    moxa2_simulation=True,
                    moxa3_simulation=True,
                ),
                store,
            )
            commands: list[str] = []

            class FakeEspClient:
                def __init__(self, host, port, timeout_s):
                    pass

                def diagnostics(self):
                    return {"queue_depth": 0, "active_line": "", "priority_until_at": 0.0}

                def exchange_line(self, line, **_kwargs):
                    commands.append(str(line))
                    if str(line) == "IO SNAPSHOT?":
                        return '{"points":{"GPIO0":0,"I0.0":1,"I0.7":1,"I0.8":1}}'
                    return "ACK"

            with patch("mas004_rpi_databridge.io_runtime.EspPlcClient", FakeEspClient):
                runtime.override_output("esp32_plc58__GPIO0", True, source="test-ui")
                commands.clear()
                result = runtime.refresh(include_points=False, device_codes={"esp32_plc58"})

            self.assertIn("IO SNAPSHOT?", commands)
            self.assertIn("IO SET GPIO0=1", commands)
            self.assertTrue(result["devices"][0]["override_enforced"])
            point = store.get_point("esp32_plc58__GPIO0")
            self.assertEqual("1", point["value"])
            self.assertEqual("override", point["quality"])
            self.assertTrue(point["override_active"])

    def test_gpio0_can_be_overridden_after_led_moved_to_tx1(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    esp_simulation=True,
                    raspi_io_simulation=True,
                    moxa1_simulation=True,
                    moxa2_simulation=True,
                    moxa3_simulation=True,
                ),
                store,
            )

            result = runtime.override_output("esp32_plc58__GPIO0", True, source="test-ui")
            self.assertTrue(result["ok"])
            point = store.get_point("esp32_plc58__GPIO0")
            self.assertEqual("1", point["value"])
            self.assertTrue(point["override_active"])

    def test_unchanged_live_output_is_not_written_again(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            store.upsert_value("moxa_e1213_3__DIO0", "1", "live", "test")
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    moxa3_host="127.0.0.1",
                    moxa3_port=1,
                    moxa3_simulation=False,
                ),
                store,
            )

            result = runtime.write_output("moxa_e1213_3__DIO0", True)

            self.assertTrue(result["ok"])
            self.assertTrue(result["skipped_unchanged"])

    def test_unchanged_io_value_updates_timestamp_only_on_slow_heartbeat(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))

            with patch("mas004_rpi_databridge.io_master.now_ts", return_value=100.0):
                changed = store.upsert_value("raspi_plc21__I0_6", "1", "live", "raspi")
            self.assertTrue(changed)
            first_ts = store.get_point("raspi_plc21__I0_6")["updated_ts"]

            with patch("mas004_rpi_databridge.io_master.now_ts", return_value=105.0):
                changed = store.upsert_value("raspi_plc21__I0_6", "1", "live", "raspi")
            self.assertFalse(changed)
            self.assertEqual(first_ts, store.get_point("raspi_plc21__I0_6")["updated_ts"])

            with patch("mas004_rpi_databridge.io_master.now_ts", return_value=111.0):
                changed = store.upsert_value("raspi_plc21__I0_6", "1", "live", "raspi")
            self.assertFalse(changed)
            self.assertEqual(111.0, store.get_point("raspi_plc21__I0_6")["updated_ts"])

    def test_batch_io_upsert_uses_one_heartbeat_window(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))

            items = [
                ("raspi_plc21__I0_6", "1", "live", "raspi"),
                ("raspi_plc21__I0_7", "0", "live", "raspi"),
            ]
            with patch("mas004_rpi_databridge.io_master.now_ts", return_value=100.0):
                self.assertEqual(2, store.upsert_values(items))
            first_i06_ts = store.get_point("raspi_plc21__I0_6")["updated_ts"]
            first_i07_ts = store.get_point("raspi_plc21__I0_7")["updated_ts"]

            with patch("mas004_rpi_databridge.io_master.now_ts", return_value=105.0):
                self.assertEqual(0, store.upsert_values(items))
            self.assertEqual(first_i06_ts, store.get_point("raspi_plc21__I0_6")["updated_ts"])
            self.assertEqual(first_i07_ts, store.get_point("raspi_plc21__I0_7")["updated_ts"])

            with patch("mas004_rpi_databridge.io_master.now_ts", return_value=106.0):
                self.assertEqual(1, store.upsert_values([("raspi_plc21__I0_7", "1", "live", "raspi")]))
            self.assertEqual(106.0, store.get_point("raspi_plc21__I0_7")["updated_ts"])

    def test_moxa_best_effort_write_failure_does_not_raise(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    moxa3_host="127.0.0.1",
                    moxa3_port=1,
                    moxa3_simulation=False,
                    moxa_timeout_s=0.3,
                ),
                store,
            )

            result = runtime.write_output("moxa_e1213_3__DIO0", True, best_effort=True)

            self.assertFalse(result["ok"])
            self.assertTrue(result["best_effort"])
            self.assertEqual("offline", store.get_point("moxa_e1213_3__DIO0")["quality"])

    def test_moxa_cooldown_refresh_keeps_previous_live_quality(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            store.upsert_value("moxa_e1213_3__DIO0", "1", "live", "test")
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    moxa3_host="127.0.0.1",
                    moxa3_port=1,
                    moxa3_simulation=False,
                    moxa_timeout_s=0.3,
                ),
                store,
            )

            with patch.object(runtime, "_moxa_call", side_effect=_MoxaEndpointCooldown("cooldown")):
                result = runtime.refresh(include_points=False, device_codes={"moxa_e1213_3"})

            self.assertTrue(result["devices"][0]["cooldown"])
            self.assertTrue(result["devices"][0]["debounced"])
            self.assertEqual("live", store.get_point("moxa_e1213_3__DIO0")["quality"])
            self.assertEqual("1", store.get_point("moxa_e1213_3__DIO0")["value"])

    def test_moxa_protocol_error_does_not_cooldown_shared_endpoint(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    moxa3_host="127.0.0.1",
                    moxa3_port=502,
                    moxa3_simulation=False,
                    moxa_timeout_s=0.3,
                ),
                store,
            )

            class FakeMoxa:
                calls = 0
                labels = []

                def read_outputs(self, labels=None):
                    self.labels.append(tuple(labels or ()))
                    self.calls += 1
                    if self.calls == 1:
                        raise MoxaProtocolError("MOXA exception 2")
                    return {"DIO0": 1}

            fake = FakeMoxa()
            with patch.object(runtime, "_moxa_client", return_value=fake):
                first = runtime.refresh(include_points=False, device_codes={"moxa_e1213_3"})
                second = runtime.refresh(include_points=False, device_codes={"moxa_e1213_3"})

            self.assertFalse(first["devices"][0]["reachable"])
            self.assertTrue(second["devices"][0]["reachable"])
            self.assertEqual([("DIO0",), ("DIO0",)], fake.labels)
            self.assertEqual("live", store.get_point("moxa_e1213_3__DIO0")["quality"])
            self.assertEqual("1", store.get_point("moxa_e1213_3__DIO0")["value"])

    def test_full_refresh_handles_esp_before_local_raspi_io(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    esp_simulation=False,
                    raspi_io_simulation=True,
                    moxa1_simulation=True,
                    moxa2_simulation=True,
                    moxa3_simulation=True,
                ),
                store,
            )
            order = []

            def fake_refresh(device_code, device_points):
                order.append(device_code)
                return {"device_code": device_code, "changed": 0}

            with patch.object(runtime, "_refresh_device", side_effect=fake_refresh):
                runtime.refresh(include_points=False)

            self.assertEqual("esp32_plc58", order[0])
            self.assertEqual("raspi_plc21", order[-1])

    def test_esp_io_snapshot_uses_short_broker_wait_timeout(self):
        io_runtime_module._ESP_IO_COOLDOWN_UNTIL = 0.0
        io_runtime_module._ESP_IO_COOLDOWN_ERROR = ""
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    esp_host="127.0.0.1",
                    esp_port=3010,
                    esp_simulation=False,
                    esp_connect_timeout_s=1.5,
                    esp_io_snapshot_timeout_s=0.75,
                    esp_read_timeout_s=2.0,
                ),
                store,
            )
            captured = {}

            class FakeEspClient:
                def __init__(self, host, port, timeout_s):
                    pass

                def diagnostics(self):
                    return {"queue_depth": 0, "active_line": "", "priority_until_at": 0.0}

                def exchange_line(self, line, read_timeout_s=None, **kwargs):
                    captured["line"] = line
                    captured["read_timeout_s"] = read_timeout_s
                    captured.update(kwargs)
                    return '{"points":{"I0.0":1,"I0.7":1,"I0.8":1}}'

            with patch("mas004_rpi_databridge.io_runtime.EspPlcClient", FakeEspClient):
                result = runtime.refresh(include_points=False, device_codes={"esp32_plc58"})

            self.assertTrue(result["devices"][0]["reachable"])
            self.assertEqual("IO SNAPSHOT?", captured["line"])
            self.assertEqual(0.75, captured["read_timeout_s"])
            self.assertLessEqual(captured["wait_timeout_s"], 2.0)
            self.assertGreaterEqual(captured["wait_timeout_s"], 0.5)

    def test_esp_io_snapshot_timeout_closes_broker_socket_for_reconnect(self):
        io_runtime_module._ESP_IO_COOLDOWN_UNTIL = 0.0
        io_runtime_module._ESP_IO_COOLDOWN_ERROR = ""
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "io.xlsx"
            db_path = Path(tmpdir) / "io.sqlite3"
            self.build_workbook(workbook_path)

            store = IoStore(DB(str(db_path)))
            store.import_xlsx(str(workbook_path))
            runtime = IoRuntime(
                Settings(
                    db_path=str(db_path),
                    peer_base_url="",
                    shared_secret="",
                    esp_host="127.0.0.1",
                    esp_port=3010,
                    esp_simulation=False,
                    esp_connect_timeout_s=1.5,
                    esp_io_snapshot_timeout_s=0.75,
                    esp_read_timeout_s=2.0,
                ),
                store,
            )
            closed = []

            class FakeEspClient:
                def __init__(self, host, port, timeout_s):
                    pass

                def diagnostics(self):
                    return {"queue_depth": 0, "active_line": "", "priority_until_at": 0.0}

                def exchange_line(self, *_args, **_kwargs):
                    raise TimeoutError("ESP command broker request timed out")

                def close(self):
                    closed.append(True)

            with patch("mas004_rpi_databridge.io_runtime.EspPlcClient", FakeEspClient):
                result = runtime.refresh(include_points=False, device_codes={"esp32_plc58"})

            self.assertFalse(result["devices"][0]["reachable"])
            self.assertEqual([True], closed)


if __name__ == "__main__":
    unittest.main()
