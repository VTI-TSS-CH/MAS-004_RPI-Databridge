import tempfile
import unittest
from pathlib import Path

import openpyxl

from mas004_rpi_databridge.config import Settings
from mas004_rpi_databridge.db import DB
from mas004_rpi_databridge.motor_master_sync import (
    apply_motor_setup_master_config_to_client,
    reapply_motor_setup_master_to_params,
    sync_motor_master_values,
)
from mas004_rpi_databridge.params import ParamStore, SHEET_NAME


class MotorMasterSyncTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.root = Path(self.tmp.name)
        self.db = DB(str(self.root / "databridge.db"))
        self.params = ParamStore(self.db)
        self.cfg = Settings(
            db_path=str(self.root / "databridge.db"),
            master_params_xlsx_path=str(self.root / "Parameterliste_master.xlsx"),
        )

    def tearDown(self):
        self.tmp.cleanup()

    def _insert_param(self, pkey: str, ptype: str, pid: str, default_v: str, min_v: float, max_v: float):
        with self.db._conn() as c:
            c.execute(
                """INSERT INTO params(
                     pkey,ptype,pid,min_v,max_v,default_v,unit,rw,esp_rw,dtype,
                     name,format_relevant,message,possible_cause,effects,remedy,ai_instructions,updated_ts
                   ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    pkey,
                    ptype,
                    pid,
                    min_v,
                    max_v,
                    default_v,
                    "1/10mm",
                    "R/W",
                    "R/W",
                    "int",
                    pkey,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    1.0,
                ),
            )

    def _write_workbook(self, path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Params_Type.", "Param ID.", "Min.", "Max.", "Default Value"])
        ws.append(["MAP", "0063", 0, 910, "910"])
        ws.append(["MAS", "0016", 0, 910, "910"])
        wb.save(path)

    def test_motor_setup_master_updates_db_snapshot_and_excel_limits(self):
        self._insert_param("MAP0063", "MAP", "0063", "910", 0, 910)
        self._insert_param("MAS0016", "MAS", "0016", "910", 0, 910)
        workbook = self.root / "Parameterliste_master.xlsx"
        self._write_workbook(workbook)

        motor = {
            "config": {
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {
                "link_ok": True,
                "feedback_tenths_mm": 1000,
                "target_tenths_mm": 1000,
                "command_tenths_mm": 1000,
            },
        }

        result = sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            motor,
            {},
            workbook_paths=[str(workbook)],
            allow_protected_position_param_write=True,
        )

        self.assertTrue(result["ok"], result)
        with self.db._conn() as c:
            rows = {
                row[0]: row[1:]
                for row in c.execute(
                    "SELECT pkey, default_v, min_v, max_v FROM params WHERE pkey IN ('MAP0063','MAS0016')"
                ).fetchall()
            }
            values = dict(c.execute("SELECT pkey,value FROM param_values").fetchall())
            snapshot = c.execute("SELECT config_json,state_json FROM motor_setup_master WHERE motor_id=8").fetchone()
        self.assertEqual(("1000", 100.0, 1000.0), rows["MAP0063"])
        self.assertEqual(("1000", 100.0, 1000.0), rows["MAS0016"])
        self.assertEqual("1000", values["MAP0063"])
        self.assertEqual("1000", values["MAS0016"])
        self.assertIsNotNone(snapshot)
        self.assertIn('"max_tenths_mm": 1000', snapshot[0])

        wb = openpyxl.load_workbook(workbook)
        ws = wb[SHEET_NAME]
        self.assertEqual(1000.0, float(ws.cell(2, 4).value))
        self.assertEqual("1000", str(ws.cell(2, 5).value))
        self.assertEqual(1000.0, float(ws.cell(3, 4).value))
        self.assertEqual("1000", str(ws.cell(3, 5).value))

    def test_position_axis_master_sync_without_motor_setup_context_is_blocked(self):
        self._insert_param("MAP0063", "MAP", "0063", "1000", 100, 1000)
        self._insert_param("MAS0016", "MAS", "0016", "1000", 100, 1000)
        workbook = self.root / "Parameterliste_master.xlsx"
        self._write_workbook(workbook)

        motor = {
            "config": {
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 0,
                "max_tenths_mm": 910,
            },
            "state": {
                "link_ok": True,
                "feedback_tenths_mm": 910,
                "target_tenths_mm": 910,
                "command_tenths_mm": 910,
            },
        }

        result = sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            motor,
            {},
            workbook_paths=[str(workbook)],
        )

        self.assertTrue(result["ok"], result)
        self.assertTrue(result["protected_position_param_write_blocked"])
        with self.db._conn() as c:
            rows = {
                row[0]: row[1:]
                for row in c.execute(
                    "SELECT pkey, default_v, min_v, max_v FROM params WHERE pkey IN ('MAP0063','MAS0016')"
                ).fetchall()
            }
            values = dict(c.execute("SELECT pkey,value FROM param_values").fetchall())
            snapshot = c.execute("SELECT config_json,state_json FROM motor_setup_master WHERE motor_id=8").fetchone()
        self.assertEqual(("1000", 100.0, 1000.0), rows["MAP0063"])
        self.assertEqual(("1000", 100.0, 1000.0), rows["MAS0016"])
        self.assertEqual({}, values)
        self.assertIsNone(snapshot)

    def test_plain_config_sync_does_not_overwrite_position_defaults(self):
        self._insert_param("MAP0063", "MAP", "0063", "910", 0, 910)
        self._insert_param("MAS0016", "MAS", "0016", "910", 0, 910)
        workbook = self.root / "Parameterliste_master.xlsx"
        self._write_workbook(workbook)

        motor = {
            "config": {
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {
                "link_ok": True,
                "feedback_tenths_mm": 306,
                "target_tenths_mm": 650,
                "command_tenths_mm": 306,
            },
        }

        result = sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            motor,
            {},
            workbook_paths=[str(workbook)],
            sync_position_default=False,
            allow_protected_position_param_write=True,
        )

        self.assertTrue(result["ok"], result)
        with self.db._conn() as c:
            rows = {
                row[0]: row[1:]
                for row in c.execute(
                    "SELECT pkey, default_v, min_v, max_v FROM params WHERE pkey IN ('MAP0063','MAS0016')"
                ).fetchall()
            }
            values = dict(c.execute("SELECT pkey,value FROM param_values").fetchall())
        self.assertEqual(("910", 100.0, 1000.0), rows["MAP0063"])
        self.assertEqual(("910", 100.0, 1000.0), rows["MAS0016"])
        self.assertEqual({}, values)

        wb = openpyxl.load_workbook(workbook)
        ws = wb[SHEET_NAME]
        self.assertEqual(1000.0, float(ws.cell(2, 4).value))
        self.assertEqual("910", str(ws.cell(2, 5).value))
        self.assertEqual(1000.0, float(ws.cell(3, 4).value))
        self.assertEqual("910", str(ws.cell(3, 5).value))

    def test_position_default_is_not_taken_from_stale_status_after_reboot(self):
        self._insert_param("MAP0060", "MAP", "0060", "0", 0, 1000)
        self._insert_param("MAS0017", "MAS", "0017", "0", 0, 1000)

        motor = {
            "config": {
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 0,
                "max_tenths_mm": 1000,
            },
            "state": {
                "link_ok": False,
                "feedback_tenths_mm": 0,
                "target_tenths_mm": 700,
                "command_tenths_mm": 700,
                "last_error": "Motor live polling disabled until MOTOR POLL=1",
            },
        }

        result = sync_motor_master_values(
            self.params,
            self.cfg,
            5,
            motor,
            {},
            sync_position_default=True,
            allow_protected_position_param_write=True,
        )

        self.assertTrue(result["ok"], result)
        self.assertTrue(result["position_default_skipped"])
        with self.db._conn() as c:
            rows = {
                row[0]: row[1:]
                for row in c.execute(
                    "SELECT pkey, default_v, min_v, max_v FROM params WHERE pkey IN ('MAP0060','MAS0017')"
                ).fetchall()
            }
            values = dict(c.execute("SELECT pkey,value FROM param_values").fetchall())
        self.assertEqual(("0", 0.0, 1000.0), rows["MAP0060"])
        self.assertEqual(("0", 0.0, 1000.0), rows["MAS0017"])
        self.assertEqual({}, values)

    def test_plain_config_sync_preserves_setup_master_position_snapshot(self):
        self._insert_param("MAP0063", "MAP", "0063", "1000", 0, 1000)
        self._insert_param("MAS0016", "MAS", "0016", "1000", 0, 1000)

        initial_motor = {
            "config": {
                "speed_mm_s": 18,
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {
                "link_ok": True,
                "feedback_tenths_mm": 1000,
                "target_tenths_mm": 1000,
                "command_tenths_mm": 1000,
            },
        }
        sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            initial_motor,
            {},
            allow_protected_position_param_write=True,
        )

        config_only_motor = {
            "config": {
                "speed_mm_s": 22,
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {
                "link_ok": True,
                "feedback_tenths_mm": 650,
                "target_tenths_mm": 650,
                "command_tenths_mm": 650,
            },
        }
        sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            config_only_motor,
            {},
            sync_position_default=False,
            allow_protected_position_param_write=True,
        )

        with self.db._conn() as c:
            snapshot = c.execute("SELECT config_json,state_json FROM motor_setup_master WHERE motor_id=8").fetchone()
        self.assertIsNotNone(snapshot)
        self.assertIn('"speed_mm_s": 22', snapshot[0])
        self.assertIn('"feedback_tenths_mm": 1000', snapshot[1])
        self.assertNotIn('"feedback_tenths_mm": 650', snapshot[1])

    def test_reapply_motor_setup_master_does_not_rewrite_position_axis_params(self):
        self._insert_param("MAP0063", "MAP", "0063", "910", 0, 910)
        self._insert_param("MAS0016", "MAS", "0016", "910", 0, 910)
        workbook = self.root / "Parameterliste_master.xlsx"
        self._write_workbook(workbook)

        motor = {
            "config": {
                "speed_mm_s": 18,
                "current_pct": 70,
                "hold_current_pct": 40,
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {
                "link_ok": True,
                "feedback_tenths_mm": 1000,
                "target_tenths_mm": 1000,
                "command_tenths_mm": 1000,
            },
        }
        sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            motor,
            {},
            workbook_paths=[str(workbook)],
            allow_protected_position_param_write=True,
        )

        with self.db._conn() as c:
            c.execute("UPDATE params SET default_v='910', max_v=910 WHERE pkey IN ('MAP0063','MAS0016')")
            c.execute("DELETE FROM param_values")

        result = reapply_motor_setup_master_to_params(
            self.params,
            self.cfg,
            {},
            workbook_paths=[str(workbook)],
        )

        self.assertTrue(result["ok"], result)
        self.assertEqual("motor_setup_only_for_position_axis_params", result["results"][0]["reason"])
        with self.db._conn() as c:
            rows = {
                row[0]: row[1:]
                for row in c.execute(
                    "SELECT pkey, default_v, min_v, max_v FROM params WHERE pkey IN ('MAP0063','MAS0016')"
                ).fetchall()
            }
            values = dict(c.execute("SELECT pkey,value FROM param_values").fetchall())
        self.assertEqual(("910", 100.0, 910.0), rows["MAP0063"])
        self.assertEqual(("910", 100.0, 910.0), rows["MAS0016"])
        self.assertEqual({}, values)

    def test_apply_motor_setup_master_config_to_client_blocks_position_axis_writes_by_default(self):
        self._insert_param("MAP0063", "MAP", "0063", "1000", 100, 1000)
        self._insert_param("MAS0016", "MAS", "0016", "1000", 100, 1000)
        motor = {
            "config": {
                "steps_per_mm": 1250,
                "speed_mm_s": 18,
                "accel_mm_s2": 20,
                "decel_mm_s2": 20,
                "current_pct": 70,
                "hold_current_pct": 40,
                "invert_direction": False,
                "zero_offset_steps": -106368,
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {"link_ok": True, "feedback_tenths_mm": 1000},
        }
        sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            motor,
            {},
            allow_protected_position_param_write=True,
        )

        class FakeClient:
            def __init__(self):
                self.calls = []

            def set_config(self, motor_id, values):
                self.calls.append(("set_config", motor_id, values))
                return {"ok": True, "reply": "ACK_SET"}

            def save(self, motor_id, **kwargs):
                self.calls.append(("save", motor_id, kwargs))
                return {"ok": True, "reply": "ACK_SAVE"}

        client = FakeClient()
        result = apply_motor_setup_master_config_to_client(self.params, client, 8)

        self.assertTrue(result["ok"], result)
        self.assertTrue(result["blocked"])
        self.assertEqual("motor_setup_master_restore_disabled_for_position_axis", result["reason"])
        self.assertEqual([], client.calls)
        self.assertFalse(result["restored"])

    def test_apply_motor_setup_master_config_to_client_stays_blocked_even_with_machine_setup_flag(self):
        self._insert_param("MAP0063", "MAP", "0063", "1000", 100, 1000)
        self._insert_param("MAS0016", "MAS", "0016", "1000", 100, 1000)
        motor = {
            "config": {
                "steps_per_mm": 1250,
                "speed_mm_s": 18,
                "accel_mm_s2": 20,
                "decel_mm_s2": 20,
                "current_pct": 70,
                "hold_current_pct": 40,
                "invert_direction": False,
                "zero_offset_steps": -106368,
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {"link_ok": True, "feedback_tenths_mm": 1000},
        }
        sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            motor,
            {},
            allow_protected_position_param_write=True,
        )

        class FakeClient:
            def __init__(self):
                self.calls = []

            def set_config(self, motor_id, values, **kwargs):
                self.calls.append(("set_config", motor_id, values, kwargs))
                return {"ok": True, "reply": "ACK_SET"}

            def save(self, motor_id, **kwargs):
                self.calls.append(("save", motor_id, kwargs))
                return {"ok": True, "reply": "ACK_SAVE"}

        client = FakeClient()
        result = apply_motor_setup_master_config_to_client(
            self.params,
            client,
            8,
            allow_machine_setup_write=True,
        )

        self.assertTrue(result["ok"], result)
        self.assertTrue(result["blocked"])
        self.assertEqual("motor_setup_master_restore_disabled_for_position_axis", result["reason"])
        self.assertEqual([], client.calls)
        self.assertFalse(result["position_restored"])
        self.assertFalse(result["config_changed"])

    def test_apply_motor_setup_master_keeps_zero_offset_when_position_restore_is_disabled_for_moves(self):
        self._insert_param("MAP0063", "MAP", "0063", "1000", 100, 1000)
        self._insert_param("MAS0016", "MAS", "0016", "1000", 100, 1000)
        motor = {
            "config": {
                "steps_per_mm": 1250,
                "speed_mm_s": 18,
                "accel_mm_s2": 20,
                "decel_mm_s2": 20,
                "current_pct": 70,
                "hold_current_pct": 40,
                "invert_direction": False,
                "zero_offset_steps": -106368,
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {"link_ok": True, "feedback_tenths_mm": 1000},
        }
        sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            motor,
            {},
            allow_protected_position_param_write=True,
        )

        class FakeClient:
            def __init__(self):
                self.calls = []

            def set_config(self, motor_id, values):
                self.calls.append(("set_config", motor_id, values))
                return {"ok": True, "reply": "ACK_SET"}

            def save(self, motor_id):
                self.calls.append(("save", motor_id))
                return {"ok": True, "reply": "ACK_SAVE"}

        client = FakeClient()
        result = apply_motor_setup_master_config_to_client(self.params, client, 8, restore_position=False)

        self.assertTrue(result["ok"], result)
        self.assertTrue(result["blocked"])
        self.assertEqual([], client.calls)
        self.assertFalse(result["position_restored"])

    def test_apply_motor_setup_master_skips_write_when_esp_config_is_current(self):
        motor = {
            "config": {
                "steps_per_mm": 1250,
                "speed_mm_s": 18,
                "accel_mm_s2": 20,
                "decel_mm_s2": 20,
                "current_pct": 70,
                "hold_current_pct": 40,
                "invert_direction": False,
                "zero_offset_steps": -106368,
                "min_enabled": True,
                "max_enabled": True,
                "min_tenths_mm": 100,
                "max_tenths_mm": 1000,
            },
            "state": {"link_ok": True, "feedback_tenths_mm": 635},
        }
        sync_motor_master_values(
            self.params,
            self.cfg,
            8,
            motor,
            {},
            allow_protected_position_param_write=True,
        )

        class FakeClient:
            def __init__(self):
                self.calls = []

            def config(self, motor_id):
                self.calls.append(("config", motor_id))
                return {"ok": True, "config": dict(motor["config"])}

            def set_config(self, motor_id, values):
                self.calls.append(("set_config", motor_id, values))
                return {"ok": True, "reply": "ACK_SET"}

            def save(self, motor_id):
                self.calls.append(("save", motor_id))
                return {"ok": True, "reply": "ACK_SAVE"}

        client = FakeClient()
        result = apply_motor_setup_master_config_to_client(self.params, client, 8, restore_position=False)

        self.assertTrue(result["ok"], result)
        self.assertTrue(result["blocked"])
        self.assertEqual([], client.calls)
        self.assertFalse(result["config_changed"])
        self.assertFalse(result["restored"])

    def test_apply_motor_setup_master_never_restores_zero_offset_or_steps_for_motor3(self):
        motor = {
            "config": {
                "steps_per_mm": 31.627315,
                "speed_mm_s": 100,
                "accel_mm_s2": 300,
                "decel_mm_s2": 300,
                "current_pct": 100,
                "hold_current_pct": 50,
                "invert_direction": True,
                "zero_offset_steps": -7252052,
            },
            "state": {"feedback_tenths_mm": 0},
        }
        sync_motor_master_values(self.params, self.cfg, 3, motor, {})

        class FakeClient:
            def __init__(self):
                self.calls = []

            def set_config(self, motor_id, values):
                self.calls.append(("set_config", motor_id, values))
                return {"ok": True, "reply": "ACK_SET"}

            def save(self, motor_id):
                self.calls.append(("save", motor_id))
                return {"ok": True, "reply": "ACK_SAVE"}

            def set_current_position_mm(self, motor_id, value):
                self.calls.append(("set_current_position_mm", motor_id, value))
                return {"ok": True, "reply": "ACK_SET_POSITION_MM"}

        client = FakeClient()
        result = apply_motor_setup_master_config_to_client(self.params, client, 3)

        self.assertTrue(result["ok"], result)
        self.assertEqual("set_config", client.calls[0][0])
        self.assertEqual(3, client.calls[0][1])
        self.assertNotIn("steps_per_mm", client.calls[0][2])
        self.assertNotIn("zero_offset_steps", client.calls[0][2])
        self.assertNotIn("min_enabled", client.calls[0][2])
        self.assertEqual(("save", 3), client.calls[1])
        self.assertEqual(2, len(client.calls))
        self.assertFalse(result["position_restored"])


if __name__ == "__main__":
    unittest.main()
